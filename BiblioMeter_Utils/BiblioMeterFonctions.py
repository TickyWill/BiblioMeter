__all__ = ['common_files',
           '_get_indexer',
           'concatenate_dat',
           'df_with_no_more_doubles',
           'get_the_doubles',
           'complete_deduplicate_and_save_articles',
           'deduplicated_all_but_articles',
           'you_got_OTPed',
           'liste_de_validation',
           'get_unique_numbers',
           'filtrer_par_departement', 
           'consolidation_anonymat', 
           'ajout_OTP']

def common_files(path_first_corpus,path_second_corpus):
    
    '''The `common_files_new` builds a list of the names of the files present in the parsing folder 
    of two corpuses referenced as first_corpus and second_corpus. 
    
    Args : 
        path_first_corpus (path) : path of the folder where the files of the first_corpus are saved.
        path_second_corpus (path) : path of the folder where the files of the second_corpus are saved.
        
    Returns :
        (list): The list of common files.
        
    '''
        
    # Standard library imports
    import os
    
    # 3rd party library imports
    import pandas as pd
    
    def _list_dat(path_corpus):    
        list_dat =[file for file in os.listdir(path_corpus) if file.endswith('.dat')]            
        return list_dat
    
    list_dir_first_corpus=set(_list_dat(path_first_corpus))
    list_dir_second_corpus=set(_list_dat(path_second_corpus))

    common_list = list_dir_first_corpus.intersection(list_dir_second_corpus)    
    
    return common_list
    
def _get_indexer(path):
    
    '''Returns the length of articles.dat from wos or scopus, to later be used to 
    increment when concatening before getting rid of doublons. If path isn't specified
    then the default path will be PATH_SCOPUS_PARSING which is where parsed.dat documents
    from the scopus corpuses are stored.
    
    Args : 
        path (string) : the path leading to where the articles.dat document is saved     
        
    Returns :
        The number by which we need to increment Pub_id when concatening articles.dat 
        from wos to scopus'''
    
    # Standard library imports
    import os
    from pathlib import Path
    
    # Local library imports   
    
    # 3rd party library imports
    import pandas as pd

    df = pd.read_csv(path, sep="\t") # TO DO
    indexer = len(df)
    
    return indexer

def concatenate_dat(document_name, indexer, path_wos, path_scopus, path_concat):
    
    '''Concatenates the documents.dat having the same name from the wos parsing and the 
    scopus parsing.
    
    Args : 
        document_name (string) : name of the document wanting to be concatenated3
        indexer (int) : number by which we increment the Pub_ids
        path_wos (string) : path leading to where the .dat documents are saved
        path_scopus (string) : path leading to where the .dat documents are saved
        path_3 (string) : path leading to where the new concatenated .dat will be saved
        
    Returns :
        A .dat file saved in path_concat'''
    
    # Standard library imports
    from pathlib import Path
    
    # Local library imports
        
    # 3rd party library imports
    import pandas as pd
    
    df_scopus = pd.read_csv(path_wos / Path(document_name), sep="\t")
    
    df_wos = pd.read_csv(path_scopus / Path(document_name), sep="\t")
    df_wos['Pub_id']=df_wos['Pub_id']+indexer
    
    list_df = [df_wos,df_scopus]
    df_inter = pd.concat(list_df,ignore_index=False)
    df_inter.set_index('Pub_id',inplace=True)
    df_inter.sort_index(inplace=True)
    
    df_inter.to_csv(path_concat / Path(document_name),
                    index=True,
                    columns=df_inter.columns.tolist(),
                    sep='\t',
                    header=True)
    
def df_with_no_more_doubles(path_concat):
    
    '''Uses the concatenated articles.dat document and applies a succesion of filters
    to get rid of doubled information and returns a df containing an index of Pub_ids
    with no doubled lines.
    
    Args :
        path (string) : path leading to where the .dat document is saved
        
    Returns :
        A df with no duplicates but unfull information and a list with the index of this df'''
    
    # Standard library imports
    from pathlib import Path
    
    # Local library imports   
    
    # 3rd party library imports
    import pandas as pd
    
    df_articles_concat = pd.read_csv(path_concat,
                                     sep="\t",
                                     index_col='Pub_id')

    # Get rid of unusable lines (Title and Document Type both empty as well as DOI)
    filtre_Title_DT = (df_articles_concat['Title'].isna()) & (df_articles_concat['Document_type'].isna())
    filtre_Title_DOI = (df_articles_concat['Title'].isna()) & (df_articles_concat['DOI'].isna())
    df_articles_concat=df_articles_concat[~filtre_Title_DOI]
    df_articles_concat=df_articles_concat[~filtre_Title_DT]

    # On récupère un indice unique des duplicats sur le DOI, on traitera les cas particulier après.
    # On les retire donc pour les gérer séparement en les rajouter à la DF une fois traités
    # Pour ce faire on va utiliser DateFrame.drop_duplicates pour récupérer l'index
    # On les rajoutera après

    filtre_DOI_NA = (df_articles_concat['DOI'].isna())
    df_inter_wos = df_articles_concat[~filtre_DOI_NA]
    df_inter_wos = df_inter_wos.drop_duplicates(subset=['DOI'],
                                            keep='first')

    # On rajoute les articles sans DOI
    df_inter_scopus = pd.concat([df_inter_wos,df_articles_concat[filtre_DOI_NA]])

    # Pour gérer les DOI isna(), on va simplement récépérer la DataFrame qui ne possède que les DOI isna()
    # On fera un filtre sur la colonne Title et Document Type (sinon risque de perdre de l'information)

    df_no_doubles = df_inter_scopus.drop_duplicates(subset=['Document_type','Title'], 
                                               keep='first')

    indices_of_duplicates = df_no_doubles.index.tolist()
    
    return [df_no_doubles, indices_of_duplicates]

def get_the_doubles(indice, df_DOI, df_no_DOI, df_articles_concat):
    
    '''_____
    
    Args : 
       _____
        
    Returns :
        A list of the duplicats of indice'''
    
    # Standard library imports
    from pathlib import Path
    
    # Local library imports   
    
    # 3rd party library imports
    import pandas as pd
    
    list_df_dup=[]
    
    filt_inter_DOI = (df_DOI['DOI'] == df_articles_concat['DOI'].loc[indice]) # Renvoie un filtre intermédiaire des mêmes DOI
    df_inter_DOI = df_DOI[filt_inter_DOI] # On récupère une DF avec les doublons, pour permettre de choisir quoi garder ensuite

    filt_inter_Title_DT = (df_no_DOI['Title'] == df_articles_concat['Title'].loc[indice]) & (df_no_DOI['Document_type'] == df_articles_concat['Document_type'].loc[indice])
    df_inter_Title_DT = df_no_DOI[filt_inter_Title_DT]

    if df_inter_DOI.index.tolist() != df_inter_Title_DT.index.tolist():
        df_inter = pd.concat([df_inter_DOI,df_inter_Title_DT])
        
    return [df_inter,df_inter.index.tolist()]

def complete_deduplicate_and_save_articles(list_df_dup, indices_of_duplicates, path):
    
    '''_____
    
    Args : 
       _____
    
    Returns :
        Nothing, but saves documents in path'''
    # Standard library imports
    from pathlib import Path
    
    # Local library imports   
    import BiblioMeter_Utils as bmu
    
    # 3rd party library imports
    import pandas as pd

    df_dup_full_unique=pd.DataFrame()
    tour = 0

    for i in range(len(list_df_dup)):
        working_df = list_df_dup[i]
        nombre_duplication = working_df.shape[0]
        nombre_colonne = working_df.shape[1]
        if nombre_duplication != 1:
            for j in range(nombre_colonne):
                if working_df.iloc[[0],[j]].isna().bool():
                    for k in range(1,nombre_duplication):
                        if working_df.iloc[[k],[j]].isna().bool():
                            working_df.iloc[[0],[j]] = working_df.iloc[[k],[j]]

        if tour == 0:
            tour = 1
            df_dup_full_unique = working_df
        else:
            df_dup_full_unique = pd.concat([df_dup_full_unique,working_df])

    df_dup_full_unique.reset_index(inplace = True)

    df_AAH=pd.DataFrame()        
    df_AAH = df_AAH.append([df_dup_full_unique[df_dup_full_unique['Pub_id'] == i] for i in indices_of_duplicates])

    df_AAH.to_csv(path, 
                  index=False, 
                  columns=df_AAH.columns.tolist(), 
                  sep='\t', 
                  header=True)    

    print('Etape terminée')
    
def deduplicated_all_but_articles(file_name, indices_of_duplicates, path_concat, path_dedupli):
    
    '''_____
    
    Args : 
       _____
        
    Returns :
        Nothing, but saves documents in path'''
    
    # Standard library imports
    from pathlib import Path
    
    # Local library imports   
    
    # 3rd party library imports
    import pandas as pd
    
    exported_df = pd.read_csv(path_concat / Path(file_name), sep="\t")
    
    filt = (exported_df['Pub_id'].isin(indices_of_duplicates))
    exported_df=exported_df[filt]

    if file_name == 'authors.dat':
        exported_df.sort_values(['Pub_id','Idx_author'], inplace=True)
    if file_name == 'addresses.dat':
        exported_df.sort_values(['Pub_id','Idx_address'], inplace=True)
        
    # Terminer les cas particuliers sur le tri des lignes
    
    exported_df.to_csv(path_dedupli / Path(file_name),
                    index=False,
                    columns=exported_df.columns.tolist(),
                    sep='\t',
                    header=True)
    
def you_got_OTPed(df,i):
    
    '''_____
    
    Args : 
       _____
        
    Returns :
        Adds OTP to df'''
    
    # Standard library imports
    from pathlib import Path
    
    # Local library imports   
    import BiblioMeter_Utils as bmu
    
    # 3rd party library imports
    import pandas as pd
    from openpyxl.utils.cell import get_column_letter
    
    # Global variable imports
    from .BiblioMeterGlobalsVariables import OTP_STRING
    
    columns_to_keep = ['Pub_id', 
                   'Idx_author', 
                   'Authors',  
                   'DOI', 
                   'ISSN', 
                   'LITEN_France',
                   'Secondary_institutions', 
                   'Document_type',
                   'Matricule', 
                   'Nom', 
                   'Prénom', 
                   'Dpt/DOB (lib court)', 
                   'Service (lib court)', 
                   'Laboratoire (lib court)', 
                   'Laboratoire (lib long)',
                   'List_of_OTP',
                   'HOMONYM']
    
    if 'List_of_OTP' in df.columns:
        #if df['Pub_id'].iloc[i-1] == df['Pub_id'].iloc[i] and df['Dpt/DOB (lib court)'].iloc[i-1] == df['Dpt/DOB (lib court)'].iloc[i]: VERSION AMAL
        if df['Pub_id'].iloc[i-1] == df['Pub_id'].iloc[i]: # VERSION DE JP
            
            lien_otp = '='+get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(i+1) # ---------------------------> rendre robuste 

            #truc = '"'+','.join(lien_otp)+'"'
            
            df['List_of_OTP'].iloc[i] = lien_otp
            
        else:
            df['List_of_OTP'].iloc[i] = OTP_STRING[df.iloc[i]['Dpt/DOB (lib court)']]
    else:
        df['List_of_OTP'] = pd.Series()
        df['List_of_OTP'].iloc[i] = OTP_STRING[df.iloc[i]['Dpt/DOB (lib court)']]
    
    return df

def liste_de_validation(df,i):
    
    '''_____
    
    Args : 
       _____
        
    Returns :
        Adds OTP to df'''
    
    # Local imports
    import BiblioMeter_Utils as bmu
    
    # Global variable imports
    from .BiblioMeterGlobalsVariables import OTP_LIST
    
    validation_list = OTP_LIST[df.iloc[i]['Dpt/DOB (lib court)']]
    
    truc = '"'+','.join(validation_list)+'"'
    
    return truc

def get_unique_numbers(numbers):

    list_of_unique_numbers = []

    unique_numbers = set(numbers)

    for number in unique_numbers:
        list_of_unique_numbers.append(number)

    return list_of_unique_numbers

def filtrer_par_departement(in_path, out_path):
    
    '''
    Faire attention à ce que path mène à un fichier submit
    
    Args : 
    path : chemin vers fichier submit
    dep_to_keep : liste de 0 et 1 disant quel département inclure
        
    Returns :
    un fichier excel
    '''
    
    # Standard imports
    from pathlib import Path

    # Local imports
    from BiblioMeter_GUI.Globals_GUI import SET_OTP

    # 3rd party imports
    import pandas as pd

    ### Charger les df et ajouter les 4 colonnes ###################################################################################################################
    df_DTNM = pd.read_excel(in_path / Path('fichier_ajout_OTP_DTNM.xlsx'))
    df_DTS = pd.read_excel(in_path / Path('fichier_ajout_OTP_DTS.xlsx'))
    df_DEHT = pd.read_excel(in_path / Path('fichier_ajout_OTP_DEHT.xlsx'))
    df_DTCH = pd.read_excel(in_path / Path('fichier_ajout_OTP_DTCH.xlsx'))
    
    df_OTP = df_DTNM.copy()
    df_OTP = df_OTP.append(df_DTS)
    df_OTP = df_OTP.append(df_DEHT)
    df_OTP = df_OTP.append(df_DTCH)
    
    df_OTP.fillna('', inplace=True)
    df_OTP.set_index('Pub_id', inplace = True)

    #data = [0] * len(df_OTP)
    #df_OTP['DTNM'] = data
    #df_OTP['DTS'] = data
    #df_OTP['DTCH'] = data
    #df_OTP['DEHT'] = data
    ################################################################################################################################################################
#
    #for i in df_OTP.index.unique().to_list():
#
        #if isinstance(df_OTP.loc[i], pd.Series):
            #df_inter_pub_id = df_OTP.loc[i].to_frame().T
        #else:
            #df_inter_pub_id = df_OTP.loc[i]
#
        #for j in df_inter_pub_id['Idx_author']:
#
            #filtre_inter_author = df_inter_pub_id['Idx_author'] == j
            #df_inter_inter = df_inter_pub_id[filtre_inter_author]
#
            #if df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DTNM':
                #df_OTP.loc[i,'DTNM'] = 1
#
            #elif df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DTS':
                #df_OTP.loc[i,'DTS'] = 1
#
            #elif df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DEHT':
                #df_OTP.loc[i,'DEHT'] = 1
#
            #else:
                #df_OTP.loc[i,'DTCH'] = 1
    
    df_OTP.sort_values(['Pub_id','Idx_author'], inplace = True)
    df_OTP.drop(['Idx_author'], axis=1, inplace = True)
    df_OTP.reset_index(inplace = True)
    df_OTP.drop_duplicates(subset = ['Pub_id'], inplace = True)
    df_OTP.set_index('Pub_id', inplace = True)
    df_OTP.to_excel(out_path)
    

def consolidation_anonymat(in_path, out_path):

    # Standard library imports
    from pathlib import Path
    import math

    # 3rd party import
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioMeter_GUI.Globals_GUI import SUBMIT_FILE_NAME
    from BiblioMeter_GUI.Globals_GUI import ORPHAN_FILE_NAME

    ###########################
    # Getting the submit file #
    ###########################

    df_submit = pd.read_excel(in_path)

    # Remettre les Pub_ID dans l'ordre
    df_submit.sort_values(by=['Pub_id', 'Dpt/DOB (lib court)'], ascending=False, inplace = True)
    df_submit.sort_values(by=['Pub_id', 'Idx_author'], ascending=True, inplace = True)

    ############################################
    # Getting rid of the columns we don't want #
    ############################################

    # TO DO : METTRE EN VARIABLE GLOBAL
    columns_to_keep = ['Pub_id', 'Idx_author', 'Title', 'Journal',
                       'Authors',  
                       'DOI', 
                       'ISSN', 
                       'Document_type',
                       'Matricule', 
                       'Nom', 
                       'Prénom', 
                       'Dpt/DOB (lib court)', 
                       'Service (lib court)', 
                       'Laboratoire (lib court)',
                       'HOMONYM']

    df_submit = df_submit[columns_to_keep].copy()

    # La colonne 0/1 pour afficher ou pas
    #list_de_1 = [1] * len(df_submit)
    #df_submit['Affichage'] = list_de_1

    columns_to_keep = df_submit.columns.to_list()

    # Liste unique des Pub_id conservé
    list_of_Pub_id = df_submit['Pub_id'].to_list()
    list_of_Pub_id = bmu.get_unique_numbers(list_of_Pub_id)

    #########################
    # Ouverture du workbook #
    #########################

    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    yellow_ft = PatternFill(fgColor = '00FFFF00', fill_type = "solid")
    red_ft = PatternFill(fgColor = '00FF0000', fill_type = "solid")
    blue_ft = PatternFill(fgColor = '0000FFFF', fill_type = "solid")
    bd = Side(style='medium', color="000000")
    active_color = 'red'

    for indice, r in enumerate(dataframe_to_rows(df_submit, index=False, header=True)):
        ws.append(r)
        last_row = ws[ws.max_row]
        if r[columns_to_keep.index('HOMONYM')] == 'HOMONYM' and indice > 0:
            cell = last_row[columns_to_keep.index('Nom')]
            cell.fill = yellow_ft
            cell = last_row[columns_to_keep.index('Prénom')]
            cell.fill = yellow_ft

            column_letter = get_column_letter(columns_to_keep.index('Nom')+1)
            excel_cell_id = column_letter + str(indice+1)

        if ws.max_row > 1 and r[0] in list_of_Pub_id:
            if list_of_Pub_id.index(r[0])%2 == 0:
                cell = last_row[0]
                cell.fill = red_ft
                if active_color != 'red':
                    for cell_bis in last_row:
                        cell_bis.border = Border(top=bd)
                        active_color = 'red'
            else:
                cell = last_row[0]
                cell.fill = blue_ft
                if active_color != 'blue':
                    for cell_bis in last_row:
                        cell_bis.border = Border(top=bd)
                        active_color = 'blue'


    for cell in ws['A'] + ws[1]:
        cell.font = Font(bold=True)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(out_path)
    

def ajout_OTP(in_path, out_path):

    '''
    '''
    
    # Standard library imports
    from pathlib import Path

    # Local library imports
    import BiblioMeter_Utils as bmu

    # 3rd party import
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter

    df_submit = pd.read_excel(in_path)
    
    df_submit.fillna('', inplace=True)
    df_submit.set_index('Pub_id', inplace = True)
        
    data = [0] * len(df_submit)
    df_submit['DTNM'] = data
    df_submit['DTS'] = data
    df_submit['DTCH'] = data
    df_submit['DEHT'] = data
    ###############################################################################################################################################################

    for i in df_submit.index.unique().to_list():
        
        if isinstance(df_submit.loc[i], pd.Series):
            df_inter_pub_id = df_submit.loc[i].to_frame().T
        else:
            df_inter_pub_id = df_submit.loc[i]

        for j in df_inter_pub_id['Idx_author']:

            filtre_inter_author = df_inter_pub_id['Idx_author'] == j
            df_inter_inter = df_inter_pub_id[filtre_inter_author]

            if df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DTNM':
                df_submit.loc[i,'DTNM'] = 1

            elif df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DTS':
                df_submit.loc[i,'DTS'] = 1

            elif df_inter_inter['Dpt/DOB (lib court)'].to_list()[0] == 'DEHT':
                df_submit.loc[i,'DEHT'] = 1

            else:
                df_submit.loc[i,'DTCH'] = 1
                
    df_submit.sort_values(['Pub_id','Idx_author'], inplace = True)
    df_submit.reset_index(inplace = True)
    df_submit.drop_duplicates(subset = ['Pub_id'], inplace = True)
    
    _=[bmu.you_got_OTPed(df_submit,i) for i in range(len(df_submit))]

    columns_to_keep = df_submit.columns.to_list()

    filtre_DTNM = df_submit['Dpt/DOB (lib court)'] == 'DTNM'
    filtre_DTS = df_submit['Dpt/DOB (lib court)'] == 'DTS'
    filtre_DEHT = df_submit['Dpt/DOB (lib court)'] == 'DEHT'
    filtre_DTCH = (df_submit['Dpt/DOB (lib court)'] == 'DTCH') | (df_submit['Dpt/DOB (lib court)'] == 'DTBH')

    #########################
    # Ouverture du workbook #
    #########################

    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    df_DTNM = df_submit[filtre_DTNM]

    for indice, r in enumerate(dataframe_to_rows(df_DTNM, index=False, header=True)):
        ws.append(r)
    
    df_DTNM.reset_index(inplace = True)
    
    for r in range(0,ws.max_row-2):
            if df_DTNM['Pub_id'].iloc[r-1] == df_DTNM['Pub_id'].iloc[r]: 
                'Yo'
            else:        
                validation_list = bmu.liste_de_validation(df_submit,r)

                data_val = DataValidation(type="list",formula1=validation_list, showErrorMessage=False)
                ws.add_data_validation(data_val)

                data_val.add(ws[get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(r+2)])

    wb.save(out_path / Path(f'fichier_ajout_OTP_DTNM.xlsx'))


    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    df_DTS = df_submit[filtre_DTS]

    for indice, r in enumerate(dataframe_to_rows(df_DTS, index=False, header=True)):
        ws.append(r)
        
    df_DTS.reset_index(inplace = True)

    for r in range(0,ws.max_row-2):
            if df_DTS['Pub_id'].iloc[r-1] == df_DTS['Pub_id'].iloc[r]: 
                'Yo'
            else:        
                validation_list = bmu.liste_de_validation(df_submit,r)

                data_val = DataValidation(type="list",formula1=validation_list, showErrorMessage=False)
                ws.add_data_validation(data_val)

                data_val.add(ws[get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(r+2)])

    wb.save(out_path / Path(f'fichier_ajout_OTP_DTS.xlsx'))


    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    df_DEHT = df_submit[filtre_DEHT]

    for indice, r in enumerate(dataframe_to_rows(df_DEHT, index=False, header=True)):
        ws.append(r)
        
    df_DEHT.reset_index(inplace = True)

    for r in range(0,ws.max_row-2):
            if df_DEHT['Pub_id'].iloc[r-1] == df_DEHT['Pub_id'].iloc[r]: 
                'Yo'
            else:        
                validation_list = bmu.liste_de_validation(df_submit,r)

                data_val = DataValidation(type="list",formula1=validation_list, showErrorMessage=False)
                ws.add_data_validation(data_val)

                data_val.add(ws[get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(r+2)])

    wb.save(out_path / Path(f'fichier_ajout_OTP_DEHT.xlsx'))


    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    df_DTCH = df_submit[filtre_DTCH]

    for indice, r in enumerate(dataframe_to_rows(df_DTCH, index=False, header=True)):
        ws.append(r)
        
    df_DTCH.reset_index(inplace = True)

    for r in range(0,ws.max_row-2):
            if df_DTCH['Pub_id'].iloc[r-1] == df_DTCH['Pub_id'].iloc[r]: 
                'Yo'
            else:        
                validation_list = bmu.liste_de_validation(df_submit,r)

                data_val = DataValidation(type="list",formula1=validation_list, showErrorMessage=False)
                ws.add_data_validation(data_val)

                data_val.add(ws[get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(r+2)])

    wb.save(out_path / Path(f'fichier_ajout_OTP_DTCH.xlsx'))

