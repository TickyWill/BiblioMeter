"""Module of functions for using publications attributes
such as homonyms and OTPs.

"""

__all__ = ['save_homonyms',
           'save_otps',
           'save_shaped_homonyms_file',
           'set_saved_homonyms',
           'set_saved_otps',
          ]


# Standard library imports
from pathlib import Path

# 3rd party imports
import pandas as pd
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.utils.dataframe import dataframe_to_rows \
    as openpyxl_dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation \
    as openpyxl_DataValidation
from openpyxl.utils import get_column_letter \
    as openpyxl_get_column_letter
from openpyxl.styles import PatternFill as openpyxl_PatternFill
from openpyxl.styles import Alignment as openpyxl_Alignment
from openpyxl.styles import Border as openpyxl_Border
from openpyxl.styles import Side as openpyxl_Side
import BiblioParsing as bp

# Local imports
import bmfuncts.employees_globals as eg
import bmfuncts.institute_globals as ig
import bmfuncts.pub_globals as pg
from bmfuncts.rename_cols import build_col_conversion_dic
from bmfuncts.rename_cols import set_col_attr
from bmfuncts.useful_functs import mise_en_page


def save_shaped_homonyms_file(homonyms_df, out_path):
    """Saves as openpyxl workbook the dataframe for resolving 
    homonymies by the user.

    Args:
        homonyms_df (dtaframe): Data for resolving homonymies.
        out_path (path): Full path for saving the created workbook.
    """
    # Setting useful column names
    col_homonyms = list(homonyms_df.columns)

    # Useful aliases of renamed columns names
    name_alias      = col_homonyms[12]
    firstname_alias = col_homonyms[13]
    homonym_alias   = col_homonyms[18]

    wb = openpyxl_Workbook()
    ws = wb.active
    ws.title = 'Consolidation Homonymes'
    yellow_ft = openpyxl_PatternFill(fgColor = pg.ROW_COLORS['highlight'],
                                     fill_type = "solid")

    for indice, r in enumerate(openpyxl_dataframe_to_rows(homonyms_df,
                                                          index = False, header = True)):
        ws.append(r)
        last_row = ws[ws.max_row]
        if r[col_homonyms.index(homonym_alias)] == pg.HOMONYM_FLAG and indice > 0:
            cell      = last_row[col_homonyms.index(name_alias)]
            cell.fill = yellow_ft
            cell      = last_row[col_homonyms.index(firstname_alias)]
            cell.fill = yellow_ft

    wb.save(out_path)


def save_homonyms(institute, org_tup, bibliometer_path, corpus_year):
    """Saves the history of the resolved homonyms by the user.

    First, builds the dataframe to save with the following columns:

        - Hash-ID of the publication for which homonyms have been solved.
        - The personal number of the kept author among the homonyms.

    Finally, saves the dataframe as Excel file.

    Args:
        institute (str): Institute name.
        org_tup (tup): Contains Institute parameters.
        bibliometer_path (path): Full path to working folder.
        corpus_year (str): 4 digits year of the corpus.
    Returns:
        (str): End message.
    """

    #  Setting useful col names
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    submit_col_rename_dic = col_rename_tup[1]
    pub_id_col_alias = submit_col_rename_dic[bp.COL_NAMES["pub_id"]]
    author_idx_col_alias = submit_col_rename_dic[bp.COL_NAMES["authors"][1]]
    homonyms_col_alias = submit_col_rename_dic[pg.COL_NAMES_BONUS['homonym']]
    matricule_col_alias = submit_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['matricule']]

    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    homonyms_folder_alias    = pg.ARCHI_YEAR["homonymes folder"]
    homonyms_file_base_alias = pg.ARCHI_YEAR["homonymes file name base"]
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_homonyms_file_alias = pg.ARCHI_YEAR["kept homonyms file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"]
    homonyms_file_alias      = homonyms_file_base_alias + ' ' + corpus_year + ".xlsx"

    # Setting useful paths
    corpus_year_path        = bibliometer_path / Path(corpus_year)
    bdd_mensuelle_path      = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path       = bdd_mensuelle_path / Path(hash_id_file_alias)
    homonyms_folder_path    = corpus_year_path / Path(homonyms_folder_alias)
    homonyms_file_path      = homonyms_folder_path / Path(homonyms_file_alias)
    history_folder_path     = corpus_year_path / Path(history_folder_alias)
    kept_homonyms_file_path = history_folder_path / Path(kept_homonyms_file_alias)

    # Getting the hash_id dataframe
    hash_id_df  = pd.read_excel(hash_id_file_path)

    # Getting the dataframe of homonyms to solve
    pub_df = pd.read_excel(homonyms_file_path)

    # Building dataframe of pub_id and kept personal numbers for solved homonyms
    temp_df = pub_df[pub_df[homonyms_col_alias]==pg.HOMONYM_FLAG]
    homonyms_df = pd.DataFrame(columns=temp_df.columns)
    for _, pub_id_df in temp_df.groupby(pub_id_col_alias):
        for _, author_df in pub_id_df.groupby(author_idx_col_alias):
            if len(author_df)==1:
                homonyms_df = pd.concat([homonyms_df, author_df])
    kept_matricules_df = homonyms_df[[pub_id_col_alias, matricule_col_alias]]

    # Building hash_id and kept matricules df
    homonyms_history_df = pd.merge(hash_id_df,
                                   kept_matricules_df,
                                   how = 'inner',
                                   on = pub_id_col_alias)
    homonyms_history_df.drop(columns = [pub_id_col_alias], inplace = True)
    homonyms_history_df = homonyms_history_df.astype(str)

    # Concatenating with the dataframe of already saved solved homonyms
    if kept_homonyms_file_path.is_file():
        existing_homonyms_history_df = pd.read_excel(kept_homonyms_file_path)
        homonyms_history_df = pd.concat([existing_homonyms_history_df, homonyms_history_df])
    homonyms_history_df = homonyms_history_df.astype('str')
    homonyms_history_df.drop_duplicates(inplace = True)

    # Saving the concatenated dataframe
    homonyms_history_df.to_excel(kept_homonyms_file_path, index = False)

    message = "History of homonyms resolution saved"
    return message


def set_saved_homonyms(institute, org_tup, bibliometer_path,
                       corpus_year, actual_homonym_status):
    """Resolves the homonyms from the history of the resolved homonyms 
    before submiting the file for resolving remaining homonyms to the user.

    First, builds the dataframe with solved homonyms and homonyms remaining \
    to be solved.

    Finally, saves the dataframe through the `save_shaped_homonyms_file` \
    internal function.

    Args:
        institute (str): Institute name.
        org_tup (tup): Contains Institute parameters.
        bibliometer_path (path): Full path to working folder.
        corpus_year (str): 4 digits year of the corpus.
        actual_homonym_status (bool): True if homonyms exists.
    Returns:
        (tup): Tuple = (End message (str), actualized homonyms \
        status (bool).
    """

    #  Setting useful col names
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    submit_col_rename_dic = col_rename_tup[1]
    pub_id_col_alias = submit_col_rename_dic[bp.COL_NAMES["pub_id"]]
    author_idx_col_alias = submit_col_rename_dic[bp.COL_NAMES["authors"][1]]
    homonyms_col_alias = submit_col_rename_dic[pg.COL_NAMES_BONUS['homonym']]
    matricule_col_alias = submit_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['matricule']]
    hash_id_col_alias   = pg.COL_HASH['hash_id']

    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    homonyms_folder_alias    = pg.ARCHI_YEAR["homonymes folder"]
    homonyms_file_base_alias = pg.ARCHI_YEAR["homonymes file name base"]
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_homonyms_file_alias = pg.ARCHI_YEAR["kept homonyms file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"]
    homonyms_file_alias      = homonyms_file_base_alias + ' ' + corpus_year + ".xlsx"

    # Setting useful paths
    corpus_year_path        = bibliometer_path / Path(corpus_year)
    bdd_mensuelle_path      = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path       = bdd_mensuelle_path / Path(hash_id_file_alias)
    homonyms_folder_path    = corpus_year_path / Path(homonyms_folder_alias)
    homonyms_file_path      = homonyms_folder_path / Path(homonyms_file_alias)
    history_folder_path     = corpus_year_path / Path(history_folder_alias)
    kept_homonyms_file_path = history_folder_path / Path(kept_homonyms_file_alias)

    if kept_homonyms_file_path.is_file():

        # Getting the kept homonyms dataframe
        homonyms_history_df = pd.read_excel(kept_homonyms_file_path)

        # Getting the hash_id dataframe
        hash_id_df = pd.read_excel(hash_id_file_path)

        # Building dataframe of pub_id and personal number to keep related to hash_id
        mats_to_keep_df = pd.merge(hash_id_df,
                                   homonyms_history_df,
                                   how='inner',
                                   on=hash_id_col_alias,)
        mats_to_keep_df = mats_to_keep_df.astype(str)
        mats_to_keep_df.drop(columns=[hash_id_col_alias], inplace=True)

        # Getting the resolved homonyms dataframe to be updated
        homonyms_df = pd.read_excel(homonyms_file_path)
        homonyms_df[matricule_col_alias] = homonyms_df[matricule_col_alias].astype(str)

        # Building the updated homonyms dataframe
        homonyms_df_new = pd.DataFrame(columns=homonyms_df.columns)

        for pub_id, pub_id_homonyms_df in homonyms_df.groupby(pub_id_col_alias):
            for _, author_df in pub_id_homonyms_df.groupby(author_idx_col_alias):
                if len(author_df)==1:
                    # Keeping row of authors without homonyms
                    homonyms_df_new = pd.concat([homonyms_df_new, author_df])
                else:
                    pub_id_mats_to_keep_df = mats_to_keep_df[mats_to_keep_df[pub_id_col_alias]\
                                                             ==pub_id]
                    pub_id_mats_to_keep_list = list(pub_id_mats_to_keep_df[matricule_col_alias])
                    mats_to_check_list = list(author_df[matricule_col_alias])
                    mats_to_keep_list = [x for x in mats_to_check_list\
                                         if x in pub_id_mats_to_keep_list]

                    if mats_to_keep_list:
                        # Keeping only row of matricule to keep when homonymies have been resolved
                        mat_to_keep = mats_to_keep_list[0]
                        new_author_df = author_df[author_df[matricule_col_alias]\
                                                  ==mat_to_keep].copy()
                        new_author_df[homonyms_col_alias] = "_"
                        homonyms_df_new = pd.concat([homonyms_df_new, new_author_df])
                    else:
                        # Keeping all rows when homonymies have not been resolved
                        homonyms_df_new = pd.concat([homonyms_df_new, author_df])

        # Setting actual homonyms status
        actual_homonym_status = False
        if pg.HOMONYM_FLAG in homonyms_df_new[homonyms_col_alias].to_list():
            actual_homonym_status = True
       # Saving updated homonyms_df
        save_shaped_homonyms_file(homonyms_df_new, homonyms_file_path)
        message = "Already resolved homonyms used"
    else:
        message = "No already resolved homonyms available"
    return message, actual_homonym_status


def save_otps(institute, org_tup, bibliometer_path, corpus_year):
    """Saves the history of the attributed OTPs by the user.

    First, builds the dataframe to save with 2 sheets:

    - A sheet which name is given by 'SHEET_SAVE_OTP' global at key 'hash_OTP' \
    with the following columns:

        - Hash-ID of the publication for which OTPs have been attributed. 
        - The OTPs value attributed.

    - A sheet which name is given by 'SHEET_SAVE_OTP' global at key 'doi_OTP' \
    with the following columns:

        - Full name (last name + first name initials) of the first author \
        of the publication for which OTPs have been attributed.
        - DOI of the publication for which OTPs have been attributed.
        - The OTPs value attributed.

    Finally, saves the dataframe as Excel file.

    Args:
        institute (str): Institute name.
        org_tup (tup): Contains Institute parameters.
        bibliometer_path (path): Full path to working folder.
        corpus_year (str): 4 digits year of the corpus.
    Returns:
        (str): End message.
    """

    #  Setting useful col names
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    all_col_rename_dic = col_rename_tup[2]

    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    pub_list_folder_alias    = pg.ARCHI_YEAR["pub list folder"]
    pub_list_file_base_alias = pg.ARCHI_YEAR["pub list file name base"]
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_otps_file_alias     = pg.ARCHI_YEAR["kept OTPs file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"]
    pub_list_file_alias      = pub_list_file_base_alias + f' {corpus_year}.xlsx'

    # Setting useful column name aliases
    pub_id_alias         = all_col_rename_dic[bp.COL_NAMES['pub_id']]
    author_col_alias     = all_col_rename_dic[bp.COL_NAMES['articles'][1]]
    doi_col_alias        = all_col_rename_dic[bp.COL_NAMES['articles'][6]]
    otp_list_col_alias   = all_col_rename_dic[pg.COL_NAMES_BONUS['list OTP']]
    otp_col_alias        = pg.COL_NAMES_BONUS['final OTP']
    hash_otp_sheet_alias = pg.SHEET_SAVE_OTP['hash_OTP']
    doi_otp_sheet_alias  = pg.SHEET_SAVE_OTP['doi_OTP']

    # Setting useful paths
    corpus_year_path     = bibliometer_path / Path(corpus_year)
    bdd_mensuelle_path   = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path    = bdd_mensuelle_path / Path(hash_id_file_alias)
    pub_list_folder_path = corpus_year_path / Path(pub_list_folder_alias)
    pub_list_file_path   = pub_list_folder_path / Path(pub_list_file_alias)
    history_folder_path  = corpus_year_path / Path(history_folder_alias)
    kept_otps_file_path  = history_folder_path / Path(kept_otps_file_alias)

    # Getting the hash_id dataframe
    hash_id_df  = pd.read_excel(hash_id_file_path)

    # Getting the dataframe of consolidated pub list with OTPs
    pub_df = pd.read_excel(pub_list_file_path)

    # Building set OTPs df
    if otp_col_alias in pub_df.columns:
        otp_col = otp_col_alias
    else:
        otp_col = otp_list_col_alias
    otps_df = pub_df[[pub_id_alias, author_col_alias, doi_col_alias, otp_col]].copy()
    otps_df = otps_df.fillna(0)
    otps_df = otps_df.astype(str)
    set_otps_df = otps_df.copy()
    sep = ","
    for idx,row in otps_df.iterrows():
        if sep in row[otp_col] or row[otp_col] == "0":
            set_otps_df = set_otps_df.drop(idx)

    # Building hash_id and kept otp df
    hash_otps_history_df = pd.merge(hash_id_df,
                                    set_otps_df,
                                    how = 'inner',
                                    on = pub_id_alias)
    hash_otps_history_df.drop(columns = [pub_id_alias, author_col_alias, doi_col_alias],
                              inplace = True)
    hash_otps_history_df.rename(columns = {otp_col:otp_col_alias}, inplace = True)

    # Building DOI and kept otp df
    doi_otps_history_df = set_otps_df[[author_col_alias, doi_col_alias, otp_col]].copy()
    doi_otps_history_df.rename(columns = {otp_col:otp_col_alias}, inplace = True)

    # Concatenating with the dataframes of already saved solved OTPs by hash_id and by DOI
    if kept_otps_file_path.is_file():
        existing_otps_history_dict = pd.read_excel(kept_otps_file_path, sheet_name = None)

        existing_hash_otps_history_df = existing_otps_history_dict[hash_otp_sheet_alias]
        if len(existing_hash_otps_history_df)-1:
            hash_otps_history_df = pd.concat([existing_hash_otps_history_df, hash_otps_history_df])
        hash_otps_history_df = hash_otps_history_df.astype('str')
        hash_otps_history_df.drop_duplicates(inplace = True)

        existing_doi_otps_history_df = existing_otps_history_dict[doi_otp_sheet_alias]
        if len(existing_doi_otps_history_df)-1:
            doi_otps_history_df = pd.concat([existing_doi_otps_history_df, doi_otps_history_df])
        doi_otps_history_df = doi_otps_history_df.astype('str')
        doi_otps_history_df.drop_duplicates(inplace = True)

    with pd.ExcelWriter(kept_otps_file_path,  # https://github.com/PyCQA/pylint/issues/3060 pylint: disable=abstract-class-instantiated
                        mode = 'a', if_sheet_exists = 'replace') as writer:
        hash_otps_history_df.to_excel(writer, sheet_name = hash_otp_sheet_alias, index = False)
        doi_otps_history_df.to_excel(writer, sheet_name = doi_otp_sheet_alias, index = False)

    message = "History of kept OTPs saved"
    return message

def _re_save_dpt_otp_file(institute, org_tup, dpt, otp_set_dpt_df, otp_to_set_dpt_df,
                          dpt_otp_list, excel_dpt_path, otp_list_col, columns_list):

    """Rebuilds and saves the openpyxl workbook for the department labelled 'dpt'.

    A data validation list is added to the cells 'otp_cell_alias' only when 
    the OTP in not already attributed.

    The openpyxl workbook is created through the `mise_en_page` function imported from 
    the `bmfuncts.useful_functs` module and is re-configured in the same way as in this 
    function after being modified and before being saved.

    The columns attributes for formatting the workbook are defined through the `set_col_attr` 
    function imported from `bmfuncts.rename_cols` module.

    Args:
        institute (str): Institute name.
        org_tup (tup): Contains Institute parameters.
        dpt (str): Department of the Institute.
        otp_set_dpt_df (dataframe): Data of the already attributed OTPs for the department.
        otp_to_set_dpt_df (dataframe): Data of the OTPs still to be attributed for the department.
        dpt_otp_list (list): String containing the OTPs of the department seperated by semicolons.
        excel_dpt_path (path): Full path to where the workbook is saved. 
        otp_list_col (str): Name of the column that contains the OTPs list.
        columns_list (list): Columns to be formatted.
    """

    # Setting useful column sizes and cell colors
    col_attr, col_set_list = set_col_attr(institute, org_tup)
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'],
                                        fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'],
                                        fill_type = "solid")]

    # Building validation list of OTP for 'dpt' department
    validation_list = '"'+','.join(dpt_otp_list) + '"'
    data_val = openpyxl_DataValidation(type = "list",
                                       formula1 = validation_list,
                                       showErrorMessage = False)

    # Initializing df_dpt_new with the publications which otp is not yet set
    df_dpt_new = otp_to_set_dpt_df.copy()

    # Adding a column containing OTPs of 'dpt' department
    df_dpt_new[otp_list_col] = validation_list

    # Creating and formatting the openpyxl workbook
    wb, ws = mise_en_page(institute, org_tup, df_dpt_new)

    # Setting num of first col and first row in EXCEL files
    excel_first_col_num = 1
    excel_first_row_num = 2

    # Getting the column letter for the OTPs column
    otp_alias_df_index = list(df_dpt_new.columns).index(otp_list_col)
    otp_alias_excel_index = otp_alias_df_index + excel_first_col_num
    otp_alias_column_letter = openpyxl_get_column_letter(otp_alias_excel_index)

    # Activating the validation data list in the OTPs column of df_dpt_new
    df_dpt_len = len(df_dpt_new)
    if df_dpt_len:
        # Adding a validation data list
        ws.add_data_validation(data_val)
        for df_index_row in range(len(df_dpt_new)):
            otp_cell_alias = otp_alias_column_letter + \
                             str(df_index_row + excel_first_row_num)
            data_val.add(ws[otp_cell_alias])

    # Appending rows of the publications which OTP is already set to df_dpt_new
    # and coloring the rows alternatively
    if len(otp_set_dpt_df):
        idx = 1 # continuously incremented index vs row index which is not
        for _, row in otp_set_dpt_df.iterrows():
            ws.append(row.values.flatten().tolist())
            # Coloring the row
            idx_row = df_dpt_len + idx
            idx += 1
            last_row = ws[ws.max_row]
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color

    # Reshaping the alignment and the border of the columns
    for idx_col, col in enumerate(columns_list):
        if col not in col_set_list:
            col_attr[col] = col_attr['else']
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        for cell in ws[column_letter]:
            cell.alignment = openpyxl_Alignment(horizontal = col_attr[col][1],
                                                vertical = "center")
            cell.border = openpyxl_Border(left = openpyxl_Side(border_style = 'thick',
                                                               color = 'FFFFFF'),
                                          right = openpyxl_Side(border_style = 'thick',
                                                                color = 'FFFFFF'))

    # Setting the worksheet label
    ws.title = pg.OTP_SHEET_NAME_BASE + " " +  dpt

    # Saving the workbook
    wb.save(excel_dpt_path)


def set_saved_otps(institute, org_tup, bibliometer_path, corpus_year):
    """Attributes the OTPs from the history of the attributed OTPs 
    before submiting to the user the file for attributing the not yet 
    attributed OTPs.

    Loops on department to:

        1. Build the dataframe with already attributed OTPs \
    and OTPs remaining to be attributed. 
        2. Save the file to be submitted to the user through the \
    `_re_save_dpt_otp_file` internal function.

    Args:
        institute (str): Institute name.
        org_tup (tup): Contains Institute parameters.
        bibliometer_path (path): Full path to working folder.
        corpus_year (str): 4 digits year of the corpus.
    Returns:
        (str): End message giving the status of the OTPs attribution.
    """

    # Setting institute parameters
    dpt_attributs_dict  = org_tup[2]

    # Setting useful col names
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    all_col_rename_dic = col_rename_tup[2]

    # Setting useful folder and file aliases
    bdd_mensuelle_alias  = pg.ARCHI_YEAR["bdd mensuelle"]
    otp_folder_alias     = pg.ARCHI_YEAR["OTP folder"]
    otp_file_base_alias  = pg.ARCHI_YEAR["OTP file name base"]
    history_folder_alias = pg.ARCHI_YEAR["history folder"]
    kept_otps_file_alias = pg.ARCHI_YEAR["kept OTPs file name"]
    hash_id_file_alias   = pg.ARCHI_YEAR["hash_id file name"]

    # Setting useful column names aliases
    hash_id_col_alias    = pg.COL_HASH['hash_id']
    pub_id_alias         = all_col_rename_dic[bp.COL_NAMES['pub_id']]
    author_col_alias     = all_col_rename_dic[bp.COL_NAMES['articles'][1]]
    doi_col_alias        = all_col_rename_dic[bp.COL_NAMES['articles'][6]]
    otp_list_col_alias   = all_col_rename_dic[pg.COL_NAMES_BONUS['list OTP']]
    otp_col_alias        = pg.COL_NAMES_BONUS['final OTP']
    hash_otp_sheet_alias = pg.SHEET_SAVE_OTP['hash_OTP']
    doi_otp_sheet_alias  = pg.SHEET_SAVE_OTP['doi_OTP']

    # Setting useful paths
    corpus_year_path    = bibliometer_path / Path(corpus_year)
    bdd_mensuelle_path  = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path   = bdd_mensuelle_path / Path(hash_id_file_alias)
    history_folder_path = corpus_year_path / Path(history_folder_alias)
    kept_otps_file_path = history_folder_path / Path(kept_otps_file_alias)
    otp_folder_path     = corpus_year_path / Path(otp_folder_alias)

    if kept_otps_file_path.is_file():
        # Getting the hash_id dataframe
        hash_id_df = pd.read_excel(hash_id_file_path)

        # Getting the kept OTPs dataframe by hash_id
        hash_otp_history_df = pd.read_excel(kept_otps_file_path,
                                            sheet_name = hash_otp_sheet_alias)

        # Building df of pub_id and OTPs to set related to hash_id
        pub_id_otp_to_set_df = pd.merge(hash_id_df,
                                        hash_otp_history_df,
                                        how = 'inner',
                                        on  = hash_id_col_alias)

        pub_id_otp_to_set_df = pub_id_otp_to_set_df.astype(str)
        pub_id_otp_to_set_df.drop(columns = [hash_id_col_alias], inplace = True)
        pub_id_to_check_list = [str(row[pub_id_alias]) for _,row
                                in pub_id_otp_to_set_df.iterrows()]
        otp_to_set_list      = [str(row[otp_col_alias]) for _,row
                                in pub_id_otp_to_set_df.iterrows()]

        # Getting the kept OTPs dataframe by DOI and first author
        doi_otp_history_df = pd.read_excel(kept_otps_file_path,
                                           sheet_name = doi_otp_sheet_alias)
        author_to_check_list = doi_otp_history_df[author_col_alias].to_list()
        doi_to_check_list    = doi_otp_history_df[doi_col_alias].to_list()
        doi_otp_to_set_list  = doi_otp_history_df[otp_col_alias].to_list()

        # Setting departments list
        dpt_list = list(dpt_attributs_dict.keys())

        # Setting the already attributed OTPs for each department
        for dpt in sorted(dpt_list):
            # Setting the full path of the EXCEl file for the 'dpt' department
            otp_file_name_dpt = f'{otp_file_base_alias}_{dpt}.xlsx'
            otp_file_name_dpt_path = otp_folder_path / Path(otp_file_name_dpt)

            # Getting the pub list for department dpt
            dpt_pub_df = pd.read_excel(otp_file_name_dpt_path)

            # Setting the pub-id list and the DOIs list for department dpt
            dept_pub_id_list = dpt_pub_df[pub_id_alias].to_list()
            dept_doi_list    = dpt_pub_df[doi_col_alias].to_list()

            # Setting columns list
            col_list = list(dpt_pub_df.columns)

            # Building the 'otp_set_dpt_pub_df' dataframe of publication with OTP set
            otp_set_dpt_pub_df = pd.DataFrame(columns = col_list)

            # Building the 'otp_to_set_dpt_pub_df' dataframe of publication
            # with OTP still to be defined
            otp_to_set_dpt_pub_df = dpt_pub_df.copy()
            otp_to_set_dpt_pub_df.drop(columns = [otp_list_col_alias], inplace = True)

            for otp_idx, pub_id_to_check in enumerate(pub_id_to_check_list):
                otp_to_set = otp_to_set_list[otp_idx]

                if pub_id_to_check in dept_pub_id_list:
                    pub_id_idx = [i for i,e in enumerate(dept_pub_id_list)
                                  if e == pub_id_to_check][0]
                    dpt_pub_df.loc[pub_id_idx, otp_list_col_alias] = otp_to_set
                    dpt_pub_id_to_check_df = dpt_pub_df[dpt_pub_df[pub_id_alias] == pub_id_to_check]
                    otp_set_dpt_pub_df = pd.concat([otp_set_dpt_pub_df, dpt_pub_id_to_check_df])
                    otp_to_set_dpt_pub_df.drop(index = pub_id_idx, inplace = True)

            # Si tous les OTPs non affectés compléter avec DOI_otp
            if len(otp_to_set_dpt_pub_df):

                # Setting the DOIs list of otp_to_set_dpt_pub_df
                otp_to_set_doi_list  = otp_to_set_dpt_pub_df[doi_col_alias].to_list()

                for otp_idx, doi_to_check in enumerate(doi_to_check_list):
                    doi_otp_to_set = doi_otp_to_set_list[otp_idx]

                    if doi_to_check in otp_to_set_doi_list:
                        dpt_doi_idx_list = [i for i,e in enumerate(dept_doi_list)
                                            if e == doi_to_check]

                        if doi_to_check != bp.UNKNOWN:
                            doi_idx = dpt_doi_idx_list[0]
                            dpt_pub_df_to_add = \
                                dpt_pub_df[dpt_pub_df[doi_col_alias] == doi_to_check]
                            dpt_pub_df_to_add.loc[doi_idx, otp_list_col_alias] = doi_otp_to_set
                            otp_set_dpt_pub_df = pd.concat([otp_set_dpt_pub_df,
                                                            dpt_pub_df_to_add])
                            otp_to_set_dpt_pub_df.drop(index = doi_idx, inplace = True)
                        else:
                            # Managing case of unknown DOIs first author name
                            new_otp_to_set_dpt_pub_df = \
                                otp_to_set_dpt_pub_df[otp_to_set_dpt_pub_df[doi_col_alias] \
                                                      == doi_to_check]
                            otp_to_set_auth_list = \
                                new_otp_to_set_dpt_pub_df[author_col_alias].to_list()

                            new_doi_otp_history_df = \
                                doi_otp_history_df[doi_otp_history_df[doi_col_alias] \
                                                   == doi_to_check]
                            author_to_check_list = \
                                new_doi_otp_history_df[author_col_alias].to_list()
                            auth_otp_to_set_list = \
                                new_doi_otp_history_df[otp_col_alias].to_list()

                            for auth_otp_idx, auth_to_check in enumerate(author_to_check_list):
                                auth_otp_to_set = auth_otp_to_set_list[auth_otp_idx]

                                if auth_to_check in otp_to_set_auth_list:
                                    dpt_pub_df_to_add_init = dpt_pub_df[dpt_pub_df[doi_col_alias] \
                                                                        == doi_to_check]
                                    dept_auth_list = dpt_pub_df[author_col_alias].to_list()
                                    dpt_auth_idx_list = [i for i,e in enumerate(dept_auth_list)
                                                         if e == auth_to_check]

                                    for auth_idx in dpt_auth_idx_list:
                                        auth_idx_to_replace_list = []
                                        if auth_idx in otp_to_set_dpt_pub_df.index:
                                            auth_idx_to_replace_list.append(auth_idx)

                                            if auth_idx in otp_set_dpt_pub_df.index:
                                                otp_set_dpt_pub_df.drop(index = auth_idx,
                                                                        inplace = True)

                                        for auth_idx_to_replace in auth_idx_to_replace_list:
                                            dpt_pub_df_to_add_init.loc[auth_idx_to_replace, \
                                                otp_list_col_alias] = auth_otp_to_set

                                            dpt_pub_df_to_add = \
                                                dpt_pub_df_to_add_init\
                                                    [dpt_pub_df_to_add_init[author_col_alias] \
                                                     == auth_to_check]
                                            dpt_pub_df_to_add.loc[auth_idx_to_replace, \
                                                otp_list_col_alias] = auth_otp_to_set
                                            otp_set_dpt_pub_df = pd.concat([otp_set_dpt_pub_df,
                                                                            dpt_pub_df_to_add])
                                            otp_set_dpt_pub_df.drop_duplicates(inplace = True)
                                            otp_to_set_dpt_pub_df.drop(index = auth_idx_to_replace,
                                                                       inplace = True)

            # Setting the list of OTPs for the 'dpt' department
            dpt_otp_list = dpt_attributs_dict[dpt][ig.DPT_OTP_KEY]

            # Resetting validation list for OTPs when not already set and saving the file
            _re_save_dpt_otp_file(institute, org_tup, dpt, otp_set_dpt_pub_df,
                                  otp_to_set_dpt_pub_df, dpt_otp_list,
                                  otp_file_name_dpt_path, otp_list_col_alias,
                                  col_list)

        message = "Already set OTPS used"
    else:
        message = "No already set OTPS available"
    return message
