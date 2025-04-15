"""Module of functions for formatting files as openpyxl workbooks 
used by several modules of package `bmfuncts`.

ToDo: Refactoring of the module (same lines repeated in several 
functions).
"""


__all__ = ['align_cell',
           'build_cell_fill_patterns',
           'build_data_val',
           'color_row',           
           'format_heading',
           'format_page',
           'format_wb_sheet',
           'get_col_letter',
           'save_formatted_df_to_xlsx',
           'set_col_width',
           'set_df_attributes',
          ]

# Standard Library imports
from pathlib import Path

# 3rd party imports
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.styles import Font as openpyxl_Font
from openpyxl.styles import PatternFill as openpyxl_PatternFill
from openpyxl.styles import Alignment as openpyxl_Alignment
from openpyxl.styles import Border as openpyxl_Border
from openpyxl.styles import Side as openpyxl_Side
from openpyxl.utils.dataframe import dataframe_to_rows \
    as openpyxl_dataframe_to_rows
from openpyxl.utils import get_column_letter \
    as openpyxl_get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation \
    as openpyxl_DataValidation

# local imports
import bmfuncts.pub_globals as pg


def get_col_letter(df, col, xl_idx_base):
    """Gets the letter or couple of letters targeting
    a column index of a dataframe taking into account 
    the base of the columns indexes in openpyxl objects.
    
    Args:
        df (dataframe): Data of which column letter is got.
        col (str): Name of column of which letter is got.
        xl_idx_base (int): Base of columns and row indexes \
        in openpyxl objects.
    Returns:
        (str): Letter (or couple of letters) targetting the column
        in openpyxl.
    """
    df_col_index = list(df.columns).index(col)
    xl_col_index = df_col_index + xl_idx_base
    col_letter = openpyxl_get_column_letter(xl_col_index)
    return col_letter


def build_data_val(values_list):
    """Builds a validation list and a list-data-validation rule.

    Args:
        values_list (list): List of values (str) to be used \
        as validation list.
    Returns:
        (tup): (Validation description (str), \
        list-data-validation rule (openpyxl.DataValidation)).
    """
    validation_list = '"'+','.join(values_list) + '"'
    data_val = openpyxl_DataValidation(type = "list",
                                       formula1 = validation_list,
                                       showErrorMessage = False)
    return validation_list, data_val


def build_cell_fill_patterns():
    """Builds list of openpyxl patterns for filling cells using 'ROW_COLORS' global."""
    # Setting cell colors
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'],
                                        fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'],
                                        fill_type = "solid")]
    return cell_colors


def color_row(ws, idx_row, cell_colors):
    """Colors alternately rows in an openpyxl sheet.

    Args:
        ws (openpyxl worksheet): The worksheet where cells are colored.
        idx_row (int): Row index to be colored.
        cell_colors (list): List of openpyxl.PatternFill objects.
    Returns:
        (openpyxl worksheet): The openpyxl worksheet where cells \
        have been colored.
    """
    last_row = ws[ws.max_row]
    if idx_row >= 1:
        cell_color = cell_colors[idx_row%2]
        for cell in last_row:
            cell.fill = cell_color
    return ws


def align_cell(ws, columns_list, col_attr, xl_idx_base):
    """Sets cell alignment and border using dict of column attributes 
    of an openpyxl sheet.

    Args:
        ws (openpyxl worksheet): The worksheet where columns are to be formatted.
        columns_list (list): List of columns names (str) to be formatted.
        col_attr (dict): The columns attributes as dict keyed by column names (str) \
        and valued by the attributes lists of each column composed \
        by [horizontal alignment (str), width (int)].
        xl_idx_base (int): Base of columns and row indexes in openpyxl objects.
    Returns:
        (openpyxl worksheet): The openpyxl worksheet where cells \
        have been aligned.
    """
    borders = openpyxl_Border(left=openpyxl_Side(border_style='thick',
                                                 color='FFFFFF'),
                              right=openpyxl_Side(border_style='thick',
                                                  color='FFFFFF'))
    wrap_text = False
    for col_idx, col in enumerate(columns_list):
        column_letter = openpyxl_get_column_letter(col_idx + xl_idx_base)
        if col_idx==len(columns_list):
            wrap_text = True
        for cell in ws[column_letter]:
            cell.alignment = openpyxl_Alignment(wrap_text=wrap_text,
                                                horizontal=col_attr[col][1],
                                                vertical="center")
            cell.border = borders
    return ws


def format_heading(ws, df_title):
    """Sets the format of the columns heading of an openpyxl sheet.

    Args:
        ws (openpyxl worksheet): The worksheet where columns headings \
        are to be formatted.
        df_title (str): Name of the data type to be formatted.
    Returns:
        (openpyxl worksheet): The openpyxl worksheet where heading cells \
        have been formatted.
    """
    head_font = openpyxl_Font(bold=True)
    head_align = openpyxl_Alignment(wrap_text=True, horizontal="center",
                                    vertical="center")
    pub_alias = pg.DF_TITLES_LIST[0]
    cells_list = ws['A'] + ws[1]
    if df_title!=pub_alias:
        cells_list = ws[1]
    for cell in cells_list:
        cell.font = head_font
        cell.alignment = head_align
    return ws


def set_col_width(ws, columns_list, col_attr, col_idx_init, xl_idx_base):
    """Sets the columns width of an openpyxl sheet.

    Args:
        ws (openpyxl worksheet): The worksheet where columns are to be formatted.
        columns_list (list): List of columns names (str) to be formatted.
        col_attr (dict): The columns attributes as dict keyed by column names (str) \
        and valued by the attributes lists of each column composed \
        by [horizontal alignment (str), width (int)].
        col_idx_init (int): Num of first column to be formatted.
        xl_idx_base (int): Base of columns and row indexes in openpyxl objects.
    Returns:
        (openpyxl worksheet): The cells colored openpyxl worksheet."""

    for col_idx, col in enumerate(columns_list):
        if col_idx>=col_idx_init:
            col_letter = openpyxl_get_column_letter(col_idx + xl_idx_base)
            if col in col_attr.keys():
                ws.column_dimensions[col_letter].width = col_attr[col][0]
            else:
                ws.column_dimensions[col_letter].width = 20
    return ws


def _set_base_attributes(cols_list):
    """Sets the dict for setting the default attributes of columns 
    in terms of width and alignment to be used for formating 
    dataframes before openpyxl save.

    Args:
        cols_list (list): List of columns names (str) for \
        which attributes are to be defined.
    Returns:
        (tup): (The columns attributes as dict keyed by column names (str) \
        and valued by the attributes lists of each column composed \
        by [horizontal alignment (str), width (int)], \
        The rows attributes as dict keyed by "first_row" and "other_rows" \
        and valued by rows height (int), Num of first column to be formatted (int)).
    """
    col_attr_dict = {}
    for col in cols_list:
        col_attr_dict[col] = [15, "center"]
    row_heights_dict = {'first_row': 50,
                        'other_rows': 15}
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_if_issn_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the missing IFs or ISSNs data to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    sub_attr_list = [[15, "center"], [40, "left"]]
    cols_nb = len(cols_list)
    set_attr_nb = len(sub_attr_list)
    add_attr_nb = cols_nb - set_attr_nb
    attr_list = sub_attr_list \
              + [[15, "center"]] * add_attr_nb
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':50,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_invalid_list_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the invalid publications list data to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    sub_attr_list = [[15, "center"], [10, "center"]] \
                  + [[15, "center"]] * 2 \
                  + [[20, "center"]] \
                  + [[40, "left"]] * 4 \
                  + [[20, "center"], [20, "left"],
                     [55, 'left'], [15, "center"]]
    cols_nb = len(cols_list)
    set_attr_nb = len(sub_attr_list)
    dept_nb = cols_nb - set_attr_nb - 1
    attr_list = sub_attr_list \
              + [[10, "center"]] * dept_nb \
              + [[15, "center"]] * 1
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':50,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_pub_list_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the publications list data to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    sub_attr_list = [[15, "center"], [10, "center"]] \
                  + [[15, "center"]] * 2 \
                  + [[20, "center"]] \
                  + [[40, "left"]] * 4 \
                  + [[20, "center"], [20, "left"],
                     [55, 'left'], [15, "center"]]
    cols_nb = len(cols_list)
    set_attr_nb = len(sub_attr_list)
    dept_nb = cols_nb - set_attr_nb - 3
    attr_list = sub_attr_list \
              + [[10, "center"]] * dept_nb \
              + [[15, "center"]] * 3
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':50,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_def_otp_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the data to be saved for setting OTP per 
    publication by the user.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    sub_attr_list = [[15, "center"], [10, "center"]] \
                  + [[15, "center"]] * 2 \
                  + [[20, "center"]] \
                  + [[40, "left"]] * 4 \
                  + [[20, "center"], [20, "left"], [55, 'left']] \
                  + [[15, "center"]] * 3 \
                  + [[25, "center"]] \
                  + [[15, "center"]] * 3
    cols_nb = len(cols_list)
    set_attr_nb = len(sub_attr_list)
    dept_nb = cols_nb - set_attr_nb - 1
    attr_list = sub_attr_list \
              + [[10, "center"]] * dept_nb \
              + [[85, "center"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':50,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_auth_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the authors data to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[12, "center"]] * 3 \
              + [[30, "left"]] * 3 \
              + [[15, "center"]] * 4
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_auth_stat_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the statistics data to be saved for authors 
    scientific production.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[15, "center"], [15, "center"], [30, "left"],
                 [50, "left"],[15, "center"], [95, "left"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_attr_dict(cols_list, widths_list, last_cols_nb=1):
    """ Sets the width and horizontal alignement of each column 
    to be used for formatting the data to be saved.

    The specified widths are at least the following:
    
        - The width of the first column.
        - The width of the other columns than first and last ones.
        - The width of the last columns.

    If the specified-widths number is of only 2, the last column is centered. 
    If the specified width value of the first column is less than 15, 
    the first column is centered.

    Args:
        cols_list (list): The columns names (str) of the data.
        widths_list (list): The list of specified widths (int) \
        of columns.
        last_cols_nb (int): The number of last columns to be considered \
        (optional, default = 1).
    Returns:
        (dict): The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str)).
    """
    # Computing number of other columns than first and last ones
    cols_nb = len(cols_list)
    first_and_last_cols_nb = 1 + last_cols_nb
    other_cols_nb = cols_nb - (first_and_last_cols_nb)

    # Setting first column alignement
    first_col_width = widths_list[0]
    first_col_align = "left"
    if first_col_width<=15:
        first_col_align = "center"

    # Setting last columns alignement
    spec_widths_nb = len(widths_list)
    last_col_align = "left"
    if spec_widths_nb==2:
        last_col_align = "center"

    # Setting width and alignement of each column as dict
    other_cols_width = widths_list[1]
    last_cols_width = widths_list[-1]
    attr_list = [[first_col_width, first_col_align]] \
              + [[other_cols_width, "center"]] * other_cols_nb \
              + [[last_cols_width, last_col_align]] * last_cols_nb
    col_attr_dict = dict(zip(cols_list, attr_list))
    return col_attr_dict


def _set_if_db_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the impact-factors (IFs) data to be saved 
    for the update of the IFs database.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols
    widths_list = [60, 15]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':20,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_kpi_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the key performance indicators (KPIs) data to be saved 
    for the update of the KPIs database.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict using col-widths list
    # with widths-list order: first col, other cols
    widths_list = [35, 15]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_if_ana_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the analysis results to be saved for the 
    impact-factors data.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols
    widths_list = [80, 15]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_kw_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the analysis results to be saved for the 
    keywords data.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols
    widths_list = [50, 15]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_geo_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the analysis results to be saved for the 
    geographical data.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols, last col
    widths_list = [32, 15, 100]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_norm_inst_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the normalized-institutions data to be saved.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols, last col
    widths_list = [12, 15, 100]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_raw_inst_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the raw-institutions data to be saved.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols, last cols
    widths_list = [12, 15, 100]
    col_attr_dict = _set_attr_dict(cols_list, widths_list, last_cols_nb=2)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_distrib_inst_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the distributed-institutions data to be saved.

    The widths and horizontal alignement of each column are 
    set through `_set_attr_dict` internal function.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    # with widths-list order: first col, other cols, last col
    widths_list = [12, 15, 15]
    col_attr_dict = _set_attr_dict(cols_list, widths_list)

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':20}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_inst_country_pub_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the data of publication IDs per country and 
    per institution to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[30, "left"], [25, "center"],
                 [15, "center"], [95, "left"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_pub_country_inst_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the data of institutions per country and per 
    publication ID to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[15, "center"], [25, "center"],
                 [15, "center"], [95, "left"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_country_inst_pub_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the data of publication IDs per institutions 
    and per country to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[30, "left"], [15, "center"],
                 [95, "left"], [15, "center"],
                 [95, "left"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def _set_doctype_stat_attributes(cols_list):
    """Sets the widths and horizontal alignement of each column 
    and the heights of the first row and other rows to be used 
    for formatting the statistics data of doctype-analysis results 
    to be saved.

    Args:
        cols_list (list): The columns names (str) of the data.
    Returns:
        (tup): (The dict keyed by columns names (str) and valued (list) \
        by the width (int) and the horizontal alignement (str), 
        The dict keyed by given row types (str) and valued by row heights (int), 
        The value (int) for initializing columns index).
    """
    # Setting col-attributes dict
    attr_list = [[80, "left"], [25, "center"], [15, "center"],
                 [95, "left"], [15, "center"]]
    col_attr_dict = dict(zip(cols_list, attr_list))

    # Setting row-heights dict
    row_heights_dict = {'first_row':30,
                        'other_rows':15}

    # Setting value to initialize columns index
    col_idx_init = 0
    return col_attr_dict, row_heights_dict, col_idx_init


def set_df_attributes(df_title, df_cols_list):
    """Sets the attributes for formating a given data type as openpyxl object.

    Args:
        df_title (str): Name of the data type to be formatted.
        df_cols_list (list): List of columns names (str) for \
        which attributes are to be defined.
    Returns:
        (tup): (The columns attributes as dict keyed by column names (str) \
        and valued by the attributes lists of each column composed \
        by [horizontal alignment (str), width (int)], \
        The rows attributes as dict keyed by "first_row" and "other_rows" \
        and valued by rows height (int), Num of first column to be formatted (int)).
    """
    # Setting useful aliases
    pub_list_alias = pg.DF_TITLES_LIST[0]
    def_otp_alias = pg.DF_TITLES_LIST[2]
    if_db_alias = pg.DF_TITLES_LIST[3]
    auth_alias = pg.DF_TITLES_LIST[4]
    auth_stat_alias = pg.DF_TITLES_LIST[5]
    kpi_alias = pg.DF_TITLES_LIST[6]
    kw_alias = pg.DF_TITLES_LIST[7]
    geo_alias = pg.DF_TITLES_LIST[8]
    norm_inst_alias = pg.DF_TITLES_LIST[9]
    if_ana_alias = pg.DF_TITLES_LIST[10]
    distrib_inst_alias = pg.DF_TITLES_LIST[11]
    inst_country_pub_alias = pg.DF_TITLES_LIST[12]
    doctype_stat_alias = pg.DF_TITLES_LIST[13]
    pub_country_inst_alias = pg.DF_TITLES_LIST[14]
    country_inst_pub_alias = pg.DF_TITLES_LIST[15]
    raw_inst_alias = pg.DF_TITLES_LIST[16]
    invalids_alias = pg.DF_TITLES_LIST[17]
    missing_if_issn_alias = pg.DF_TITLES_LIST[18]

    if df_title==pub_list_alias:
        attr_tup = _set_pub_list_attributes(df_cols_list)

    elif df_title==invalids_alias:
        attr_tup = _set_invalid_list_attributes(df_cols_list)

    elif df_title==missing_if_issn_alias:
        attr_tup = _set_if_issn_attributes(df_cols_list)

    elif df_title==def_otp_alias:
        attr_tup = _set_def_otp_attributes(df_cols_list)

    elif df_title==if_db_alias:
        attr_tup = _set_if_db_attributes(df_cols_list)

    elif df_title==auth_alias:
        attr_tup = _set_auth_attributes(df_cols_list)

    elif df_title==auth_stat_alias:
        attr_tup = _set_auth_stat_attributes(df_cols_list)

    elif df_title==kpi_alias:
        attr_tup = _set_kpi_attributes(df_cols_list)

    elif df_title==if_ana_alias:
        attr_tup = _set_if_ana_attributes(df_cols_list)

    elif df_title==kw_alias:
        attr_tup = _set_kw_attributes(df_cols_list)

    elif df_title==geo_alias:
        attr_tup = _set_geo_attributes(df_cols_list)

    elif df_title==norm_inst_alias:
        attr_tup = _set_norm_inst_attributes(df_cols_list)

    elif df_title==inst_country_pub_alias:
        attr_tup = _set_inst_country_pub_attributes(df_cols_list)

    elif df_title==pub_country_inst_alias:
        attr_tup = _set_pub_country_inst_attributes(df_cols_list)

    elif df_title==country_inst_pub_alias:
        attr_tup = _set_country_inst_pub_attributes(df_cols_list)

    elif df_title==distrib_inst_alias:
        attr_tup = _set_distrib_inst_attributes(df_cols_list)

    elif df_title==doctype_stat_alias:
        attr_tup = _set_doctype_stat_attributes(df_cols_list)

    elif df_title==raw_inst_alias:
        attr_tup = _set_raw_inst_attributes(df_cols_list)

    else:
        attr_tup = _set_base_attributes(df_cols_list)
    return attr_tup


def format_page(df, df_title, wb=None, header=True,
                cell_colors=None, idx_wrap=None):
    """Formats a worksheet of an openpyxl workbook using 
    columns attributes got through the `set_df_attributes`  
    internal function.

    When the workbook wb is not None, this is applied 
    to the active worksheet of the passed workbook. 
    If the workbook wb is None, then the workbook is created.

    Args:
        df (dataframe): The dataframe to be formatted.
        df_title (str): Name of data to be formatted for setting \
        columns attributes, to be specified using the 'DF_TITLES_LIST' \
        global defined in `bmfuncts.pub_globals` module.
        wb (openpyxl workbook): Worbook of the worksheet \
        to be formatted (default = None).
        header (bool): Value of the 'header' arg of the \
        'openpyxl_dataframe_to_rows' function imported from \
        the openpyxl package (default = True).
        cell_colors (list): List of openpyxl.PatternFill objects \
        (default = None).
        idx_wrap (int): The optional maximum index of the rows \
        for which text is wraped in the last column.
    Returns:
        (tup): (worbook of the formatted worksheet (openpyxl workbook), \
        formatted active sheet).
    """
    # Setting base of columns and row indexes in openpyxl objects
    xl_idx_base = pg.XL_INDEX_BASE

    # Setting list of cell colors
    if not cell_colors:
        cell_colors = build_cell_fill_patterns()

    # Setting useful df attributes
    df_cols_list = df.columns
    attrib_tup = set_df_attributes(df_title, df_cols_list)
    col_attr_dict, row_heights_dict, col_idx_init = attrib_tup

    # Initialize wb as a openpyxl workbook and ws its active worksheet
    if not wb:
        wb = openpyxl_Workbook()
        header = True
    ws = wb.active

    # Coloring alternately rows in ws
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=header)
    for idx_row, row in enumerate(ws_rows):
        ws.append(row)
        ws = color_row(ws, idx_row, cell_colors)

    # Setting cell alignment and border in ws
    ws = align_cell(ws, df_cols_list, col_attr_dict, xl_idx_base)

    # Setting the format of the columns heading
    ws = format_heading(ws, df_title)

    # Setting the columns width
    ws = set_col_width(ws, df_cols_list, col_attr_dict,
                       col_idx_init, xl_idx_base)

    # Setting height of rows
    for idx_row in range(ws.max_row):
        if idx_row==0:
            height = row_heights_dict['first_row']
        elif idx_wrap and idx_row<=idx_wrap: # Auto Height of data row
            height = None
        else:
            height = row_heights_dict['other_rows']
        ws.row_dimensions[idx_row + 1].height = height
    return wb, ws


def format_wb_sheet(sheet_name, df, df_title, wb, first, idx_wrap=None):
    """Formats impact-factors (IFs) sheet in the 'wb' openpyxl workbook 
    as first sheet of the workbook if first is True.

    This done through the `format_page` function imported from 
    the `bmfuncts.useful_functs` module.

    Args:
        sheet_name (str): 4-digits IFs sheet-name.
        df (dataframe): Data to be saved.
        df_title (str): Name of data to be formatted for setting \
        columns attributes, to be specified using the 'DF_TITLES_LIST' \
        global defined in `bmfuncts.pub_globals` module.
        wb (openpyxl workbook): Workbook to be updated with the 'sheet_name' sheet.
        first (bool): True if the sheet to add is the first of the workbook.
        idx_wrap (int): The optional maximum index of the rows \
        for which text is wraped in the last column.
    Returns:
        (openpyxl workbook): The updated workbook with the 'sheet_name' sheet.
    """
    if first:
        wb, ws = format_page(df, df_title, wb=wb, idx_wrap=idx_wrap)
        ws.title = sheet_name
    else:
        wb.create_sheet(sheet_name)
        wb.active = wb[sheet_name]
        wb, ws = format_page(df, df_title, wb=wb, idx_wrap=idx_wrap)
    return wb


def save_formatted_df_to_xlsx(save_path, item_filename, item_df,
                              item_df_title, sheet_name, idx_wrap=None):
    """Formats the 'item_df' dataframe through `format_page` function imported 
    from the `bmfuncts.format_files` module and saves it as xlsx workbook.

    Args:
        save_path (path): Full path to the folder where the data will be saved.
        item_filename (str): Name of the file for saving the data.
        item_df (dataframe): Data to be saved.
        item_df_title (str): Name of data to be formatted for setting \
        columns attributes, to be specified using the 'DF_TITLES_LIST' \
        global defined in `bmfuncts.pub_globals` module.
        sheet_name (str): 4-digits IFs sheet-name. 
        idx_wrap (int): The optional maximum index of the rows \
        for which text is wraped in the last column.
    """
    item_xlsx_file = item_filename
    item_xlsx_path = save_path / Path(item_xlsx_file)
    wb, ws = format_page(item_df, item_df_title, idx_wrap=idx_wrap)
    ws.title = sheet_name
    wb.save(item_xlsx_path)
