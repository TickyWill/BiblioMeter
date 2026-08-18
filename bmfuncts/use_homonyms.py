"""Module of functions for using publications attributes
such as homonyms and OTPs.

"""

__all__ = ['save_homonyms',
           'set_saved_homonyms',
           'solve_homonyms',
          ]


# Standard library imports
from pathlib import Path

# 3rd party imports
import pandas as pd
from openpyxl import Workbook as openpyxl_Workbook
from openpyxl.utils.dataframe import dataframe_to_rows \
    as openpyxl_dataframe_to_rows
from openpyxl.styles import PatternFill as openpyxl_PatternFill

# Local imports
import bmfuncts.pub_globals as bm_pg
from bmfuncts.rename_cols import set_homonym_col_names
from bmfuncts.useful_functs import concat_dfs
from bmfuncts.useful_functs import print_step_text


def _set_use_homonyms_cols(institute, org_tup):
    """Builds a dict setting selected columns names for the process 
    of homonymies resolution.

    This is done through the `set_homonym_col_names` function imported from the 
    `bmfuncts.rename_cols` module.

    Args:
        institute (str): Institute's name.
        org_tup (tup): Contains parameters of Institute's organization.
    Returns:
        (tup): The built dict and the full list of final column names \
        got from the `set_homonym_col_names` function.
    """
    # Setting useful column names from homonyms file
    homonyms_col_dic = set_homonym_col_names(institute, org_tup)

    use_homonyms_cols_dic = {'hash_id_col'  : bm_pg.COL_HASH['hash_id'],
                             'pub_id_col'   : homonyms_col_dic['pub_id'],
                             'author_id_col': homonyms_col_dic['author_id'],
                             'mat_col'      : homonyms_col_dic['matricul'],
                             'lastname_col' : homonyms_col_dic['last_name'],
                             'firstname_col': homonyms_col_dic['first_name'],
                             'homonyms_col' : homonyms_col_dic['homonym'],
                            }

    homonyms_cols_list = list(homonyms_col_dic.values())
    return use_homonyms_cols_dic, homonyms_cols_list


def _save_shaped_homonyms_file(homonyms_df, save_cols_list, out_path):
    """Saves as openpyxl workbook the dataframe for resolving 
    homonymies by the user.

    Args:
        homonyms_df (dataframe): Data for resolving homonymies.
        save_cols_list (list): Names (str) of columns \
        where values are highlighted when homonyms exists.
        out_path (path): Full path for saving the created workbook.
    """
    # Setting useful column names
    homonyms_cols = list(homonyms_df.columns)

    # Setting col names to highlight from save_cols_list arg
    lastname_col, firstname_col, homonyms_col = save_cols_list

    wb = openpyxl_Workbook()
    ws = wb.active
    ws.title = 'Consolidation Homonymes'
    yellow_ft = openpyxl_PatternFill(fgColor=bm_pg.ROW_COLORS['highlight'],
                                     fill_type="solid")

    for idx, row in enumerate(openpyxl_dataframe_to_rows(homonyms_df,
                                                         index=False,
                                                         header=True)):
        ws.append(row)
        last_row = ws[ws.max_row]
        if row[homonyms_cols.index(homonyms_col)]==bm_pg.HOMONYM_FLAG and idx>0:
            cell = last_row[homonyms_cols.index(lastname_col)]
            cell.fill = yellow_ft
            cell = last_row[homonyms_cols.index(firstname_col)]
            cell.fill = yellow_ft

    wb.save(out_path)


def solve_homonyms(solve_homonyms_params, in_path, out_path):
    """Creates the file for homonyms solving by the user.

    First, a dataframe is built from specific columns of the list of publications 
    merged with employees database given by the file pointed by 'in_path' path. 
    In these data the homonyms are tagged by the 'HOMONYM_FLAG' global 
    imported from `bmfuncts.pub_globals` module. 
    Then these data are saved as EXCEL file pointed by 'out_path' path 
    through `_save_shaped_homonyms_file` internal function.

    Args:
        solve_homonyms_params (list): The list composed of the print parameters (list), \
        of the Institute's name (str) and of the org_tup (tup) that contains parameters \
        of Institute's organization used here to set column names for homonyms.
        in_path (path): The full path to the input file of list of publications \
        merged with employees database.
        out_path (path): The full path to the output file for homonyms solving \
        by the user.
    Returns:
        (bool): The homonyms' status, True if homonyms are found.
    """
    # Setting parameters value from 'solve_homonyms_params'
    print_params, institute, org_tup = solve_homonyms_params

    print_step_text("\nBuilding data for homonyms resolution...", print_params)

    # Setting useful col names
    use_homonyms_cols_dic, homonyms_cols_list = _set_use_homonyms_cols(institute, org_tup)
    homonyms_col = use_homonyms_cols_dic['homonyms_col']
    col_keys = ['lastname_col', 'firstname_col', 'homonyms_col']
    save_cols_list = [use_homonyms_cols_dic[key] for key in col_keys]

    # Reading the merge file
    merge_df = pd.read_excel(in_path)

    # Getting rid of the columns we don't want
    homonyms_df = merge_df[homonyms_cols_list].copy()

    # Setting homonyms status
    homonyms_status = False
    if bm_pg.HOMONYM_FLAG in homonyms_df[homonyms_col].to_list():
        homonyms_status = True

    # Saving shaped homonyms_df
    print_step_text("  - Saving homonyms data as shaped openpyxl file...", print_params)
    _save_shaped_homonyms_file(homonyms_df, save_cols_list, out_path)
    step_text = ("  - File for solving homonymies saved"
                 f"\n  - Homonyms existing before using resolution history: {homonyms_status}")
    print_step_text(step_text, print_params)
    return homonyms_status


def _set_homonyms_file_params(wf_path, corpus_year):
    """Builds the files parameters for the process of using resolved homonyms.

    Args:
        wf_path (path): The full path to the working folder.
        corpus_year (str): The 4-digits year of the corpus.
    Returns:
        (tup): (the full path to the file of Hash IDs data, \
        the full path to the file of homonyms resolution, \
        the full path to the file of history of homonyms resolution).
    """
    # Setting useful folder and file aliases
    bdd_mensuelle_alias = bm_pg.ARCHI_YEAR["bdd mensuelle"]
    homonyms_folder_alias = bm_pg.ARCHI_YEAR["homonymes folder"]
    homonyms_file_base_alias = bm_pg.ARCHI_YEAR["homonymes file name base"]
    history_folder_alias = bm_pg.ARCHI_YEAR["history folder"]
    kept_homonyms_file_alias = bm_pg.ARCHI_YEAR["kept homonyms file name"]
    hash_id_file_alias = bm_pg.ARCHI_YEAR["hash_id file name"]
    homonyms_file_alias = homonyms_file_base_alias + ' ' + corpus_year + ".xlsx"

    # Setting useful paths
    corpus_year_path = wf_path / Path(corpus_year)
    bdd_mensuelle_path = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path = bdd_mensuelle_path / Path(hash_id_file_alias)
    homonyms_folder_path = corpus_year_path / Path(homonyms_folder_alias)
    homonyms_file_path = homonyms_folder_path / Path(homonyms_file_alias)
    history_folder_path = corpus_year_path / Path(history_folder_alias)
    kept_homonyms_file_path = history_folder_path / Path(kept_homonyms_file_alias)

    return hash_id_file_path, homonyms_file_path, kept_homonyms_file_path


def save_homonyms(save_homonyms_params):
    """Saves the history of the resolved homonyms by the user.

    First, builds the dataframe to save with the following columns:

        - Hash-ID of the publication for which homonyms have been solved.
        - The personal number of the kept author among the homonyms.

    Finally, saves the dataframe as Excel file.

    Args:
        save_homonyms_params (list): The list composed of the 4 digits \
        year of the corpus (str), of the print parameters (list), \
        of the Institute's name (str), of the org_tup (tup) that contains \
        parameters of Institute's organization and of the full path \
        to working folder (path).
    """
    # Setting params values from save_homonyms_params
    corpus_year, print_params, institute, org_tup, wf_path = save_homonyms_params
    print_step_text("\nUpdating history of solved homonyms...", print_params)

    # Setting useful col names
    use_homonyms_cols_dic, _ = _set_use_homonyms_cols(institute, org_tup)
    col_keys = ['pub_id_col', 'author_id_col', 'mat_col', 'homonyms_col']
    (pub_id_col, author_id_col,
     mat_col, homonyms_col) = [use_homonyms_cols_dic[key] for key in col_keys]

    # Setting useful paths
    (hash_id_file_path, homonyms_file_path,
     kept_homonyms_file_path) = _set_homonyms_file_params(wf_path, corpus_year)

    # Getting the hash_id dataframe
    hash_id_df = pd.read_excel(hash_id_file_path)

    # Getting the dataframe of homonyms to solve
    pub_df = pd.read_excel(homonyms_file_path)

    # Building dataframe of pub_id and kept personal numbers for solved homonyms
    temp_df = pub_df[pub_df[homonyms_col]==bm_pg.HOMONYM_FLAG]
    homonyms_df = pd.DataFrame(columns=temp_df.columns)
    for _, pub_id_df in temp_df.groupby(pub_id_col):
        for _, author_df in pub_id_df.groupby(author_id_col):
            if len(author_df)==1:
                homonyms_df = concat_dfs([homonyms_df, author_df])
    kept_matricules_df = homonyms_df[[pub_id_col, mat_col]]

    # Building hash_id and kept matricules df
    homonyms_history_df = pd.merge(hash_id_df,
                                   kept_matricules_df,
                                   how='inner',
                                   on=pub_id_col)
    homonyms_history_df = homonyms_history_df.drop(columns=[pub_id_col])
    homonyms_history_df = homonyms_history_df.astype(str)

    # Concatenating with the dataframe of already saved solved homonyms
    if kept_homonyms_file_path.is_file():
        existing_homonyms_history_df = pd.read_excel(kept_homonyms_file_path)
        homonyms_history_df = concat_dfs([existing_homonyms_history_df, homonyms_history_df])
    homonyms_history_df = homonyms_history_df.astype('str')
    homonyms_history_df = homonyms_history_df.drop_duplicates()

    # Saving the concatenated dataframe
    homonyms_history_df.to_excel(kept_homonyms_file_path, index=False)
    print_step_text("  - History of homonyms resolution saved", print_params)


def set_saved_homonyms(set_homonyms_params, homonyms_status):
    """Resolves the homonyms from the history of the resolved homonyms 
    before submitting the file for resolving remaining homonyms to the user.

    First, builds the dataframe with solved homonyms and homonyms remaining \
    to be solved. 
    Then, saves the dataframe through the `_save_shaped_homonyms_file` \
    internal function.

    Args:
        save_homonyms_params (list): The list composed of the 4 digits \
        year of the corpus (str), of the print parameters (list), \
        of the Institute's name (str), of the org_tup (tup) that contains \
        parameters of Institute's organization and of the full path \
        to working folder (path).
        homonyms_status (bool): True if homonyms exist.
    Returns:
        (bool): Updated homonyms' status.
    """
    # Setting params values from set_homonyms_params
    corpus_year, print_params, institute, org_tup, wf_path = set_homonyms_params
    print_step_text("\nUsing history of resolved homonyms...", print_params)

    # Setting useful col names
    use_homonyms_cols_dic, _ = _set_use_homonyms_cols(institute, org_tup)
    col_keys = ['hash_id_col', 'pub_id_col', 'author_id_col', 'mat_col', 'homonyms_col']
    (hash_id_col, pub_id_col, author_id_col,
     mat_col, homonyms_col) = [use_homonyms_cols_dic[key] for key in col_keys]
    col_keys = ['lastname_col', 'firstname_col', 'homonyms_col']
    save_cols_list = [use_homonyms_cols_dic[key] for key in col_keys]

    # Setting useful paths
    (hash_id_file_path, homonyms_file_path,
     kept_homonyms_file_path) = _set_homonyms_file_params(wf_path, corpus_year)

    if kept_homonyms_file_path.is_file():

        # Getting the kept homonyms dataframe
        homonyms_history_df = pd.read_excel(kept_homonyms_file_path)

        # Getting the hash_id dataframe
        hash_id_df = pd.read_excel(hash_id_file_path)

        # Building dataframe of pub_id and personal number to keep related to hash_id
        mats_to_keep_df = pd.merge(hash_id_df, homonyms_history_df, how='inner',
                                   on=hash_id_col,)
        mats_to_keep_df = mats_to_keep_df.astype(str)
        mats_to_keep_df = mats_to_keep_df.drop(columns=[hash_id_col])

        # Getting the resolved homonyms dataframe to be updated
        homonyms_df = pd.read_excel(homonyms_file_path)
        homonyms_df[mat_col] = homonyms_df[mat_col].astype(str)

        # Building the updated homonyms dataframe
        homonyms_df_new = pd.DataFrame(columns=homonyms_df.columns)

        for pub_id, pub_id_homonyms_df in homonyms_df.groupby(pub_id_col):
            for _, author_df in pub_id_homonyms_df.groupby(author_id_col):
                if len(author_df)==1:
                    # Keeping row of authors without homonyms
                    homonyms_df_new = concat_dfs([homonyms_df_new, author_df])
                else:
                    pub_id_mats_to_keep_df = mats_to_keep_df[mats_to_keep_df[pub_id_col]\
                                                             ==pub_id]
                    pub_id_mats_to_keep_list = list(pub_id_mats_to_keep_df[mat_col])
                    mats_to_check_list = list(author_df[mat_col])
                    mats_to_keep_list = [x for x in mats_to_check_list\
                                         if x in pub_id_mats_to_keep_list]

                    if mats_to_keep_list:
                        # Keeping only row of matricule to keep when homonymies have been resolved
                        mat_to_keep = mats_to_keep_list[0]
                        new_author_df = author_df[author_df[mat_col]\
                                                  ==mat_to_keep].copy()
                        new_author_df[homonyms_col] = "_"
                        homonyms_df_new = concat_dfs([homonyms_df_new, new_author_df])
                    else:
                        # Keeping all rows when homonymies have not been resolved
                        homonyms_df_new = concat_dfs([homonyms_df_new, author_df])

        # Setting actual homonyms status
        homonyms_status = False
        if bm_pg.HOMONYM_FLAG in homonyms_df_new[homonyms_col].to_list():
            homonyms_status = True
        # Saving updated homonyms_df
        _save_shaped_homonyms_file(homonyms_df_new, save_cols_list, homonyms_file_path)
        step_text = ("  - History of resolved homonyms used"
                     f"\n  - Homonyms remains after using resolution history: {homonyms_status}")
    else:
        step_text = "  - No history of resolved homonyms available"
    print_step_text(step_text, print_params)
    return homonyms_status
