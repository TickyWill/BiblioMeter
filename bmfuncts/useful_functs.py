"""Module of useful functions used by several modules of package `bmfuncts`."""

__all__ = ['check_dedup_parsing_available',
           'create_archi',
           'create_folder',
           'format_df_4_excel',
           'mise_en_page',
           'read_parsing_dict',
           'save_fails_dict',
           'save_parsing_dict',
           'set_rawdata',
          ]


# Standard library imports
import json
import os
import shutil
from pathlib import Path

# 3rd party imports
import BiblioParsing as bp
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows \
    as openpyxl_dataframe_to_rows
from openpyxl.utils import get_column_letter \
    as openpyxl_get_column_letter
from openpyxl.styles import Font as openpyxl_Font
from openpyxl.styles import PatternFill as openpyxl_PatternFill
from openpyxl.styles import Alignment as openpyxl_Alignment
from openpyxl.styles import Border as openpyxl_Border
from openpyxl.styles import Side as openpyxl_Side

# local imports
import bmfuncts.pub_globals as pg
from bmfuncts.rename_cols import set_col_attr
from bmfuncts.config_utils import set_user_config


def check_dedup_parsing_available(bibliometer_path, year):
    """
    """
    # Setting default returned status
    dedup_parsing_status = False

    # Getting the full paths of the working folder architecture for the corpus "year select"
    config_tup = set_user_config(bibliometer_path, year, pg.BDD_LIST)
    parsing_path_dict = config_tup[1]

    # Setting parsing files extension of saved results
    parsing_save_extent = pg.TSV_SAVE_EXTENT

    # Setting path of deduplicated parsings
    dedup_parsing_path = parsing_path_dict['dedup']
    if os.path.isdir(dedup_parsing_path):
        dedup_files_list = []
        for path, _, files in os.walk(dedup_parsing_path):
            dedup_files_list.extend(Path(path) / Path(file) for file in files
                                    if file.endswith(parsing_save_extent))
        if len(dedup_files_list) != 0:
            dedup_parsing_status = True
    return dedup_parsing_status


def _get_database_file_path(database_folder_path, database_file_end):
    """The function lists the files ending with "database_file_end"
    present in the folder targeted by "database_folder_path".
    Then it selects the most recent one.

    Args:
        database_folder_path (path): The path to the folder where files
                                     with names ending with 'database_file_end'
                                     will be searched.
        database_file_end (str): Ending of the names of the files to be searched.
    Returns:
        (path): Path targeting the file found and selected.
    """

    # Listing the available files ending with database_file_end
    list_data_base = []
    for path, _, files in os.walk(database_folder_path):
        list_data_base.extend(Path(path) / Path(file) for file in files
                              if file.endswith(database_file_end))

    if list_data_base:
        # Selecting the most recent file with raw_extent extension
        list_data_base.sort(key = os.path.getmtime, reverse = True)
        database_file_path = list_data_base[0]
    else:
        database_file_path = None
    return database_file_path


def _set_database_extract_info(bibliometer_path, datatype, database):
    """The function builds the path to database extractions and
    the file names ending that are specific to the data type 'datatype'.
    It also sets the folder name of the empty files.
    To do that it uses the global 'ARCHI_EXTRACT' defined
    in 'pub_globals' module.

    Args:
        bibliometer_path (path): The path to the working folder.
        datatype (str): The data type selected for the analysis.
        database (str): The database selected for the analysis.

    Returns:
        (path, str, path): (Path to database extractions,
                            File name ending ,
                            Path to the folder of empty files).
    """

    # Setting useful aliases
    extraction_folder    = pg.ARCHI_EXTRACT["root"]
    empty_file_folder    = pg.ARCHI_EXTRACT["empty-file folder"]
    database_folder      = pg.ARCHI_EXTRACT[database]["root"]
    database_file_base   = pg.ARCHI_EXTRACT[database][datatype]
    database_file_extent = pg.ARCHI_EXTRACT[database]["file_extent"]
    database_file_end    = database_file_base + database_file_extent

    # Setting useful paths
    extraction_folder_path = bibliometer_path / Path(extraction_folder)
    database_folder_path   = extraction_folder_path / Path(database_folder)

    return database_folder_path, database_file_end, empty_file_folder


def set_rawdata(bibliometer_path, datatype, years_list, database):
    """The function sets the rawdata to be used for the data type 'datatype' analysis.
    It copies the files ending with 'database_file_end' from database folder
    targeted by the path 'database_folder_path' to the rawdata folder
    targeted by the path 'rawdata_path'.
    When the database is Scopus and the data type to be analysed is restricted to WoS,
    empty files ending with 'database_file_end' are used as Scopus rawdata.
    """

    # Getting database extractions info
    return_tup = _set_database_extract_info(bibliometer_path, datatype, database)
    database_folder_path, database_file_end, empty_file_folder = return_tup

    last_year_database_file_end = database_file_end
    if datatype == pg.DATATYPE_LIST[1] and database == bp.SCOPUS:
        last_year_datatype = pg.DATATYPE_LIST[0]
        return_tup = _set_database_extract_info(bibliometer_path, last_year_datatype,
                                                database)
        _, last_year_database_file_end, _ = return_tup

    # Cycling on year
    for year in years_list:
        if database == bp.SCOPUS and datatype == pg.DATATYPE_LIST[2]:
            year_database_folder_path = database_folder_path / Path(empty_file_folder)
        else:
            year_database_folder_path = database_folder_path / Path(year)
            if database == bp.SCOPUS and datatype == pg.DATATYPE_LIST[1] and year == years_list[-1]:
                database_file_end = last_year_database_file_end

        year_database_file_path = _get_database_file_path(year_database_folder_path,
                                                          database_file_end)

        rawdata_path_dict, _, _ = set_user_config(bibliometer_path, year, pg.BDD_LIST)
        rawdata_path = rawdata_path_dict[database]
        if os.path.exists(rawdata_path):
            shutil.rmtree(rawdata_path)
        os.makedirs(rawdata_path)
        shutil.copy2(year_database_file_path, rawdata_path)

    message = f"\n{database} rawdata set for {datatype} data type."
    return message


def mise_en_page(institute, org_tup, df,
                 wb = None, if_database = None):
    """When the workbook wb is not None, this is applied
    to the active worksheet of the passed workbook.
    If the workbook wb is None, then the workbook is created.
    """

    # Setting useful column sizes
    col_attr, col_set_list = set_col_attr(institute, org_tup)
    columns_list = list(df.columns)
    for col in columns_list:
        if col not in col_set_list:
            col_attr[col] = col_attr['else']

    # Setting list of cell colors
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'], fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'], fill_type = "solid")]

    # Initialize wb as a workbook and ws its active worksheet
    if not wb:
        wb = Workbook()
    ws = wb.active
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)

    # Coloring alternately rows in ws using list of cell colors cell_colors
    for idx_row, row in enumerate(ws_rows):
        ws.append(row)
        last_row = ws[ws.max_row]
        if idx_row >= 1:
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color

    # Setting cell alignment and border using dict of column attributes col_attr
    if if_database:
        align_list = ["left", "center","center","center"]
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            for cell in ws[column_letter]:
                cell.alignment = openpyxl_Alignment(horizontal=align_list[idx_col],
                                                    vertical="center")
                cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick',
                                                                 color='FFFFFF'),
                                              right=openpyxl_Side(border_style='thick',
                                                                  color='FFFFFF'))
    else:
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            for cell in ws[column_letter]:
                cell.alignment = openpyxl_Alignment(horizontal=col_attr[col][1],
                                                    vertical="center")
                cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick',
                                                                 color='FFFFFF'),
                                              right=openpyxl_Side(border_style='thick',
                                                                  color='FFFFFF'))

    # Setting the format of the columns heading
    cells_list = ws['A'] + ws[1]
    if if_database:
        cells_list = ws[1]
    for cell in cells_list:
        cell.font = openpyxl_Font(bold=True)
        cell.alignment = openpyxl_Alignment(wrap_text=True, horizontal="center",
                                            vertical="center")

    # Setting de columns width using dict of column attributes col_attr if if_database = None
    if if_database:
        col_width_list = [60,15,15,15]
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            ws.column_dimensions[column_letter].width = col_width_list[idx_col]
    else:
        for idx_col, col in enumerate(columns_list):
            if idx_col >= 1:
                column_letter = openpyxl_get_column_letter(idx_col + 1)
                if col in col_attr.keys():
                    ws.column_dimensions[column_letter].width = col_attr[col][0]
                else:
                    ws.column_dimensions[column_letter].width = 20


    # Setting height of first row
    first_row_num = 1
    if if_database:
        ws.row_dimensions[first_row_num].height = 20
    else:
        ws.row_dimensions[first_row_num].height = 50

    return wb, ws


def format_df_4_excel(df, first_col_width, last_col_width = None):
    """
    """

    # Setting list of cell colors
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'], fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'], fill_type = "solid")]

    # Setting useful column attributes
    columns_list = list(df.columns)
    col_attr = {}
    col_attr[columns_list[0]] = [first_col_width, "left"]
    if last_col_width:
        col_attr[columns_list[-1]] = [last_col_width, "left"]
    else:
        col_attr[columns_list[-1]] = [15, "center"]
    for col in columns_list[1:-1]:
        col_attr[col] = [15, "center"]

    # Initializing wb as a workbook and ws its active worksheet
    wb = Workbook()
    ws = wb.active
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)

    # Coloring alternatly rows in ws using list of cell colors cell_colors
    for idx_row, row in enumerate(ws_rows):
        ws.append(row)
        last_row = ws[ws.max_row]
        if idx_row >= 1:
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color

    # Setting cell alignement and border using dict of column attributes col_attr
    for idx_col, col in enumerate(columns_list):
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        for cell in ws[column_letter]:
            cell.alignment = openpyxl_Alignment(horizontal=col_attr[col][1], vertical="center")
            cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick', color='FFFFFF'),
                                          right=openpyxl_Side(border_style='thick', color='FFFFFF'))

    # Setting the format of the columns heading
    cells_list = ws[1]
    for cell in cells_list:
        cell.font = openpyxl_Font(bold=True)
        cell.alignment = openpyxl_Alignment(wrap_text=True, horizontal="center", vertical="center")

    # Setting de columns width using dict of column attributes col_attr
    for idx_col, col in enumerate(columns_list):
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        ws.column_dimensions[column_letter].width = col_attr[col][0]

    # Setting height of rows
    height = 30
    for idx_row in range(ws.max_row):
        row_num = idx_row + 1
        if row_num > 1:
            height = 20
        ws.row_dimensions[row_num].height = height

    return wb, ws


def create_folder(root_path, folder, verbose = False):
    """ """

    folder_path = root_path / Path(folder)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        message = f"{folder_path} created"
    else:
        message = f"{folder_path} already exists"

    if verbose:
        print(message)
    return folder_path


def create_archi(bibliometer_path, corpus_year_folder, verbose = False):
    """The `create_archi` function creates a corpus folder with the required architecture.
    It uses the global "ARCHI_YEAR" for the names of the sub_folders.

    Args:
        bibliometer_path (path): The full path of the working folder.
        corpus_year_folder (str): The name of the folder of the corpus.

    Returns:
        (str): The message giving the folder creation status.

    """
    # Setting useful alias
    archi_alias = pg.ARCHI_YEAR

    corpus_year_folder_path = create_folder(bibliometer_path, corpus_year_folder, verbose = verbose)
    _ = create_folder(corpus_year_folder_path, archi_alias["bdd mensuelle"], verbose = verbose)
    _ = create_folder(corpus_year_folder_path, archi_alias["homonymes folder"], verbose = verbose)
    _ = create_folder(corpus_year_folder_path, archi_alias["OTP folder"], verbose = verbose)
    _ = create_folder(corpus_year_folder_path, archi_alias["pub list folder"], verbose = verbose)
    _ = create_folder(corpus_year_folder_path, archi_alias["history folder"], verbose = verbose)

    analysis_folder = create_folder(corpus_year_folder_path, archi_alias["analyses"],
                                    verbose = verbose)
    _ = create_folder(analysis_folder, archi_alias["if analysis"], verbose = verbose)
    _ = create_folder(analysis_folder, archi_alias["keywords analysis"], verbose = verbose)
    _ = create_folder(analysis_folder, archi_alias["subjects analysis"], verbose = verbose)
    _ = create_folder(analysis_folder, archi_alias["countries analysis"], verbose = verbose)
    _ = create_folder(analysis_folder, archi_alias["institutions analysis"], verbose = verbose)

    corpus_folder = create_folder(corpus_year_folder_path, archi_alias["corpus"], verbose = verbose)

    concat_folder = create_folder(corpus_folder, archi_alias["concat"], verbose = verbose)
    _ = create_folder(concat_folder, archi_alias["parsing"], verbose = verbose)

    dedup_folder = create_folder(corpus_folder, archi_alias["dedup"], verbose = verbose)
    _ = create_folder(dedup_folder, archi_alias["parsing"], verbose = verbose)

    scopus_folder = create_folder(corpus_folder, archi_alias["scopus"], verbose = verbose)
    _ = create_folder(scopus_folder, archi_alias["parsing"], verbose = verbose)
    _ = create_folder(scopus_folder, archi_alias["rawdata"], verbose = verbose)

    wos_folder = create_folder(corpus_folder, archi_alias["wos"], verbose = verbose)
    _ = create_folder(wos_folder, archi_alias["parsing"], verbose = verbose)
    _ = create_folder(wos_folder, archi_alias["rawdata"], verbose = verbose)

    message = f"Architecture created for {corpus_year_folder} folder"
    return message
        
        
def save_parsing_dict(parsing_dict, parsing_path,
                      item_filename_dict, save_extent,
                      dedup_infos=None):
    """
    """
    # Internal functions
    def set_item_path(item, parsing_path):
        item_file = item_filename_dict[item] + "." + save_extent
        item_path = parsing_path / Path(item_file)
        return item_path

    # Setting parameters for the specific case of deduplication results
    if dedup_infos:
        bibliometer_path, datatype, corpus_year = dedup_infos
        
        # Setting aliases for final saving deduplication results
        results_root_alias       = pg.ARCHI_RESULTS["root"]
        results_folder_alias     = pg.ARCHI_RESULTS[datatype]
        results_sub_folder_alias = pg.ARCHI_RESULTS["dedup_parsing"]

        # Setting path for final saving deduplication results
        results_root_path   = bibliometer_path / Path(results_root_alias)
        results_folder_path = results_root_path / Path(results_folder_alias)
        year_target_folder_path    = results_folder_path / Path(corpus_year)
        target_parsing_folder_path = year_target_folder_path / Path(results_sub_folder_alias)
        
        # Checking availability of required final results folders
        if not os.path.exists(year_target_folder_path):
            os.makedirs(year_target_folder_path)
        if not os.path.exists(target_parsing_folder_path):
            os.makedirs(target_parsing_folder_path)

    # Cycling on parsing items
    for item in bp.PARSING_ITEMS_LIST:
        if item in parsing_dict.keys():
            item_df = parsing_dict[item]
            item_working_path = set_item_path(item, parsing_path)
            if save_extent == "xlsx":
                item_df.to_excel(item_working_path, index=False)
                if dedup_infos:
                    item_final_path = set_item_path(item, target_parsing_folder_path)
                    item_df.to_excel(item_final_path, index=False)
            elif save_extent == "dat":
                item_df.to_csv(item_working_path, index=False, sep='\t')
                if dedup_infos:
                    item_final_path = set_item_path(item, target_parsing_folder_path)
                    item_df.to_csv(item_final_path, index=False, sep='\t')
            else:
                item_df.to_csv(item_working_path, index=False, sep=',')
                if dedup_infos:
                    item_final_path = set_item_path(item, target_parsing_folder_path)
                    item_df.to_csv(item_final_path, index=False, sep=',')
        else:
            pass
    
    if dedup_infos:
        end_message = (f"Deduplication files for year {corpus_year} saved in folder: "
                       f"\n  '{target_parsing_folder_path}'")
        print(end_message)

def read_parsing_dict(parsing_path, item_filename_dict, save_extent):
    """
    """

    parsing_dict = {}
    # Cycling on parsing items
    for item in bp.PARSING_ITEMS_LIST:
        item_df = None
        if save_extent == "xlsx":
            item_xlsx_file = item_filename_dict[item] + ".xlsx"
            item_xlsx_path = parsing_path / Path(item_xlsx_file)
            if item_xlsx_path.is_file():
                try:
                    item_df = pd.read_excel(item_xlsx_path)
                except pd.errors.EmptyDataError:
                    item_df = pd.DataFrame()
        elif save_extent == "dat":
            item_tsv_file = item_filename_dict[item] + ".dat"
            item_tsv_path = parsing_path / Path(item_tsv_file)
            if item_tsv_path.is_file():
                try:
                    item_df = pd.read_csv(item_tsv_path, sep = "\t")
                except pd.errors.EmptyDataError:
                    item_df = pd.DataFrame()
        else:
            pass

        if item_df is not None:
            parsing_dict[item] = item_df
    return parsing_dict


def save_fails_dict(fails_dict, parsing_path):
    """The function `save_fails_dict` saves parsing fails in a json file
    named by the global PARSING_PERF.

    Args:
        fails_dict (dict): The dict of parsing fails.
        parsing_path (path): The full path of the parsing results folder
        where the json file is being saved.

    Returns:
        None

    """
    parsing_perf_path = parsing_path / Path(pg.PARSING_PERF)
    with open(parsing_perf_path, 'w', encoding = "utf-8") as write_json:
        json.dump(fails_dict, write_json, indent = 4)
