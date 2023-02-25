__all__ = ['add_author_job_type',
           'add_biblio_list',
           'add_if',
           'add_OTP', 
           'concatenate_pub_lists',
           'consolidate_pub_list',            
           'find_missing_if',          
           'mise_en_page',
           'solving_homonyms',
          ]

def solving_homonyms(in_path, out_path):

    # Standard library imports
    from pathlib import Path
    import math

    # 3rd party imports
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows as openpyxl_dataframe_to_rows
    from openpyxl.styles import PatternFill as openpyxl_PatternFill

    # BiblioAnalysis_Utils package globals imports    
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_USEFUL_COLS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_HOMONYMS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import HOMONYM_FLAG
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import ROW_COLORS

    # Useful alias
    pub_id_alias    = COL_NAMES['pub_id']
    idx_author      = COL_NAMES['authors'][1]
    dpt_alias       = EMPLOYEES_USEFUL_COLS['dpt']
    name_alias      = EMPLOYEES_USEFUL_COLS['name'] 
    firstname_alias = EMPLOYEES_USEFUL_COLS['first_name']
    homonym_alias   = COL_NAMES_BONUS['homonym']
    
    # Reading the submit file #
    df_submit = pd.read_excel(in_path)
    
    # Getting rid of the columns we don't want #
    df_homonyms = df_submit[COL_HOMONYMS].copy()
    
    # Sizing columns widths
    #wb, ws = mise_en_page(df_homonyms)
    
    wb = Workbook()
    ws = wb.active    
    ws.title = 'Consolidation Homonymes'    
    yellow_ft = openpyxl_PatternFill(fgColor = ROW_COLORS['highlight'], fill_type = "solid")

    for indice, r in enumerate(openpyxl_dataframe_to_rows(df_homonyms, index=False, header=True)):
        ws.append(r)
        last_row = ws[ws.max_row]
        if r[COL_HOMONYMS.index(homonym_alias)] == HOMONYM_FLAG and indice > 0:
            cell      = last_row[COL_HOMONYMS.index(name_alias)] 
            cell.fill = yellow_ft
            cell      = last_row[COL_HOMONYMS.index(firstname_alias)] 
            cell.fill = yellow_ft
            
    wb.save(out_path)
    
    end_message = f"File for solving homonymies saved in folder: \n  '{out_path}'"
    return end_message


def add_author_job_type(in_path, out_path):

    ''' The function `add_author_job_type` adds a new column containing the job type for each author
    of each publication listed in an EXCEL file and saves it.
    The job type is get from the employee information available in 3 particular columns wich names 
    are defined by the global 'EMPLOYEES_USEFUL_COLS' at keys 'category', 'status' and 'qualification'.
    The name of the new column is defined by the global "COL_NAMES_BONUS" at key 'author_type'.
    
    Args:
        in_path (path): path (including name of the file) leading to the working EXCEL file. 
        out_path (path): path (including name of the file) leading to where the file with the new column will be saved.
    
    Returns:
        None.
    
    Notes:
        The globals 'EMPLOYEES_CONVERTERS_DIC' and 'EMPLOYEES_USEFUL_COLS' are imported 
        from the module 'BiblioMeterEmployeesGlobals' of the package 'BiblioMeter_FUNCTS'. 
        The global 'COL_NAMES_BONUS'is imported from the module 'BiblioMeterGlobalsVariables' 
        of the package 'BiblioMeter_FUNCTS'.             
    '''
    # 3rd party imports
    import pandas as pd 
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import CATEGORIES_DIC 
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import QUALIFICATION_DIC
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import STATUS_DIC
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_CONVERTERS_DIC
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_USEFUL_COLS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    
    # internal functions:    
    def _get_author_type(row):
        for col_name, dic in author_types_dic.items(): 
            for key,values_list in dic.items():
                values_status = [True for value in values_list if value in row[col_name]]
                if any(values_status): return key
        return '-'

    # Setting useful aliases
    category_col_alias        = EMPLOYEES_USEFUL_COLS['category']
    status_col_alias          = EMPLOYEES_USEFUL_COLS['status']
    qualification_col_alias   = EMPLOYEES_USEFUL_COLS['qualification']     
    author_type_col_alias     = COL_NAMES_BONUS['author_type']
    
    author_types_dic = {category_col_alias      : CATEGORIES_DIC, 
                        status_col_alias        : STATUS_DIC, 
                        qualification_col_alias : QUALIFICATION_DIC}

    # Read of the excel file with dates convertion through EMPLOYEES_CONVERTERS_DIC
    submit_df = pd.read_excel(in_path, converters = EMPLOYEES_CONVERTERS_DIC)
    
    submit_df[author_type_col_alias] = submit_df.apply(_get_author_type, axis=1)
    
    submit_df.to_excel(out_path, index = False)
    
    end_message = f"Column with author job type added in file: \n  '{out_path}'"
    return end_message 


def _add_authors_name_list(in_path, out_path):    
    ''' The function ` _add_authors_name_list` adds two columns to the dataframe get 
    from the Excel file pointed by 'in_path'.
    The columns contain respectively the full name of each author as "NAME, Firstname" 
    and the institute co-authors list with their job type as 
    "NAME1, Firstame1 (job type); NAME2, Firstame2 (job type);...".
    
    Args:
        in_path (path): Fullpath of the working excel file. 
        out_path (path): Fullpath of the processed dataframe as an Excel file 
                         saved after going through its treatment.
    
    Returns:
        (str): end message recalling out_path.
    
    Notes:
        The global 'COL_NAMES' is imported from 'BiblioSpecificGlobals' module 
        of 'BiblioAnalysis_Utils' package.
        The global 'EMPLOYEES_USEFUL_COLS' is imported from 'BiblioMeterEmployeesGlobals' 
        module of 'BiblioMeter_FUNCTS' package.  
        The global 'COL_NAMES_BONUS' is imported from 'BiblioMeterGlobalsVariables' 
        module of 'BiblioMeter_FUNCTS' package.
    '''
    # 3rd party imports
    import pandas as pd
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local imports
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_USEFUL_COLS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    
    # Useful alias
    pub_id_alias          = COL_NAMES['pub_id']
    idx_authors_alias     = COL_NAMES['authors'][1]
    nom_alias             = EMPLOYEES_USEFUL_COLS['name']
    prenom_alias          = EMPLOYEES_USEFUL_COLS['first_name']
    full_name_alias       = COL_NAMES_BONUS['nom prénom']
    author_type_col_alias = COL_NAMES_BONUS['author_type']
    full_name_list_alias  = COL_NAMES_BONUS['nom prénom liste']

    # Read of the excel file
    df_in = pd.read_excel(in_path)
    
    # Add of the column 'Nom Prénom' that will be used to create the authors fullname list
    df_in[prenom_alias] = df_in[prenom_alias].apply(lambda x: x.capitalize())
    df_in[full_name_alias] = df_in[nom_alias] + ', ' + df_in[prenom_alias]
        
    df_out = pd.DataFrame()
    for pub_id, pub_id_df in df_in.groupby(pub_id_alias):
        authors_tup_list = sorted(list(set(zip(pub_id_df[idx_authors_alias], 
                                               pub_id_df[full_name_alias], 
                                               pub_id_df[author_type_col_alias]))))
        authors_str_list = [f'{x[1]} ({x[2]})' for x in  authors_tup_list]
        authors_full_str ="; ".join(authors_str_list)
        pub_id_df[full_name_list_alias] = authors_full_str
        df_out = pd.concat([df_out, pub_id_df])

    # Saving 'df_out' in an excel file 'out_path'
    df_out.to_excel(out_path, index = False)
    
    end_message = f"Column with co-authors list is added to the file: \n  '{out_path}'"
    return end_message


def add_OTP(in_path, out_path, out_file_base):
    '''    
    Args:
        in_path (path): fullpath of the working excel file. 
        out_path (path): fullpath of the saved prosseced file
    
    Returns:
        None.
    
    Notes:
        The global 'COL_NAMES' is imported from the module 'BiblioSpecificGlobals'  
        of the package 'BiblioAnalysis_Utils'.
        The functions `_add_authors_name_list` and `mise_en_page` are imported 
        from the module 'BiblioMeterFonctions' of the package 'BiblioMeter_FUNCTS'.
        The global 'EMPLOYEES_USEFUL_COLS' is imported from the module 'BiblioMeterEmployeesGlobals' 
        of the package 'BiblioMeter_FUNCTS'. 
        The globals 'COL_NAMES_BONUS', 'COL_OTP' and 'DPT_ATTRIBUTS_DICT' are imported 
        from the module 'BiblioMeterGlobalsVariables' of the package 'BiblioMeter_FUNCTS'. 
    '''
    
    # Standard library imports
    from pathlib import Path

    # 3rd party imports
    import pandas as pd
    from openpyxl.worksheet.datavalidation import DataValidation as openpyxl_DataValidation
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local library imports
    from BiblioMeter_FUNCTS.BiblioMeterFonctions import _add_authors_name_list
    from BiblioMeter_FUNCTS.BiblioMeterFonctions import mise_en_page
    
    # Local globals imports 
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_USEFUL_COLS 
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_OTP
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import DPT_ATTRIBUTS_DICT       

    # Internal functions    
    def _save_dpt_OTP_file(dpt, df_dpt, dpt_otp_list, excel_dpt_path):
        
        ''' Create and store an Excel file under 'excel_dpt_path' for the department labelled 'dpt'.
        The OPTs of the choosen department are added in a new column named 'OTP_alias' definined after the
        global `COL_NAMES_BONUS['list OTP']`. 
        A list data validation rules is added to each celles of the column
        'OTP_alias'. The data frame column are renamed after the  global 'COL_OTP'. The Excel frame is
        configurated by the `mise_en_page` function.
        
        '''
        
        # Building validation list of OTP for 'dpt' department
        validation_list = '"'+','.join(dpt_otp_list) + '"' 
        data_val = openpyxl_DataValidation(type = "list", 
                                           formula1 = validation_list, 
                                           showErrorMessage = False)
        
        # Adding a column containing OTPs of 'dpt' department
        df_dpt[OTP_alias] = validation_list
        
        # Renaming the columns 
        df_dpt = df_dpt.reindex(columns = COL_OTP)
        # to replace by :
        # COL_OTP_DICT {col_old_name : col_new_name}
        # df_dpt = df_dpt.rename(columns = COL_OTP_DICT)
        
        # Formatting the EXCEL file
        wb, ws = mise_en_page(df_dpt)
        ws.title = 'OTP ' +  dpt    # To define via a global
        
        # Setting num of first col and first row in EXCEL files
        excel_first_col_num = 1
        excel_first_row_num = 2
        
        # Getting the column letter for the OTPs column 
        OTP_alias_df_index = list(df_dpt.columns).index(OTP_alias)        
        OTP_alias_excel_index = OTP_alias_df_index + excel_first_col_num 
        OTP_alias_column_letter = openpyxl_get_column_letter(OTP_alias_excel_index)
                
        # Activating the validation data list in all cells of the OTPs column
        if len(df_dpt):
            # Adding a validation data list
            ws.add_data_validation(data_val)
            for df_index_row in range(len(df_dpt)):  
                OTP_cell_alias = OTP_alias_column_letter + str(df_index_row + excel_first_row_num)
                data_val.add(ws[OTP_cell_alias])

        wb.save(excel_dpt_path)
        
    # Setting useful aliases
    pub_id_alias = COL_NAMES['pub_id']           # Pub_id
    idx_author   = COL_NAMES['authors'][1]       # Idx_author
    dpt_alias    = EMPLOYEES_USEFUL_COLS['dpt']  # Dpt/DOB (lib court)
    OTP_alias    = COL_NAMES_BONUS['list OTP']   # Choix de l'OTP
    
    # Adding a column with a list of the authors in the file where homonymies 
    # have been solved and pointed by in_path
    end_message = _add_authors_name_list(in_path, in_path)
    print('\n ',end_message)
    
    solved_homonymies_df = pd.read_excel(in_path)
    solved_homonymies_df.fillna('', inplace = True)
    
    dpt_list = list(DPT_ATTRIBUTS_DICT.keys())
     
    # For each department adding a column containing 1 or 0 
    # depending if the author belongs or not to the department
    for dpt in  dpt_list:
        dpt_label_list = DPT_ATTRIBUTS_DICT[dpt]['dpt_label']
        solved_homonymies_df[dpt] = solved_homonymies_df[dpt_alias].apply(lambda x: 1 
                                                                          if x in dpt_label_list 
                                                                          else 0)
    
    # Building 'df_out' out of 'solved_homonymies_df' with a row per pub_id 
    # 1 or 0 is assigned to each department column depending 
    # on if at least one co-author is a member of this department,
    # the detailed information is related to the first author only
    df_out = pd.DataFrame()
    for pub_id, dg in solved_homonymies_df.groupby('Pub_id'):
        dg = dg.sort_values(by=['Idx_author'])
        x = dg[dpt_list].any().astype(int) #sum()
        dg[dpt_list] = x
        df_out = pd.concat([df_out,dg.iloc[:1]]) 
    
    # Configuring an Excel file per department with the list of OTPs
    for dpt in sorted(dpt_list):
        # Setting df_dpt with only pub_ids for which the first author
        # is from the 'dpt' department
        filtre_dpt = False
        for dpt_value in DPT_ATTRIBUTS_DICT[dpt]['dpt_label']:
            filtre_dpt = filtre_dpt | (df_out[dpt_alias] == dpt_value)
        df_dpt = df_out[filtre_dpt].copy()
        
        # Setting the list of OTPs for the 'dpt' department
        dpt_otp_list = DPT_ATTRIBUTS_DICT[dpt]['dpt_otp']
               
        # Setting the full path of the EXCEl file for the 'dpt' department
        OTP_file_name_dpt = f'{out_file_base}_{dpt}.xlsx'
        excel_dpt_path = out_path / Path(OTP_file_name_dpt)
        
        # Adding a column with validation list for OTPs and saving the file
        _save_dpt_OTP_file(dpt,df_dpt,dpt_otp_list,excel_dpt_path)

    end_message  = f"Files for setting publication OTPs per department "
    end_message += f"saved in folder: \n  '{out_path}'"
    return end_message
    

def consolidate_pub_list(in_path, out_path, in_file_base):
    
    '''    
    Args : 
        in_path
        out_path
        in_file_base
        
    Returns :
        un fichier excel
    '''
    
    # Standard imports
    from pathlib import Path
    import os
    
    # 3rd party imports
    import pandas as pd    

    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    # local globals imports
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_FINAL_LIST    
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_FINALE   
    from BiblioMeter_GUI.Globals_GUI import DPT_LABEL_DICT
    
    # internal functions
    def _set_df_OTP_dpt(dpt_label):        
        OTP_file_name_dpt_ok = in_file_base + '_' + dpt_label + '_ok.xlsx'
        dpt_path = in_path / Path(OTP_file_name_dpt_ok)       
        if not os.path.exists(dpt_path):
            OTP_file_name_dpt = in_file_base + '_' + dpt_label + '.xlsx'
            dpt_path = in_path / Path(OTP_file_name_dpt)
        dpt_df = pd.read_excel(dpt_path)
        return dpt_df
    
    # Useful alias
    pub_id_alias = COL_NAMES['pub_id']
    
    ### Charger les df et ajouter les 4 colonnes 
    dpt_label_list = list(DPT_LABEL_DICT.keys())
    OTP_df_init_status = True
    for dpt_label in dpt_label_list:
        dpt_df =  _set_df_OTP_dpt(dpt_label)
        if OTP_df_init_status:
            OTP_df = dpt_df.copy()            
        else:
            OTP_df = OTP_df.append(dpt_df)
        OTP_df_init_status = False

    # Deduplicating rows on Pub_id
    OTP_df.drop_duplicates(subset = [pub_id_alias], inplace = True)
    
    # Selecting useful columns using COL_FINAL_LIST
    consolidate_pub_list_df = OTP_df[COL_FINAL_LIST].copy()          
    
    # Renaming column names using COL_NAMES_FINAL
    consolidate_pub_list_df.rename(columns = COL_NAMES_FINALE, inplace = True)
    
    # Saving df to EXCEL file
    consolidate_pub_list_df.to_excel(out_path, index = False)
    
    end_message = f"OTPs identification integrated in file : \n  '{out_path}'"
    return end_message
    

def add_biblio_list(in_path, out_path):
    ''' The function `add_biblio_list` adds a new column containing the full reference 
    of each publication listed in an EXCEL file and saves it.
    The full reference is built by concatenating the folowing items: title, first author, year, journal, DOI.
    These items sould be available in the initial EXCEL file with columns names 
    defined by the global 'COL_NAMES' with the keys 'pub_id' and 'articles'.
    The name of the new column is defined by the global "COL_NAMES_BONUS['liste biblio']".
    
    Args:
        in_path (path): path (including name of the file) leading to the working EXCEL file. 
        out_path (path): path (including name of the file) leading to where the file with the new column will be saved.
    
    Returns:
        None.
    
    Notes:
        The global 'COL_NAMES' is imported from the module 'BiblioSpecificGlobals'  
        of the package 'BiblioAnalysis_Utils'.
        The global 'EMPLOYEES_CONVERTERS_DIC' is imported from the module 'BiblioMeterEmployeesGlobals' 
        of the package 'BiblioMeter_FUNCTS'. 
        The global 'COL_NAMES_BONUS'is imported from the module 'BiblioMeterGlobalsVariables' 
        of the package 'BiblioMeter_FUNCTS'.             
    '''
    # 3rd party imports
    import pandas as pd 
    
    # Local imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BiblioMeterEmployeesGlobals import EMPLOYEES_CONVERTERS_DIC
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS

    # Setting useful aliases
    pub_id_alias           = COL_NAMES['pub_id']
    pub_first_author_alias = COL_NAMES['articles'][1]
    pub_year_alias         = COL_NAMES['articles'][2]
    pub_journal_alias      = COL_NAMES['articles'][3]
    pub_doi_alias          = COL_NAMES['articles'][6]
    pub_title_alias        = COL_NAMES['articles'][9]   
    pub_full_ref_alias     = COL_NAMES_BONUS['liste biblio']

    # Read of the excel file with dates convertion through EMPLOYEES_CONVERTERS_DIC
    submit_df = pd.read_excel(in_path, converters = EMPLOYEES_CONVERTERS_DIC)

    articles_plus_full_ref_df = pd.DataFrame()
    for pub_id, pub_id_df in submit_df.groupby(pub_id_alias): # Split the frame into subframes with same Pub_id
        
        pub_id_first_row = pub_id_df.iloc[0]                                # Select the first row and build the full reference
        full_ref  = f'{pub_id_first_row[pub_title_alias]}, '                # add the reference's title
        full_ref += f'{pub_id_first_row[pub_first_author_alias]}. et al., ' # add the reference's first author
        full_ref += f'{pub_id_first_row[pub_journal_alias].capitalize()}, ' # add the reference's journal name
        full_ref += f'{str(pub_id_first_row[pub_year_alias])}, '            # add the reference's publication year
        full_ref += f'{pub_id_first_row[pub_doi_alias]}'                    # add the reference's DOI
        
        pub_id_df[pub_full_ref_alias] = full_ref
        articles_plus_full_ref_df = pd.concat([articles_plus_full_ref_df, pub_id_df])

    articles_plus_full_ref_df.to_excel(out_path, index = False)
    
    end_message = f"Column with full reference of publication added in file: \n  '{out_path}'"
    return end_message
    
    
def add_if(in_file_path, out_file_path, if_path, year = None):
    
    '''The function `add_if` adds two new columns containing impact factors
    to the corpus dataframe 'corpus_df' got from a file which full path is 'in_file_path'. 
    The two columns are named through 'corpus_year_if_col_alias' and 'current_year_if_col_alias'.
    The impact factors are got from an Excel workbook with a worksheet per year 
    which full path is 'if_path' and put in the IFs dataframe 'if_df'.
    The column 'corpus_year_if_col_alias' is filled with the IFs values 
    of the corpus year 'year' if available in the dataframe 'if_df', 
    else the values are set to 'not_available_if_alias'.
    The column 'current_year_if_col_alias' is filled with the IFs values 
    of the most recent year available in the dataframe 'if_df'.
    In these columns, the nan values of IFs are replaced by 'unknown_if_fill_alias'.
    If 'year' is None, this is performed for all the years found in the corpus dataframe 
    and results in the update of the two columns.
    If 'year' is not None, this is performed for the corpus dataframe of the year 'year' 
    and results in the creation of the two columns.
    
    Args:
        in_file_path (path): The full path to get the corpus dataframe 'corpus_df'. 
        out_file_path (path): The full path to save the new corpus dataframe with the two columns.
        if_path (path): The full path to get the dataframe 'if_df'.
        year (int): The year of the corpus to be appended with the two new IF columns.
    Returns:
        (str): Message indicating which file has been affected and how. 
    
    Notes:
        The global 'COL_NAMES' is imported from the module 'BiblioSpecificGlobals' 
        of the package 'BiblioAnalysis_Utils'.
        The globals 'COL_MAJ_IF', 'COL_NAMES_BONUS', 'FILL_EMPTY_KEY_WORD' 
        and 'NOT_AVAILABLE_IF' are imported from the module 'BiblioMeterGlobalsVariables' 
        of the package 'BiblioMeter_FUNCTS'.    
    '''
    
    # 3rd party imports
    import pandas as pd    
    
    # BiblioAnalysis_Utils imports 
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local imports
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_MAJ_IF
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_FINALE
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import FILL_EMPTY_KEY_WORD
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import NOT_AVAILABLE_IF
    
    # Internal functions    
    def _create_if_column(issn_column, if_dict, if_empty_word):
        ''' The function `_create_if_column` builds a dataframe column 'if_column'
        using the column 'issn_column' of this dataframe and the dict 'if_dict' 
        that make the link between ISSNs ('if_dict' keys) and IFs ('if_dict' values). 
        The 'nan' values in the column 'if_column' are replaced by 'empty_word'
        
        Args:
            issn_column (pandas serie): The column of the dataframe of interest 
                                        that contains the ISSNs values.
            if_dict (dict): The dict which keys are ISSNs and values are IFs.
            if_empty_word (str): The word that will replace nan values in column the returned column.
        
        Returns:
            (pandas serie): The column of the dataframe of interest 
                            that contains the IFs values. 
        '''
        if_column = issn_column.map(if_dict)
        if_column = if_column.fillna(if_empty_word)
        return if_column

    # Setting useful aliases
    year_col_alias            = COL_NAMES_FINALE[COL_NAMES['articles'][2]]
    issn_col_alias            = COL_NAMES['articles'][10]
    eissn_col_alias           = COL_NAMES_BONUS['EISSN']
    database_if_col_alias     = COL_NAMES_BONUS['IF clarivate']
    current_year_if_col_alias = COL_NAMES_BONUS['IF en cours']
    corpus_year_if_col_alias  = COL_NAMES_BONUS['IF année publi']
    if_update_col_list_alias  = COL_MAJ_IF
    not_available_if_alias    = NOT_AVAILABLE_IF
    unknown_if_fill_alias     = FILL_EMPTY_KEY_WORD
    unknown_alias             = FILL_EMPTY_KEY_WORD
    
    # Getting the df of the IFs database
    if_df = pd.read_excel(if_path, sheet_name = None)
    
    # Getting the df where to add IFs
    corpus_df = pd.read_excel(in_file_path)
    
    # Setting useful internal variables
    if_available_years_list = list(if_df.keys())
    if_most_recent_year = if_available_years_list[-1]
    most_recent_year_if_dict = dict(zip(if_df[if_most_recent_year][issn_col_alias], 
                                        if_df[if_most_recent_year][database_if_col_alias]))
        
    if year == None:
        # Setting type of values in 'year_col_alias' as string
        corpus_df = corpus_df.astype({year_col_alias : str})
        
        # Building the 'years' list of years in 'corpus_df' removing 'unkown_alias' years
        years = list(corpus_df[year_col_alias].unique())
        if unknown_alias in years : years.remove(unknown_alias)
        
        # Initializing 'corpus_df_bis' that will concatenate, on all years, 
        # the 'corpus_df' info added with IF values
        corpus_df_bis = pd.DataFrame()
        
        # Appending, year by year in 'years', the 'corpus_df' info completed with IF values, to 'corpus_df_bis' 
        for year in years:
            # Initialize 'inter_df' as copy of 'corpus_df' for column 'year_col_alias' value = 'year'
            inter_df = corpus_df[corpus_df[year_col_alias] == year].copy()
            
            # Adding 'current_year_if_col_alias' column to 'inter_df' 
            # with values defined by internal function '_create_if_column' 
            inter_df[current_year_if_col_alias] = _create_if_column(inter_df[issn_col_alias],
                                                                    most_recent_year_if_dict,
                                                                    unknown_if_fill_alias)            
            
            # Adding 'corpus_year_if_col_alias' column to 'inter_df' 
            if year in if_available_years_list:
                # with values defined by internal function '_create_if_column'
                year_if_dict = dict(zip(if_df[year][issn_col_alias], 
                                        if_df[year][database_if_col_alias]))
                inter_df[corpus_year_if_col_alias] = _create_if_column(inter_df[issn_col_alias],
                                                                       year_if_dict,
                                                                       unknown_if_fill_alias)
            else:
                # with 'not_available_if_alias' value 
                inter_df[corpus_year_if_col_alias] = not_available_if_alias
            
            # Appending 'inter_df' to 'corpus_df_bis' 
            corpus_df_bis = corpus_df_bis.append(inter_df)
            
        # Saving 'corpus_df_bis' as EXCEL file at full path 'out_file_path'
        corpus_df_bis.to_excel(out_file_path, columns = if_update_col_list_alias, index = False)    
        end_message = f"IFs updated in file : \n  '{out_file_path}'"
    
    else: 
        # Initialize 'corpus_df_bis' as copy of 'corpus_df'
        corpus_df_bis = corpus_df.copy()
        
        # Adding 'current_year_if_col_alias' column to 'corpus_df_bis' 
        # with values defined by internal function '_create_if_column'
        corpus_df_bis[current_year_if_col_alias] = _create_if_column(corpus_df_bis[issn_col_alias],
                                                                     most_recent_year_if_dict,
                                                                     unknown_if_fill_alias)        

        # Adding 'corpus_year_if_col_alias' column to 'corpus_df_bis' 
        if year in if_available_years_list:
            # with values defined by internal function '_create_if_column'
            current_year_if_dict = dict(zip(if_df[year][issn_col_alias], 
                                            if_df[year][database_if_col_alias]))
            corpus_df_bis[corpus_year_if_col_alias] = _create_if_column(corpus_df_bis[issn_col_alias],
                                                                        current_year_if_dict,
                                                                        unknown_if_fill_alias)
        else:
            # with 'not_available_if_alias' value
            corpus_df_bis[corpus_year_if_col_alias] = not_available_if_alias
        
        
        # Saving 'corpus_df_bis' as EXCEL file at full path 'out_file_path'
        corpus_df_bis.to_excel(out_file_path, index = False)
        end_message = f"IFs added for year {year} in file : \n  '{out_file_path}'"
    
    return end_message
                
    
def concatenate_pub_lists(bibliometer_path, years_list):
    
    """
    """
    # Standard library imports
    import os
    from pathlib import Path
    from datetime import datetime
    
    # 3rd library imports
    import pandas as pd    
    
    # Local library imports
    from BiblioMeter_GUI.Globals_GUI import ARCHI_BDD_MULTI_ANNUELLE
    from BiblioMeter_GUI.Globals_GUI import ARCHI_YEAR
    
    # Setting useful aliases
    pub_list_path_alias      = ARCHI_YEAR["pub list folder"]
    pub_list_file_base_alias = ARCHI_YEAR["pub list file name base"]
    bdd_multi_annuelle_folder_alias = ARCHI_BDD_MULTI_ANNUELLE["root"]
    bdd_multi_annuelle_file_alias   = ARCHI_BDD_MULTI_ANNUELLE["concat file name base"]
    
    # Building the concatenated dataframe of available publications lists
    df_concat = pd.DataFrame()    
    available_liste_conso = ""    
    for i in range(len(years_list)):
        try:
            year = years_list[i]
            pub_list_file_name = f"{pub_list_file_base_alias} {year}.xlsx"
            pub_list_path = bibliometer_path / Path(year) / Path(pub_list_path_alias) / Path(pub_list_file_name)
            df_inter = pd.read_excel(pub_list_path)
            df_concat = df_concat.append(df_inter)        
            available_liste_conso += f" {year}"
        
        except FileNotFoundError:
            pass
    
    # Saving the concatenated dataframe in an EXCEL file
    date = str(datetime.now())[:16].replace(':', 'h')
    out_file = f"{date} {bdd_multi_annuelle_file_alias} {os.getlogin()}_{available_liste_conso}.xlsx"
    out_path = bibliometer_path / Path(bdd_multi_annuelle_folder_alias)
    df_concat.to_excel(out_path / Path(out_file), index = False)
    
    end_message  = f"Concatenation of consolidated pub lists saved in folder: \n  '{out_path}'"
    end_message += f"\n\n under filename: \n  '{out_file}'."
    return end_message
    
            
def mise_en_page(df):
    
    ''' 
        
    '''
    
    # 3rd party imports
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows as openpyxl_dataframe_to_rows
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    from openpyxl.styles import Font as openpyxl_Font  
    from openpyxl.styles import PatternFill as openpyxl_PatternFill 
    from openpyxl.styles import Alignment as openpyxl_Alignment
    
    # Local imports
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_SIZES
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import ROW_COLORS

    cell_colors = [openpyxl_PatternFill(fgColor = ROW_COLORS['odd'], fill_type = "solid"),
                  openpyxl_PatternFill(fgColor = ROW_COLORS['even'], fill_type = "solid")]

    columns_list = list(df.columns)
    
    # Initialize wb as a workbook and ws its active worksheet
    wb = Workbook()
    ws = wb.active
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)
    
    #  Coloring alternatly rows in ws
    for idx, row in enumerate(ws_rows):
        ws.append(row)        
        last_row = ws[ws.max_row]            
        if idx >= 1:
            cell_color = cell_colors[idx%2]
            for cell in last_row:
                cell.fill = cell_color

    for cell in ws['A'] + ws[1]:
        cell.font = openpyxl_Font(bold=True)
        cell.alignment = openpyxl_Alignment(horizontal="center", vertical="center")

    for idx, col in enumerate(columns_list):
        if idx >= 1:
            column_letter = openpyxl_get_column_letter(idx + 1)
            try:
                ws.column_dimensions[column_letter].width = COL_SIZES[col]
            except:
                ws.column_dimensions[column_letter].width = 20
    
    first_row_num = 1
    ws.row_dimensions[first_row_num].height = 30

    return wb, ws


def find_missing_if(bibliometer_path, in_file_path):
    
    '''
    Genère un fichier des ISSN dont l'impact factor n'est pas dans la base de donnée.
    '''
    
    # Standard library imports
    from pathlib import Path
    
    # 3rd party imports
    import pandas as pd 
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local imports
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BiblioMeterGlobalsVariables import COL_NAMES_FINALE
    from BiblioMeter_GUI.Globals_GUI import ARCHI_IF

    # setting useful aliases
    if_issn_alias           = COL_NAMES['articles'][10]
    if_year_alias           = COL_NAMES_FINALE[COL_NAMES['articles'][2]]
    if_journal_alias        = COL_NAMES['articles'][3]    
    if_annee_publi_alias    = COL_NAMES_BONUS['IF année publi']
    if_root_folder_alias    = ARCHI_IF["root"]
    if_manquants_file_alias = ARCHI_IF["missing"] 
       
    out_folder_path = bibliometer_path / Path(if_root_folder_alias)
    out_file_path = out_folder_path / Path(if_manquants_file_alias)
    

    df = pd.read_excel(in_file_path)
    
    issn_list = list()
    for i in range(len(df)):
        if_value = df[if_annee_publi_alias][i]
        if if_value == 'unknown' or if_value == 'Not available':
            issn_value    = df[if_issn_alias][i]
            year_value    = df[if_year_alias][i]
            journal_value = df[if_journal_alias][i]
            issn_list.append([issn_value, year_value, journal_value])
    
    df = pd.DataFrame(issn_list, columns = [if_issn_alias, if_year_alias, if_journal_alias])
    df.drop_duplicates(inplace = True)
    
    df.to_excel(out_file_path, index = False)
    
    end_message  = f"List of ISSNs without IF saved in folder: \n '{out_folder_path}'"
    end_message += f"\n\n under file '{if_manquants_file_alias}'"
    return end_message