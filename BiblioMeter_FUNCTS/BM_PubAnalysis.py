__all__ = ['if_analysis',
           'keywords_analysis',          
          ]


def formatting_df(df, first_col_width):
    """
    """
    # 3rd party imports
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows as openpyxl_dataframe_to_rows
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    from openpyxl.styles import Font as openpyxl_Font  
    from openpyxl.styles import PatternFill as openpyxl_PatternFill 
    from openpyxl.styles import Alignment as openpyxl_Alignment
    from openpyxl.styles import Border as openpyxl_Border
    from openpyxl.styles import Side as openpyxl_Side
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import ROW_COLORS
    
    # Setting list of cell colors   
    cell_colors = [openpyxl_PatternFill(fgColor = ROW_COLORS['odd'], fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = ROW_COLORS['even'], fill_type = "solid")]   
    
    # Setting useful column attributes
    columns_list = list(df.columns)
    col_attr = {}
    col_attr[columns_list[0]] = [first_col_width, "left"]
    for col in columns_list[1:]:
        col_attr[col] = [15, "center"]

    # Initializing wb as a workbook and ws its active worksheet
    wb = Workbook()
    ws = wb.active
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)
        
    # Coloring alternatly rows in ws using list of cell colors cell_colors
    for idx_row, row in enumerate(ws_rows):       
        ws.append(row)      
        last_row = ws[ws.max_row]
        if idx_row >= 1:
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color 

    # Setting cell alignement and border using dict of column attributes col_attr
    for idx_col, col in enumerate(columns_list):
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        for cell in ws[column_letter]:
            cell.alignment = openpyxl_Alignment(horizontal=col_attr[col][1], vertical="center")
            cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick', color='FFFFFF'),
                                          right=openpyxl_Side(border_style='thick', color='FFFFFF'))
            
    # Setting the format of the columns heading
    cells_list = ws[1]
    for cell in cells_list:
        cell.font = openpyxl_Font(bold=True)
        cell.alignment = openpyxl_Alignment(wrap_text=True, horizontal="center", vertical="center")  
        
    # Setting de columns width using dict of column attributes col_attr
    for idx_col, col in enumerate(columns_list):        
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        ws.column_dimensions[column_letter].width = col_attr[col][0]
    
    # Setting height of rows
    height = 25
    for idx_row in range(ws.max_row):
        row_num = idx_row + 1
        if row_num >= 1: height = 20
        ws.row_dimensions[row_num].height = height

    return wb, ws  


###################################
# IFs analysis specific functions #
###################################

def _build_analysis_books_data(books_df):
    """
    """
    # Standard Library imports
    import numpy as np
    from pathlib import Path
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import COL_NAMES_IF_ANALYSIS
    from BiblioMeter_FUNCTS.BM_PubGlobals import KPI_KEYS_ORDER_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE 
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    
    # Setting useful column names aliases
    col_final_list    = set_final_col_names()
    book_col_alias    = col_final_list[6]
    doctype_col_alias = col_final_list[7]
    depts_col_list    = col_final_list[11:16] 
    
    # Setting new col names and related parameters
    chapters_nb_col_alias  = COL_NAMES_IF_ANALYSIS['articles_nb']
    
    # Building the books KPIs dict covering all departments
    books_kpi_dict = {}
    for dept in [INSTITUTE] + depts_col_list:
        
        # Building the books dataframe for "dept"
        if dept!= INSTITUTE:
            dept_books_df = books_df[books_df[dept]==1].copy()
        else:
            dept_books_df = books_df.copy()
        
        # Adding a column with number of articles per journal then droping duplicate rows
        count_books_df = dept_books_df[book_col_alias].value_counts().to_frame()
        count_books_df.rename(columns = {book_col_alias: chapters_nb_col_alias}, inplace = True)
        count_books_df.reset_index(inplace = True)
        count_books_df.rename_axis("idx", axis = 1,  inplace = True)        
        count_books_df.rename(columns = {"index": book_col_alias,}, inplace = True)        
        dept_books_df = dept_books_df.drop_duplicates([book_col_alias])
        dept_books_df = count_books_df.merge(dept_books_df, how = "outer", on = book_col_alias)
        
        # Computing KPI for department'dept'        
            ## Number of items by doctype
        nb_books = round(len(dept_books_df))
        
            ## Number of documents by doctype
        nb_chapters = round(dept_books_df[chapters_nb_col_alias].sum())
        
            ## Ratio of chapters per book
        chapters_per_book     = 0 
        chapters_per_book_max = 0
        if nb_books: 
            chapters_per_book     = round(nb_chapters / nb_books,1)      
            chapters_per_book_max = round(dept_books_df[chapters_nb_col_alias].max())
        
        # Building the dict of books KPIs of department 'dept'
        dept_kpi_dict = {}
        dept_kpi_dict = {KPI_KEYS_ORDER_DICT[2] : nb_books,
                         KPI_KEYS_ORDER_DICT[3] : nb_chapters,
                         KPI_KEYS_ORDER_DICT[4] : chapters_per_book,
                         KPI_KEYS_ORDER_DICT[5] : chapters_per_book_max,
                        }
        
        # Updating KPI dict with KPIs of department 'dept'
        books_kpi_dict[dept] = dept_kpi_dict
        
    return books_kpi_dict


def _build_analysis_if_data(corpus_year, analysis_df, if_col_dict, books_kpi_dict, 
                            if_analysis_col, if_analysis_year,
                            if_analysis_folder_path, verbose = True):
    """
    
    """
    # Standard Library imports
    import numpy as np
    from pathlib import Path
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import COL_NAMES_IF_ANALYSIS
    from BiblioMeter_FUNCTS.BM_PubGlobals import DOC_TYPE_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import KPI_KEYS_ORDER_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE 
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    
    # Setting useful aliases
    doctype_article_alias  = DOC_TYPE_DICT['Articles']
    
    # Setting useful column names aliases
    col_final_list         = set_final_col_names()
    journal_col_alias      = col_final_list[6]
    doctype_col_alias      = col_final_list[7]
    issn_col_alias         = col_final_list[10]
    depts_col_list         = col_final_list[11:16]
    journal_norm_col_alias = COL_NAMES['temp_col'][1]
    articles_nb_col_alias  = COL_NAMES_IF_ANALYSIS['articles_nb']
    
    # Setting useful columns list
    if_col_list = list(if_col_dict.keys()) 
    
    # Building the full KPIs dict covering all departments
    kpi_dict = {}    
    for dept in [INSTITUTE] + depts_col_list:
        
        # Building the IFs analysis dataframe for "dept"
        if dept!= INSTITUTE:
            dept_analysis_df = analysis_df[analysis_df[dept]==1].copy()
        else:
            dept_analysis_df = analysis_df.copy() 
        dept_analysis_df.drop(columns = depts_col_list, inplace = True)
            
        # Adding a column with number of articles per journal then droping duplicate rows
        count_journal_df = dept_analysis_df[issn_col_alias].value_counts().to_frame()
        count_journal_df.rename(columns = {issn_col_alias: articles_nb_col_alias}, inplace = True)
        count_journal_df.reset_index(inplace = True)
        count_journal_df.rename_axis("idx", axis = 1,  inplace = True)
        count_journal_df.rename(columns = {"index":issn_col_alias}, inplace = True)
        dept_analysis_df = dept_analysis_df.drop_duplicates([issn_col_alias])
        dept_analysis_df = count_journal_df.merge(dept_analysis_df, how = "outer", on = issn_col_alias)              

        # Keeping only articles of journal (not proceedings) and droping 'doctype_col_alias' column
        dept_articles_df = dept_analysis_df[dept_analysis_df[doctype_col_alias].isin(doctype_article_alias)]
        dept_articles_df = dept_articles_df.drop(columns = [doctype_col_alias])
    
        # Computing KPIs independant of IF for department'dept' 
            ## Number of items by doctype
        nb_journals_proceedings = round(len(dept_analysis_df))
        nb_journals             = round(len(dept_articles_df))
        
            ## Number of documents by doctype
        nb_articles_communications = round(dept_analysis_df[articles_nb_col_alias].sum())
        nb_articles                = round(dept_articles_df[articles_nb_col_alias].sum())
        nb_communications          = round(nb_articles_communications - nb_articles)
        communications_ratio       = 0
        if nb_articles_communications:
            communications_ratio = round(nb_communications / nb_articles_communications * 100)

            ## Ratio of articles per journal
        articles_per_journal     = 0
        articles_per_journal_max = 0
        if nb_journals: 
            articles_per_journal     = round(nb_articles / nb_journals,1)
            articles_per_journal_max = round(dept_articles_df[articles_nb_col_alias].max())
            
            ## Total number of publications
        nb_chapters     = books_kpi_dict[dept][KPI_KEYS_ORDER_DICT[3]]
        nb_publications = round(nb_chapters + nb_articles_communications)        
        
        # Completing the KPIs dict with IF independant KPIs of department 'dept'
        dept_kpi_dict = {k:v for k,v in books_kpi_dict[dept].items()}
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[1]]  = nb_publications
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[6]]  = nb_journals_proceedings
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[7]]  = nb_journals
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[8]]  = nb_articles_communications
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[9]]  = nb_communications
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[10]] = communications_ratio        
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[11]] = nb_articles
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[12]] = articles_per_journal
        dept_kpi_dict[KPI_KEYS_ORDER_DICT[13]] = articles_per_journal_max
               
        # Selecting useful columns for barchart plot
        dept_if_df = dept_articles_df[[journal_col_alias, articles_nb_col_alias, if_analysis_col]]           

        # Cleaning the dataframe 'dept_if_df'
        dept_if_df = dept_if_df.sort_values(by = [if_analysis_col, journal_col_alias], ascending = True)
        if_analysis_col_new = "IF " + if_analysis_year
        dept_if_df.rename(columns = {if_analysis_col: if_analysis_col_new}, inplace = True)
        dept_if_df = dept_if_df.reset_index().drop(columns = ["index"])

        # Computing IF useful values   
        if_moyen = sum([x[0]*x[1] for x in zip(dept_if_df[if_analysis_col_new],dept_if_df[articles_nb_col_alias])])/nb_articles
        if_max   = np.max(dept_if_df[if_analysis_col_new])
        if_min   = np.min(dept_if_df[if_analysis_col_new])

        # Computing corrected IF useful values
        dept_if_sub_df = dept_if_df[dept_if_df[if_analysis_col_new]==0]                                    
        nb_art_wo_if   = dept_if_sub_df[articles_nb_col_alias].sum()
        wo_if_ratio    = 0
        if nb_articles: wo_if_ratio = nb_art_wo_if / nb_articles * 100                       

        # Completing the KPIs dict with IF useful values
        dept_kpi_dict[if_analysis_col_new] = {KPI_KEYS_ORDER_DICT[15] : round(if_max,1),
                                              KPI_KEYS_ORDER_DICT[16] : round(if_min,1),
                                              KPI_KEYS_ORDER_DICT[17] : round(if_moyen,1),
                                              KPI_KEYS_ORDER_DICT[18] : round(nb_art_wo_if),
                                              KPI_KEYS_ORDER_DICT[19] : round(wo_if_ratio),
                                             }        
        
        # Updating KPI dict with KPIs of department 'dept'
        kpi_dict[dept] = dept_kpi_dict
        
        # Saving after formatting the updated dataframe as EXCEL file
        file_name = f'{if_analysis_col_new}-{dept}'
        dept_xlsx_file_path = Path(if_analysis_folder_path) / Path(file_name + '.xlsx')
        first_col_width = 50
        wb, ws = formatting_df(dept_if_df, first_col_width)
        ws.title = dept + ' IFs '
        wb.save(dept_xlsx_file_path)
        
        message  = f"\n    EXCEL file of {if_analysis_col_new} for {dept} department "
        message += f"saved in : \n {if_analysis_folder_path}"
        if verbose: print(message, "\n")

    return kpi_dict, if_analysis_col_new


def _update_kpi_database(bibliometer_path, corpus_year, kpi_dict, if_key, verbose = False):
    """
    """
    
    # Standard Library imports
    import os
    from pathlib import Path
    
    # 3rd parties import
    import pandas as pd
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import ARCHI_BDD_MULTI_ANNUELLE
    from BiblioMeter_FUNCTS.BM_PubGlobals import KPI_KEYS_ORDER_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE 
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    
    # Setting useful aliases
    kpi_folder_alias         = ARCHI_BDD_MULTI_ANNUELLE["root"]
    kpi_file_base_alias      = ARCHI_BDD_MULTI_ANNUELLE["IF analysis file name base"]
    
    # Setting useful column names aliases
    col_final_list        = set_final_col_names()
    depts_col_list        = col_final_list[11:16]  
    corpus_year_row_alias = KPI_KEYS_ORDER_DICT[0]

    # Building as a dataframe the column of KPIs of each 'dept'
    for dept in [INSTITUTE] + depts_col_list:
        
        # Building sub dict of KPIs dict for 'dept'
        dept_kpi_dict = {k:v for k,v in kpi_dict[dept].items()}
        
        # Building sub dict of publications KPIs of 'dept' in keys order specified by 'ordered_keys'
        ordered_keys  = [v for k,v in KPI_KEYS_ORDER_DICT.items() if k in range(1,14)]
        
        dept_pub_dict = {k: dept_kpi_dict[k] for k in ordered_keys}
        
        # Building 'dept_pub_df' using keys of 'dept_pub_dict' as indexes 
        # and setting the name of the values column to 'corpus_year'       
        dept_pub_df = pd.DataFrame.from_dict(dept_pub_dict, orient = "index", columns = [corpus_year])
        
        # Renaming the index column of 'dept_pub_df' as 'corpus_year_row_alias'
        dept_pub_df.reset_index(inplace = True)
        dept_pub_df.rename_axis("idx", axis = 1,  inplace = True)        
        dept_pub_df.rename(columns = {"index":corpus_year_row_alias,}, inplace = True)
        
        # Building sub dict of IFs KPIs of 'dept' in keys order specified by 'ordered_keys'        
        part_dept_if_dict = {k:v for k,v in dept_kpi_dict[if_key].items()}
        dept_if_dict = dict({KPI_KEYS_ORDER_DICT[14] : if_key}, **part_dept_if_dict)
        
        # Building 'dept_if_df' using keys of 'dept_if_dict' as indexes 
        # and setting the name of the values column to 'corpus_year'      
        dept_if_df = pd.DataFrame.from_dict(dept_if_dict, orient = "index", columns = [corpus_year]) 
        
        # Renaming the index column with 'corpus_year_row_alias'
        dept_if_df.reset_index(inplace = True)
        dept_if_df.rename_axis("idx", axis = 1,  inplace = True)        
        dept_if_df.rename(columns = {"index":corpus_year_row_alias,}, inplace = True) 

        ## Combining the two dataframes through rows concatenation
        dept_kpi_df = pd.DataFrame()
        dept_kpi_df = pd.concat([dept_pub_df, dept_if_df], axis = 0)
        
        # Reading as the dataframe the KPI file of 'dept' if it exists else creating it
        filename = dept + "_" + kpi_file_base_alias + ".xlsx"
        file_path = bibliometer_path / Path(kpi_folder_alias) / Path(filename)       
        if os.path.isfile(file_path):
            db_dept_kpi_df = pd.read_excel(file_path)
            # Updating the dataframe with the column to append
            if corpus_year in db_dept_kpi_df.columns:
                db_dept_kpi_df = db_dept_kpi_df.drop(columns = [corpus_year])           
            db_dept_kpi_df = db_dept_kpi_df.merge(dept_kpi_df, how = "outer", on = corpus_year_row_alias)
        else:
            db_dept_kpi_df = dept_kpi_df
            
        # Saving after formatting the updated dataframe
        first_col_width = 35
        wb, ws = formatting_df(db_dept_kpi_df, first_col_width)
        ws.title = dept + ' KPIs '
        wb.save(file_path)
        
        if dept== INSTITUTE : institute_kpi_df = db_dept_kpi_df
        
    message = f"\n    Kpi database updated and saved in folder: \n {file_path}"   
    if verbose: print(message)
    
    return institute_kpi_df
    
def _create_if_barchart(corpus_year, dept, if_df, if_col, kpi_dict, journal_col_alias, part = "all"):
    """
    """
    # Standard Library imports
    import numpy as np
    import plotly
    import plotly.express as px
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_COLOR_RANGE
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_COLOR_SCALE
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_HEIGHT
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_HEIGHT_RATIO
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_X_RANGE
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_Y_LABEL_MAX
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_Y_MAX
    from BiblioMeter_FUNCTS.BM_PubGlobals import BAR_WIDTH 
    from BiblioMeter_FUNCTS.BM_PubGlobals import COL_NAMES_IF_ANALYSIS
    from BiblioMeter_FUNCTS.BM_PubGlobals import KPI_KEYS_ORDER_DICT
    
    # internal functions               
    _shorting_journal_name = lambda x: x[:max_journal_short_name]+'...' if len(x)>max_journal_short_name else x
    
    # Setting new col names and related parameters
    journal_short_col_alias = COL_NAMES_IF_ANALYSIS['journal_short']
    articles_nb_col_alias   = COL_NAMES_IF_ANALYSIS['articles_nb']
    max_journal_short_name  = BAR_Y_LABEL_MAX
    
    # Creating columns with shortnames of journals for barchart plots
    plot_df = if_df.copy()
    plot_df[journal_short_col_alias] = plot_df[journal_col_alias].apply(_shorting_journal_name)    
    
    # Setting useful values for barchart plot and title     
    dept_kpi_dict = kpi_dict[dept]
    nb_journals          = dept_kpi_dict[KPI_KEYS_ORDER_DICT[7]]            #"journals number"
    nb_articles          = dept_kpi_dict[KPI_KEYS_ORDER_DICT[11]]           #"articles number"
    articles_per_journal = dept_kpi_dict[KPI_KEYS_ORDER_DICT[12]]           #"articles_per_journal"
    nb_articles_max      = dept_kpi_dict[KPI_KEYS_ORDER_DICT[13]]           #"max articles_per_journal"
    if_max               = dept_kpi_dict[if_col][KPI_KEYS_ORDER_DICT[15]]   #"IF max"
    if_min               = dept_kpi_dict[if_col][KPI_KEYS_ORDER_DICT[16]]   #"IF min"
    if_moyen             = dept_kpi_dict[if_col][KPI_KEYS_ORDER_DICT[17]]   #"IF mean"
    wo_if_ratio          = dept_kpi_dict[if_col][KPI_KEYS_ORDER_DICT[19]]   #"Articles-w/o-IF ratio"
       
    # Setting the first part of the barchart title
    title_base  = f"{dept} corpus {corpus_year}: "
    title_base += f"Journals = {nb_journals}, Articles = {nb_articles}, "
    title_base += f"Articles/Journal = {articles_per_journal: .1f}"
    
    # Completing the barchart title
    if_values = if_col
    if part != "all": 
        if_values += " " + part + " half"

    title_add = "<br>" + f"{if_values}: IF max = {if_max:.1f}, IF min = {if_min:.1f}, " 
    title_add += f"IF mean = {if_moyen:.1f}, Articles w/o IF = {wo_if_ratio:.0f} %" + "<br>"
    title = title_base + title_add
    
    # Setting barchart parameters   
    labels_dict       = {articles_nb_col_alias  : 'Articles number',
                         journal_short_col_alias: 'Short name'}
    nb_articles_range = BAR_X_RANGE
    barchart_width    = BAR_WIDTH
    barchart_height   = BAR_HEIGHT 
    if nb_journals <= BAR_Y_MAX or part != "all": 
        barchart_height = round(BAR_HEIGHT / BAR_HEIGHT_RATIO)
    color_range       = BAR_COLOR_RANGE
    color_scale       = BAR_COLOR_SCALE
    
    barchart = px.bar(data_frame             = plot_df,
                      x                      = articles_nb_col_alias,
                      y                      = journal_short_col_alias,
                      orientation            = 'h',
                      title                  = title,
                      color                  = if_col,
                      color_continuous_scale = color_scale,
                      range_color            = color_range,
                      labels                 = labels_dict,
                      width                  = barchart_width,
                      height                 = barchart_height,
                      hover_name             = journal_col_alias,
                      hover_data             = {journal_short_col_alias: False,
                                                if_col: ':.1f'},
                      range_x                = nb_articles_range,
                     )
    
    return barchart


def _save_dept_barchart(barchart, dept, if_col, if_analysis_folder_path, part = "all"):
    """
    """
    # Standard Library imports
    from pathlib import Path
    
    file_name = f"{if_col}-{dept}"    
    if part != "all": file_name += f"_{part}"
    
    dept_html_file_path = Path(if_analysis_folder_path) / Path(file_name + ".html")
    barchart.write_html(dept_html_file_path)

    dept_png_file_path  = Path(if_analysis_folder_path) / Path(file_name + ".png")
    barchart.write_image(dept_png_file_path)
    
    end_message  = f"\n    Barchart of {if_col} ({part} values) for {dept} "
    end_message += f"department saved in : \n {if_analysis_folder_path}"
    return end_message


def _plot_if_analysis(corpus_year, kpi_dict, if_col, if_analysis_folder_path, verbose = True):
    """
    Module internal functions: _create_if_barchart, _save_dept_barchart
     
    """
    # Standard Library imports
    import pandas as pd
    from pathlib import Path
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import KPI_KEYS_ORDER_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE 
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    
    # internal functions    
    def _create_save_barchart(dept, bar_chart_if_df, part):
        barchart = _create_if_barchart(corpus_year, dept, bar_chart_if_df, if_col, kpi_dict, journal_col_alias, part)
        message = _save_dept_barchart(barchart, dept, if_col, if_analysis_folder_path, part)
        return message
        
    # Setting useful column names aliases
    col_final_list          = set_final_col_names()
    journal_col_alias       = col_final_list[6]   
    depts_col_list          = col_final_list[11:16]
    
    for dept in [INSTITUTE] + depts_col_list:
        dept_kpi_dict = kpi_dict[dept]        
        file_name     = f'{if_col}-{dept}'
        dept_xlsx_file_path = Path(if_analysis_folder_path) / Path(file_name + '.xlsx')
        dept_if_df = pd.read_excel(dept_xlsx_file_path)

        if dept == INSTITUTE:
            # Setting two dataframes with, respectively, upper and lower values of full IF dataframe of INSTITUTE
            nb_journals       = dept_kpi_dict[KPI_KEYS_ORDER_DICT[7]]
            journal_median    = dept_if_df.loc[int(nb_journals/2), journal_col_alias]
            if_median         = dept_if_df[dept_if_df[journal_col_alias]==journal_median][if_col].values[0]
            upper_dept_if_df  = dept_if_df[dept_if_df[if_col]>=if_median]
            lower_dept_if_df  = dept_if_df[dept_if_df[if_col]<if_median]

            # creating barchart with full IF dataframe of INSTITUTE
            message = _create_save_barchart(dept, dept_if_df, "all")
            if verbose: print(message, "\n")

            # creating barchart with upper values of IF dataframe of INSTITUTE
            message = _create_save_barchart(dept, upper_dept_if_df, "upper")           
            if verbose: print(message, "\n")

            # creating barchart with upper values of IF dataframe of INSTITUTE
            message = _create_save_barchart(dept, lower_dept_if_df, "lower")  
            if verbose: print(message, "\n")

        else:
            # creating barchart with full IF dataframe of dept
            message = _create_save_barchart(dept, dept_if_df, "all")        
            if verbose: print(message, "\n")
    
    end_message = f"\n    IF analysis plots for corpus {corpus_year} saved in : \n {if_analysis_folder_path}"
    if verbose: print(end_message, "\n")
    return    
    

def if_analysis(bibliometer_path, corpus_year, if_most_recent_year, verbose = True):
    """
    
    Module internal functions: _build_analysis_if_data, _plot_if_analysis
    """
    
    # Standard Library imports
    import numpy as np
    import os
    from pathlib import Path
    
    # 3rd parties import
    import pandas as pd
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DEDUPLICATED_XLSX
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import UNKNOWN
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import ARCHI_YEAR
    from BiblioMeter_FUNCTS.BM_PubGlobals import ANALYSIS_IF
    from BiblioMeter_FUNCTS.BM_PubGlobals import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BM_PubGlobals import DOC_TYPE_DICT
    from BiblioMeter_FUNCTS.BM_PubGlobals import DOCTYPE_TO_SAVE_DICT    
    from BiblioMeter_FUNCTS.BM_PubGlobals import NOT_AVAILABLE_IF
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    from BiblioMeter_FUNCTS.BM_UpdateImpactFactors import journal_capwords
    
    # internal functions  
    
    def _unique_journal_name(init_analysis_df):
        """Sets a unique journal name by ISSN value.
        """
        analysis_df = pd.DataFrame(columns = init_analysis_df.columns)
        for _, df in init_analysis_df.groupby(by=[issn_col_alias]):
            issn_df = df.copy()
            issn = issn_df[issn_col_alias].to_list()[0]
            journal_names_list = issn_df[journal_col_alias].to_list()
            if len(journal_names_list)>1:
                if issn!=UNKNOWN:
                    journal_length_list = [len(journal) for journal in journal_names_list]
                    journal_names_dict  = dict(zip(journal_length_list,journal_names_list))
                    length_min          = min(journal_length_list)
                    issn_df[journal_col_alias] = journal_names_dict[length_min]                
                else:
                    journal_names_list = list(set(issn_df[journal_col_alias].to_list()))
                    journal_issn_list  = [issn + str(num) for num in range(len(journal_names_list))]
                    journal_names_dict = dict(zip(journal_names_list,journal_issn_list))
                    issn_df[issn_col_alias] = issn_df[journal_col_alias].copy()
                    issn_df[issn_col_alias] = issn_df[issn_col_alias].map(journal_names_dict)                
            analysis_df = pd.concat([analysis_df, issn_df], ignore_index=True)
        return analysis_df
    
    _capwords_journal_col = lambda row: journal_capwords(row[journal_col_alias])    
    _replace_no_if = lambda x: x if x != UNKNOWN and x != NOT_AVAILABLE_IF else 0
    
    # Setting useful aliases for paths and filenames building
    corpus_folder_alias            = ARCHI_YEAR["corpus"]
    dedup_folder_alias             = ARCHI_YEAR["dedup"]
    parsing_folder_alias           = ARCHI_YEAR["parsing"]
    pub_list_folder_alias          = ARCHI_YEAR["pub list folder"]
    analysis_folder_alias          = ARCHI_YEAR["analyses"]
    if_analysis_folder_alias       = ARCHI_YEAR["if analysis"]
    pub_list_filename_base         = ARCHI_YEAR["pub list file name base"]
    papers_doctype_alias           = list(DOCTYPE_TO_SAVE_DICT.keys())[0]
    books_doctype_alias            = list(DOCTYPE_TO_SAVE_DICT.keys())[1]
    dedup_articles_file_name_alias = DEDUPLICATED_XLSX
    
    # Setting useful file names 
    papers_list_filename = pub_list_filename_base + " " + corpus_year + "_" + papers_doctype_alias + ".xlsx"
    books_list_filename  = pub_list_filename_base + " " + corpus_year + "_" + books_doctype_alias + ".xlsx"

    # Setting useful paths
    year_folder_path         = bibliometer_path / Path(corpus_year)
    corpus_folder_path       = year_folder_path / Path(corpus_folder_alias) 
    dedup_folder_path        = corpus_folder_path / Path(dedup_folder_alias)
    parsing_folder_path      = dedup_folder_path / Path(parsing_folder_alias)
    dedup_articles_file_path = parsing_folder_path / Path(dedup_articles_file_name_alias)
    pub_list_folder_path     = year_folder_path / Path(pub_list_folder_alias)
    papers_list_file_path    = pub_list_folder_path / Path(papers_list_filename)
    books_list_file_path     = pub_list_folder_path / Path(books_list_filename)
    analysis_folder_path     = year_folder_path / Path(analysis_folder_alias)
    if_analysis_folder_path  = analysis_folder_path / Path(if_analysis_folder_alias)
                                       
    # Creating required output folders
    if not os.path.exists(analysis_folder_path):
        os.makedirs(analysis_folder_path) 
    if not os.path.exists(if_analysis_folder_path):
        os.makedirs(if_analysis_folder_path) 
    
    # Setting useful column names aliases
    col_final_list                     = set_final_col_names()
    journal_col_alias                  = col_final_list[6]
    doctype_col_alias                  = col_final_list[7] 
    issn_col_alias                     = col_final_list[10] 
    depts_col_list                     = col_final_list[11:16]
    journal_norm_col_alias             = COL_NAMES['temp_col'][1]
    most_recent_year_if_col_base_alias = COL_NAMES_BONUS["IF en cours"]
    corpus_year_if_col                 = COL_NAMES_BONUS['IF année publi']    
    most_recent_year_if_col_alias      = most_recent_year_if_col_base_alias + ", " + if_most_recent_year
    
    # Building the dataframe to be analysed from the file which full path is 'books_list_file_path'
    usecols = [journal_col_alias, doctype_col_alias] + depts_col_list
    books_df = pd.read_excel(books_list_file_path,
                             usecols = usecols)
    books_kpi_dict = _build_analysis_books_data(books_df) 
    
    # Setting the IF column dict
    if_col_dict = {most_recent_year_if_col_alias: if_most_recent_year,
                   corpus_year_if_col           : corpus_year}
    
    # Setting the IF analysis year and column   
    if if_most_recent_year>=corpus_year: 
        if_analysis_col  = ANALYSIS_IF
        if_analysis_year = if_col_dict[ANALYSIS_IF]
    else:
        if_analysis_col  = most_recent_year_if_col_alias
        if_analysis_year = if_most_recent_year
    
    # Building the dataframe of normalized journal names
    journal_norm_df = pd.read_excel(dedup_articles_file_path,
                                    usecols = [journal_col_alias, journal_norm_col_alias])   
    journal_norm_dict = dict(zip(journal_norm_df[journal_col_alias],journal_norm_df[journal_norm_col_alias]))
    
    # Building the dataframe to be analysed from the file which full path is 'papers_list_file_path'
    if_col_list = list(if_col_dict.keys())
    usecols = [journal_col_alias, doctype_col_alias, issn_col_alias] + depts_col_list + if_col_list
    init_analysis_df = pd.read_excel(papers_list_file_path,
                                     usecols = usecols)    
    analysis_df = _unique_journal_name(init_analysis_df)
    analysis_df[journal_norm_col_alias] = analysis_df[journal_col_alias]
    analysis_df[journal_norm_col_alias] = analysis_df[journal_norm_col_alias].map(journal_norm_dict)
    analysis_df[journal_col_alias]      = analysis_df.apply(_capwords_journal_col, axis=1)    
    for if_col, year in if_col_dict.items():
        analysis_df[if_col] = analysis_df[if_col].apply(_replace_no_if)  

    # Building the data resulting from IFs analysis and saving them as xlsx files
    kpi_dict, if_analysis_col_new = _build_analysis_if_data(corpus_year, analysis_df, if_col_dict, books_kpi_dict,
                                                            if_analysis_col, if_analysis_year,
                                                            if_analysis_folder_path, verbose = verbose)

    # Updating the KPIs database 
    institute_kpi_df = _update_kpi_database(bibliometer_path, corpus_year, kpi_dict, if_analysis_col_new, verbose = verbose)
    
    # Ploting IF analysis data as html files
    _plot_if_analysis(corpus_year, kpi_dict, if_analysis_col_new, if_analysis_folder_path, verbose = verbose)
    
    return if_analysis_folder_path, institute_kpi_df, kpi_dict


########################################
# Keywords analysis specific functions #
########################################
    
def _create_kw_analysis_data(year, analysis_df, kw_type, kw_df, usecols,
                             kw_analysis_folder_path, verbose = False):
    """
    """
    # Standard Library imports
    from pathlib import Path
    
    # 3rd parties import
    import pandas as pd 
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE
    
    # Setting useful column names aliases
    final_pub_id_col_alias   = usecols[0]    
    depts_col_list           = usecols[1]
    parsing_pub_id_col_alias = usecols[2]
    keywords_col_alias       = usecols[3]
    weight_col_alias         = usecols[4]
    
    # Analyzing the keywords for each of the department in 'depts_col_list'         
    for dept in [INSTITUTE] + depts_col_list:                                                 
        # Collecting and normalizing all the Pub_ids of the department 'dept'
        # by removing the 4 first characters corresponding to the corpus year
        # in order to make them comparable to 'parsing_pub_id_col_alias' values
        if dept!= INSTITUTE:
            dept_pub_id_list = [int(x[5:8]) for x in analysis_df[analysis_df[dept]==1][final_pub_id_col_alias].tolist()]
        else:
            dept_pub_id_list = [int(x[5:8]) for x in analysis_df[final_pub_id_col_alias].tolist()] 

        # Building the list of keywords for the keywords type 'kw_type' and the department 'dept'
        dept_kw_list = []
        for _, row in kw_df.iterrows():
            pub_id  = row[parsing_pub_id_col_alias]
            keyword = row[keywords_col_alias]              
            if pub_id in dept_pub_id_list: 
                pub_kw_list = [word.strip() for word in keyword.split(";")]
                dept_kw_list= dept_kw_list + pub_kw_list

        # Building a dataframe with the keywords and their weight for the keywords type 'kw_type'
        # and the department 'dept'
        dept_kw_df = pd.DataFrame(columns = [keywords_col_alias, weight_col_alias])
        dept_kw_set_to_list = sorted(list(set(dept_kw_list)))
        kw_drop = 0
        for idx, keyword in enumerate(dept_kw_set_to_list):
            if len(keyword)>1:
                dept_kw_df.loc[idx, keywords_col_alias] = keyword
                dept_kw_df.loc[idx, weight_col_alias]   = dept_kw_list.count(keyword)
            else:
                kw_drop +=1
        if kw_drop and dept==INSTITUTE:
            print(f"    WARNING: {kw_drop} dropped keywords of 1 character among {len(dept_kw_set_to_list)} {kw_type} ones of {INSTITUTE}")
        
        # Saving the keywords dataframe as EXCEL file
        dept_xlsx_file_path = Path(kw_analysis_folder_path) / Path(f'{dept} {year}-{kw_type}.xlsx')
        first_col_width = 50
        wb, ws = formatting_df(dept_kw_df, first_col_width)
        ws.title = dept + ' ' + kw_type
        wb.save(dept_xlsx_file_path)

    message = f"\n    Keywords of all types and all departments saved in : \n {kw_analysis_folder_path}"
    if verbose: print(message, "\n")    
    return


def keywords_cloud(txt, out, bckg, h, w, mxw, verbose = False):
    """
    Args:
        txt (str): Text which words will be plot as cloud.
        out (path): Full path of the png file that will contain the plot image.
        bckg (str): Color of the plot background.
        h (int): Height of the plot in pixels.
        w (int): Width of the plot in pixels.
        mxw (int): Maximum number of words to be plot.
        
    Returns:
        (str): Message about the completion of the image building.
        
    """
    # 3rd parties import
    from wordcloud import WordCloud
    
    wc = WordCloud(background_color = bckg,
                   height           = h,
                   width            = w,
                   max_words        = mxw)
    cloud = wc.generate(txt)
    cloud.to_file(out)
    
    message = f"\n    Wordcloud image saved in file: \n {out}"
    if verbose: print(message)
    return


def _create_kw_cloud(year, kw_type, kw_analysis_folder_path, usecols,
                     verbose = False):
    """
    """
    
    # Standard Library imports
    from pathlib import Path
    
    # 3rd parties import
    import pandas as pd 
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import UNKNOWN
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import CLOUD_BCKG
    from BiblioMeter_FUNCTS.BM_PubGlobals import CLOUD_HEIGHT
    from BiblioMeter_FUNCTS.BM_PubGlobals import CLOUD_WIDTH
    from BiblioMeter_FUNCTS.BM_PubGlobals import CLOUD_MAX_WORDS
    from BiblioMeter_FUNCTS.BM_PubGlobals import CLOUD_MAX_WORDS_LENGTH
    from BiblioMeter_FUNCTS.BM_PubGlobals import INSTITUTE
    
    # Setting the maximum length of the words for the cloud
    kw_length = CLOUD_MAX_WORDS_LENGTH
    
    # Setting useful column names aliases   
    depts_col_list           = usecols[1]
    keywords_col_alias       = usecols[3]
    weight_col_alias         = usecols[4]
    
    # creating the keywords text for each of the department in 'depts_col_list'         
    for dept in [INSTITUTE] + depts_col_list:
        
        # Getting the dataframe of keywords with their weight
        dept_xlsx_file_path = Path(kw_analysis_folder_path) / Path(f'{dept} {year}-{kw_type}.xlsx')
        dept_kw_df = pd.read_excel(dept_xlsx_file_path)
        dept_kw_df[keywords_col_alias] = dept_kw_df[keywords_col_alias].apply(lambda x: x[0:kw_length])
    
        # Building the keywords list with each keyword repeated up to its weight
        dept_kw_list = []
        for _, row in dept_kw_df.iterrows():
            keyword = row[keywords_col_alias]
            weight  = row[weight_col_alias]
            if keyword != UNKNOWN:
                keyword_list = [keyword] * weight
                dept_kw_list = dept_kw_list + keyword_list
                
        # Building the text 'dept_kw_txt' that contains the keywords       
        dept_kw_txt = ' '.join(dept_kw_list)
        dept_kw_txt.encode(encoding = 'UTF-8', errors = 'strict')
        
        # create and save the cloud image for department 'dept'
        if dept_kw_txt!='':
            dept_png_file_path = Path(kw_analysis_folder_path) / Path(f"{kw_type} {year}-{dept}.png")  
            keywords_cloud(dept_kw_txt, 
                           dept_png_file_path, 
                           CLOUD_BCKG, 
                           CLOUD_HEIGHT, 
                           CLOUD_WIDTH, 
                           CLOUD_MAX_WORDS)
            
    message = f"\n    Wordcloud images for all keywords types and all departments saved in : \n {kw_analysis_folder_path}"
    if verbose: print(message, "\n")    
    return

def keywords_analysis(bibliometer_path, year, verbose=False):
    """
    """
    
    # Standard Library imports
    import os
    from pathlib import Path
    
    # 3rd parties import
    import pandas as pd
    
    # BiblioAnalysis_Utils package imports
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    
    # Local globals imports
    from BiblioMeter_FUNCTS.BM_PubGlobals import ARCHI_YEAR
    from BiblioMeter_FUNCTS.BM_PubGlobals import COL_NAMES_BONUS
    from BiblioMeter_FUNCTS.BM_PubGlobals import DOCTYPE_TO_SAVE_DICT
    
    # Local library imports
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    
    # Setting useful aliases
    corpus_folder_alias      = ARCHI_YEAR["corpus"]
    dedup_folder_alias       = ARCHI_YEAR["dedup"]
    parsing_folder_alias     = ARCHI_YEAR["parsing"]
    pub_list_folder_alias    = ARCHI_YEAR["pub list folder"]
    analysis_folder_alias    = ARCHI_YEAR["analyses"] 
    kw_analysis_folder_alias = ARCHI_YEAR["keywords analysis"]    
    pub_list_filename_base   = ARCHI_YEAR["pub list file name base"]
    doctype_alias            = list(DOCTYPE_TO_SAVE_DICT.keys())[0]
    
    # Setting useful file names 
    pub_list_filename = pub_list_filename_base + " " + str(year) + "_" + doctype_alias + ".xlsx"

    # Setting useful paths
    year_folder_path        = bibliometer_path / Path(str(year))
    corpus_folder_path      = year_folder_path / Path(corpus_folder_alias) 
    dedup_folder_path       = corpus_folder_path / Path(dedup_folder_alias)
    parsing_folder_path     = dedup_folder_path / Path(parsing_folder_alias)
    pub_list_folder_path    = year_folder_path / Path(pub_list_folder_alias)
    pub_list_file_path      = pub_list_folder_path / Path(pub_list_filename)
    analysis_folder_path    = year_folder_path / Path(analysis_folder_alias)
    kw_analysis_folder_path = analysis_folder_path / Path(kw_analysis_folder_alias)
    
    # Creating required output folders
    if not os.path.exists(analysis_folder_path):
        os.makedirs(analysis_folder_path) 
    if not os.path.exists(kw_analysis_folder_path):
        os.makedirs(kw_analysis_folder_path)
    
    # Setting useful column names aliases
    col_final_list           = set_final_col_names()
    final_pub_id_col_alias   = col_final_list[0]
    depts_col_list           = col_final_list[11:16]
    parsing_pub_id_col_alias = COL_NAMES['pub_id'] 
    keywords_col_alias       = COL_NAMES['keywords'][1]
    weight_col_alias         = COL_NAMES_BONUS['weight']

    # Setting useful filenames dict
    kw_filename_dict = {'AK' : DIC_OUTDIR_PARSING['AK'],
                        'IK' : DIC_OUTDIR_PARSING['IK'],
                        'TK' : DIC_OUTDIR_PARSING['TK'],
                       }
    
    # Building the dataframe to be analysed from the file which full path is 'pub_list_file_path'  
    analysis_df = pd.read_excel(pub_list_file_path,
                                usecols = [final_pub_id_col_alias] + depts_col_list)

    # Plotting the words-cloud of the different kinds of keywords
    for kw_type, kw_file in kw_filename_dict.items():

        # Building the keywords dataframe for the keywords type 'kw_type' from the file 'kw_file'
        kw_file_path = parsing_folder_path / Path(kw_file)
        kw_df = pd.read_csv(kw_file_path, sep='\t') 
        kw_df[keywords_col_alias] = kw_df[keywords_col_alias].apply(lambda x: x.replace(' ','_').replace('-','_'))
        kw_df[keywords_col_alias] = kw_df[keywords_col_alias].apply(lambda x: x.replace('_(',';').replace(')','')) 
        kw_df[keywords_col_alias] = kw_df[keywords_col_alias].apply(lambda x: x.lower())
        
        # Creating keywords-analysis data and saving them as xlsx files
        usecols = [final_pub_id_col_alias,      
                   depts_col_list,          
                   parsing_pub_id_col_alias,
                   keywords_col_alias,      
                   weight_col_alias,
                  ]
        _create_kw_analysis_data(year, analysis_df, kw_type, kw_df, usecols, 
                                 kw_analysis_folder_path, verbose = verbose)
        
        # Creating keywords clouds and saving them as png images
        _create_kw_cloud(year, kw_type, kw_analysis_folder_path, usecols, verbose = verbose)
    
    return kw_analysis_folder_path


