__all__ = ['save_homonyms',
           'save_otps',
           'set_saved_homonyms',
           'set_saved_otps',
          ]


def save_homonyms(bibliometer_path, corpus_year):
    """
    """
    # Standard library imports
    from pathlib import Path
    
    # 3rd library imports
    import BiblioParsing as bp
    import pandas as pd 

    # local imports
    import BiblioMeter_FUNCTS.BM_EmployeesGlobals as eg
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg   
    
    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    homonyms_folder_alias    = pg.ARCHI_YEAR["homonymes folder"]
    homonyms_file_base_alias = pg.ARCHI_YEAR["homonymes file name base"]
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_homonyms_file_alias = pg.ARCHI_YEAR["kept homonyms file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"] 
    homonyms_file_alias      = homonyms_file_base_alias + ' ' + corpus_year + ".xlsx"
    
    # Setting useful column name aliases    
    hash_id_col_alias   = pg.COL_HASH['hash_id']
    pub_id_alias        = pg.SUBMIT_COL_RENAME_DIC[bp.COL_NAMES['pub_id']]
    homonyms_col_alias  = pg.SUBMIT_COL_RENAME_DIC[pg.COL_NAMES_BONUS['homonym']]
    matricule_col_alias = pg.SUBMIT_COL_RENAME_DIC[eg.EMPLOYEES_USEFUL_COLS['matricule']]       
    
    # Setting useful paths
    corpus_year_path        = bibliometer_path / Path(corpus_year)    
    bdd_mensuelle_path      = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path       = bdd_mensuelle_path / Path(hash_id_file_alias)
    homonyms_folder_path    = corpus_year_path / Path(homonyms_folder_alias)
    homonyms_file_path      = homonyms_folder_path / Path(homonyms_file_alias)
    history_folder_path     = corpus_year_path / Path(history_folder_alias)
    kept_homonyms_file_path = history_folder_path / Path(kept_homonyms_file_alias)
        
    # Getting the hash_id dataframe
    hash_id_df  = pd.read_excel(hash_id_file_path)
    
    # Getting the dataframe of homonyms to solve 
    pub_df = pd.read_excel(homonyms_file_path)
    
    # Building pub_id and kept matricules df for solved homonymes
    temp_df = pub_df[pub_df[homonyms_col_alias] == pg.HOMONYM_FLAG]
    not_solved_homonyms_pub_id = [pub_id for idx,pub_id in enumerate(temp_df[pub_id_alias]) 
                                  if int(temp_df[temp_df[pub_id_alias] == pub_id] 
                                         [pub_id_alias].value_counts()) > 1]
    homonyms_df = temp_df.copy().reset_index()
    for idx, pub_id in enumerate(temp_df[pub_id_alias]):
        if pub_id in not_solved_homonyms_pub_id:
            homonyms_df = homonyms_df.drop(idx)
    
    kept_matricules_df = homonyms_df[[pub_id_alias, matricule_col_alias]]
    
    # Building hash_id and kept matricules df
    homonyms_history_df = pd.merge(hash_id_df, 
                                   kept_matricules_df, 
                                   how = 'inner',
                                   on = pub_id_alias)
    homonyms_history_df.drop(columns = [pub_id_alias], inplace = True)
    homonyms_history_df = homonyms_history_df.astype(str)
    
    # Concatenating with the dataframe of already saved solved homonyms
    if kept_homonyms_file_path.is_file():
        existing_homonyms_history_df = pd.read_excel(kept_homonyms_file_path)
        homonyms_history_df = pd.concat([existing_homonyms_history_df,homonyms_history_df])
    homonyms_history_df = homonyms_history_df.astype('str')
    homonyms_history_df.drop_duplicates(inplace = True)
    
    # Saving the dataframe concatenated
    homonyms_history_df.to_excel(kept_homonyms_file_path, index = False) 

    message = f"History of homonyms resolution saved"
    return message


def set_saved_homonyms(bibliometer_path, corpus_year, actual_homonym_status):
    """
    """
    # Standard library imports
    from pathlib import Path
    
    # 3rd library imports
    import BiblioParsing as bp
    import pandas as pd 
    
    # local imports
    import BiblioMeter_FUNCTS.BM_EmployeesGlobals as eg
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_ConsolidatePubList import save_shaped_homonyms_file    
    
    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    homonyms_folder_alias    = pg.ARCHI_YEAR["homonymes folder"]
    homonyms_file_base_alias = pg.ARCHI_YEAR["homonymes file name base"]
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_homonyms_file_alias = pg.ARCHI_YEAR["kept homonyms file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"]
    homonyms_file_alias      = homonyms_file_base_alias + ' ' + corpus_year + ".xlsx"
    
    # Setting useful column name aliases
    hash_id_col_alias   = pg.COL_HASH['hash_id']
    pub_id_alias        = pg.SUBMIT_COL_RENAME_DIC[bp.COL_NAMES['pub_id']]
    homonyms_col_alias  = pg.SUBMIT_COL_RENAME_DIC[pg.COL_NAMES_BONUS['homonym']]
    matricule_col_alias = pg.SUBMIT_COL_RENAME_DIC[eg.EMPLOYEES_USEFUL_COLS['matricule']]        
    
    # Setting useful paths
    corpus_year_path        = bibliometer_path / Path(corpus_year)    
    bdd_mensuelle_path      = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path       = bdd_mensuelle_path / Path(hash_id_file_alias)
    homonyms_folder_path    = corpus_year_path / Path(homonyms_folder_alias)
    homonyms_file_path      = homonyms_folder_path / Path(homonyms_file_alias)
    history_folder_path     = corpus_year_path / Path(history_folder_alias)
    kept_homonyms_file_path = history_folder_path / Path(kept_homonyms_file_alias)
 
    if kept_homonyms_file_path.is_file():
        
        # Getting the kept homonyms dataframe
        homonyms_history_df = pd.read_excel(kept_homonyms_file_path) 
        
        # Getting the hash_id dataframe
        hash_id_df  = pd.read_excel(hash_id_file_path)
        
        # Building df of pub_id and matricule to keep related to hash_id       
        pub_id_row_to_keep_df = pd.merge(hash_id_df, 
                                         homonyms_history_df, 
                                         how = 'inner',
                                         on  = hash_id_col_alias,) 
        pub_id_row_to_keep_df = pub_id_row_to_keep_df.astype(str)
        pub_id_row_to_keep_df.drop(columns = [hash_id_col_alias], inplace = True)

        # Getting the resolved homonyms dataframe to be updated    
        homonyms_df = pd.read_excel(homonyms_file_path)  

        # Droping row of unkept matricules in homonyms_df
        homonyms_df_new = homonyms_df.copy()
        for _,row in pub_id_row_to_keep_df.iterrows(): 
            pub_id_to_check   = str(row[pub_id_alias])
            matricule_to_keep = str(row[matricule_col_alias])
            for idx in range(len(homonyms_df)):
                pub_id         = str(homonyms_df.loc[idx,pub_id_alias])
                matricule      = str(homonyms_df.loc[idx,matricule_col_alias])
                homonym_status = str(homonyms_df.loc[idx,homonyms_col_alias])  
                if pub_id == pub_id_to_check and homonym_status == pg.HOMONYM_FLAG:
                    if matricule == matricule_to_keep: 
                        homonyms_df_new.loc[idx,homonyms_col_alias] = None
                    else:
                        homonyms_df_new = homonyms_df_new.drop(idx)
        
        # Setting actual homonyms status
        actual_homonym_status = False
        if pg.HOMONYM_FLAG in homonyms_df_new[homonyms_col_alias].to_list(): actual_homonym_status = True
        
        # Saving updated homonyms_df
        save_shaped_homonyms_file(homonyms_df_new, homonyms_file_path)     
        message = f"Already resolved homonyms used"
    else:
        message = f"No already resolved homonyms available"
    return message, actual_homonym_status 


def save_otps(bibliometer_path, corpus_year):
    """
    """
    # Standard library imports
    from pathlib import Path
    
    # 3rd library imports
    import BiblioParsing as bp
    import pandas as pd 

    # local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
        
    # Setting useful folder and file aliases
    bdd_mensuelle_alias      = pg.ARCHI_YEAR["bdd mensuelle"]
    pub_list_folder_alias    = pg.ARCHI_YEAR["pub list folder"]
    pub_list_file_base_alias = pg.ARCHI_YEAR["pub list file name base"]    
    history_folder_alias     = pg.ARCHI_YEAR["history folder"]
    kept_otps_file_alias     = pg.ARCHI_YEAR["kept otps file name"]
    hash_id_file_alias       = pg.ARCHI_YEAR["hash_id file name"]
    pub_list_file_alias      = pub_list_file_base_alias + f' {corpus_year}.xlsx'
    
    # Setting useful column name aliases       
    hash_id_col_alias  = pg.COL_HASH['hash_id']
    pub_id_alias       = pg.SUBMIT_COL_RENAME_DIC[bp.COL_NAMES['pub_id']]    
    otp_col_alias      = pg.COL_NAMES_BONUS['final OTP']
    otp_list_col_alias = pg.COL_NAMES_BONUS['list OTP']
    
    # Setting useful paths
    corpus_year_path     = bibliometer_path / Path(corpus_year)    
    bdd_mensuelle_path   = corpus_year_path / Path(bdd_mensuelle_alias)
    hash_id_file_path    = bdd_mensuelle_path / Path(hash_id_file_alias)
    pub_list_folder_path = corpus_year_path / Path(pub_list_folder_alias)
    pub_list_file_path   = pub_list_folder_path / Path(pub_list_file_alias)
    history_folder_path  = corpus_year_path / Path(history_folder_alias)
    kept_otps_file_path  = history_folder_path / Path(kept_otps_file_alias)
            
    # Getting the hash_id dataframe
    hash_id_df  = pd.read_excel(hash_id_file_path)
    
    # Getting the dataframe of consolidated pub list with OTPs 
    pub_df = pd.read_excel(pub_list_file_path)
    
    # Building pub_id and otp df
    if otp_col_alias in pub_df.columns: 
        otp_col = otp_col_alias
    else:
        otp_col = otp_list_col_alias    
    kept_otps_df_init = pub_df[[pub_id_alias, otp_col]]
    kept_otps_df_init = kept_otps_df_init.fillna(0)
    kept_otps_df_init = kept_otps_df_init.astype(str)
    kept_otps_df      = kept_otps_df_init.copy()
    sep = ","
    for idx,row in kept_otps_df_init.iterrows():
        if sep in row[otp_col] or row[otp_col] == "0": 
            kept_otps_df = kept_otps_df.drop(idx)
    
    # Building hash_id and kept otp df
    otps_history_df = pd.merge(hash_id_df, 
                               kept_otps_df, 
                               how = 'inner',
                               on = pub_id_alias)
    otps_history_df.drop(columns = [pub_id_alias], inplace = True)
    otps_history_df = otps_history_df.astype(str)
    otps_history_df.rename(columns = {otp_col:otp_col_alias}, inplace = True)
    
    otps_history_df.to_excel(kept_otps_file_path, index = False) 

    message = f"History of kept otps saved"
    return message 


def _re_save_dpt_OTP_file(dpt, otp_set_dpt_df, otp_to_set_dpt_df, 
                          dpt_otp_list, excel_dpt_path, otp_list_col_alias, columns_list):

    ''' Rebuild and store the Excel file under 'excel_dpt_path' for the department 
    labelled 'dpt'. 
    A data validation list is added to each cell of the column
    'OTP_alias' only when the OTP in not already set. 
    The Excel frame is configurated in the same way as in the `mise_en_page` function.        
    '''

    # 3rd party imports
    from openpyxl.worksheet.datavalidation import DataValidation as openpyxl_DataValidation
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    from openpyxl.styles import PatternFill as openpyxl_PatternFill
    from openpyxl.styles import Alignment as openpyxl_Alignment
    from openpyxl.styles import Border as openpyxl_Border
    from openpyxl.styles import Side as openpyxl_Side
    
    # local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_ConsolidatePubList import mise_en_page
    from BiblioMeter_FUNCTS.BM_RenameCols import set_col_attr
    
    # Setting useful column sizes and cell colors
    col_attr, col_set_list = set_col_attr()
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'], fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'], fill_type = "solid")]
    
    # Building validation list of OTP for 'dpt' department
    validation_list = '"'+','.join(dpt_otp_list) + '"' 
    data_val = openpyxl_DataValidation(type = "list", 
                                       formula1 = validation_list, 
                                       showErrorMessage = False)

    # Initializing df_dpt_new with the publications which otp is not yet set
    df_dpt_new = otp_to_set_dpt_df.copy()

    # Adding a column containing OTPs of 'dpt' department
    df_dpt_new[otp_list_col_alias] = validation_list

    # Formatting the EXCEL file
    wb, ws = mise_en_page(df_dpt_new)

    # Setting num of first col and first row in EXCEL files
    excel_first_col_num = 1
    excel_first_row_num = 2

    # Getting the column letter for the OTPs column 
    OTP_alias_df_index = list(df_dpt_new.columns).index(otp_list_col_alias)        
    OTP_alias_excel_index = OTP_alias_df_index + excel_first_col_num 
    OTP_alias_column_letter = openpyxl_get_column_letter(OTP_alias_excel_index)

    # Activating the validation data list in the OTPs column of df_dpt_new 
    df_dpt_len = len(df_dpt_new)
    if df_dpt_len:            
        # Adding a validation data list
        ws.add_data_validation(data_val)
        for df_index_row in range(len(df_dpt_new)):  
            OTP_cell_alias = OTP_alias_column_letter + str(df_index_row + excel_first_row_num)
            data_val.add(ws[OTP_cell_alias])

    # Appending rows of the publications which otp is already set to df_dpt_new
    # and coloring the rows alternatively
    if len(otp_set_dpt_df):
        idx = 1 # continuously incremented index vs row index which is not 
        for _, row in otp_set_dpt_df.iterrows():
            ws.append(row.values.flatten().tolist())                
            # Coloring the row
            idx_row = df_dpt_len + idx 
            idx += 1
            last_row = ws[ws.max_row]
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color

    # Reshaping the alignment and the border of the columns 
    for idx_col, col in enumerate(columns_list):
        if col not in col_set_list: col_attr[col] = col_attr['else']
        column_letter = openpyxl_get_column_letter(idx_col + 1)
        for cell in ws[column_letter]:
            cell.alignment = openpyxl_Alignment(horizontal = col_attr[col][1], vertical = "center")
            cell.border = openpyxl_Border(left = openpyxl_Side(border_style = 'thick', color = 'FFFFFF'),
                                          right = openpyxl_Side(border_style = 'thick', color = 'FFFFFF'))

    # Setting the worksheet label
    ws.title = pg.OTP_SHEET_NAME_BASE + " " +  dpt

    # Saving the workbook
    wb.save(excel_dpt_path)


def set_saved_otps(bibliometer_path, corpus_year):
    """
    
    The OPTs of the choosen department are added in a new column 
    named 'OTP_alias' definined after the global "COL_NAMES_BONUS['list OTP']".
    """
    # Standard library imports
    from pathlib import Path
    
    # 3rd party imports
    import BiblioParsing as bp
    import pandas as pd

    # Local imports
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg 
    
    # Setting useful folder and file aliases
    bdd_mensuelle_alias  = pg.ARCHI_YEAR["bdd mensuelle"]
    otp_folder_alias     = pg.ARCHI_YEAR["OTP folder"]
    otp_file_base_alias  = pg.ARCHI_YEAR["OTP file name base"]    
    history_folder_alias = pg.ARCHI_YEAR["history folder"]
    kept_otps_file_alias = pg.ARCHI_YEAR["kept otps file name"]
    hash_id_file_alias   = pg.ARCHI_YEAR["hash_id file name"] 
    
    # Setting useful key and specific character aliases
    dpt_label_alias = 'dpt_label'
    dpt_otp_alias   = 'dpt_otp'
        
    # Setting useful column names alises       
    hash_id_col_alias  = pg.COL_HASH['hash_id']
    pub_id_alias       = pg.BM_COL_RENAME_DIC[bp.COL_NAMES['pub_id']]    
    otp_list_col_alias = pg.BM_COL_RENAME_DIC[pg.COL_NAMES_BONUS['list OTP']] 
    otp_col_alias      = pg.COL_NAMES_BONUS['final OTP']                      
    
    # Setting useful paths
    corpus_year_path    = bibliometer_path / Path(corpus_year)    
    bdd_mensuelle_path  = corpus_year_path / Path(bdd_mensuelle_alias)    
    hash_id_file_path   = bdd_mensuelle_path / Path(hash_id_file_alias)
    history_folder_path = corpus_year_path / Path(history_folder_alias)
    kept_otps_file_path = history_folder_path / Path(kept_otps_file_alias)
    otp_folder_path     = corpus_year_path / Path(otp_folder_alias)     
    
    if kept_otps_file_path.is_file():
        # Getting the hash_id dataframe
        hash_id_df = pd.read_excel(hash_id_file_path)  

        # Getting the kept otps dataframe
        otp_history_df = pd.read_excel(kept_otps_file_path)

        # Building df of pub_id and otps to set related to hash_id   
        pub_id_otp_to_set_df = pd.merge(hash_id_df, 
                                        otp_history_df, 
                                        how = 'inner',
                                        on  = hash_id_col_alias)
        
        pub_id_otp_to_set_df = pub_id_otp_to_set_df.astype(str)
        pub_id_otp_to_set_df.drop(columns = [hash_id_col_alias], inplace = True)
        pub_id_to_check_list = [str(row[pub_id_alias]) for _,row in pub_id_otp_to_set_df.iterrows()]
        otp_to_set_list      = [str(row[otp_col_alias]) for _,row in pub_id_otp_to_set_df.iterrows()]
        
        # Setting departments list
        dpt_list = list(ig.DPT_ATTRIBUTS_DICT.keys())

        # Setting the known OTPs  
        for dpt in sorted(dpt_list):
            # Setting the full path of the EXCEl file for the 'dpt' department
            otp_file_name_dpt = f'{otp_file_base_alias}_{dpt}.xlsx'
            otp_file_name_dpt_path = otp_folder_path / Path(otp_file_name_dpt)

            # Getting the pub list for departement dpt
            dpt_pub_df = pd.read_excel(otp_file_name_dpt_path)
            
            # Setting the pub-id list for department dpt
            dept_pub_id_list = dpt_pub_df[pub_id_alias].to_list()

            # Setting columns list
            col_list = list(dpt_pub_df.columns)

            # Building the otp_set_dpt_pub_df dataframe of publication with otp set
            # the otp_to_set_dpt_pub_df dataframe of publication with otp still to be defined
            otp_set_dpt_pub_df = pd.DataFrame(columns=col_list)
            otp_to_set_dpt_pub_df = dpt_pub_df.copy()
            otp_to_set_dpt_pub_df.drop(columns = [otp_list_col_alias], inplace = True)
            
            for otp_idx in range(len(pub_id_to_check_list)):
                pub_id_to_check = pub_id_to_check_list[otp_idx]
                otp_to_set      = otp_to_set_list[otp_idx]                
                
                if pub_id_to_check in dept_pub_id_list:
                    pub_id_idx = [i for i,e in enumerate(dept_pub_id_list) if e == pub_id_to_check][0]
                    dpt_pub_df.loc[pub_id_idx,otp_list_col_alias] = otp_to_set
                    otp_set_dpt_pub_df = pd.concat([otp_set_dpt_pub_df,
                                                    dpt_pub_df[dpt_pub_df[pub_id_alias] == pub_id_to_check]])
                    otp_to_set_dpt_pub_df.drop(index = pub_id_idx, inplace = True)
            
            # Setting the list of OTPs for the 'dpt' department
            dpt_otp_list = ig.DPT_ATTRIBUTS_DICT[dpt][dpt_otp_alias]

            # Resetting validation list for OTPs when not already set and saving the file
            _re_save_dpt_OTP_file(dpt, otp_set_dpt_pub_df, otp_to_set_dpt_pub_df, 
                                  dpt_otp_list, otp_file_name_dpt_path, otp_list_col_alias, col_list)
        
        message = f"Already set OTPS used"
    else:
        message = f"No already set OTPS available"
    return message


