__all__ = ['add_if',
           'add_OTP', 
           'concatenate_pub_lists',
           'consolidate_pub_list',
           'get_if_db',
           'mise_en_page',
           'save_shaped_homonyms_file',
           'solving_homonyms',
           'split_pub_list_by_doc_type',
          ]

def mise_en_page(institute, org_tup, df,                        
                 wb = None, if_database = None):    
    ''' 
    When the workbook wb is not None, this is applied 
    to the active worksheet of the passed workbook. 
    If the workbook wb is None, then the worbook is created.    
    '''
    
    # 3rd party imports
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows as openpyxl_dataframe_to_rows
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    from openpyxl.styles import Font as openpyxl_Font  
    from openpyxl.styles import PatternFill as openpyxl_PatternFill 
    from openpyxl.styles import Alignment as openpyxl_Alignment
    from openpyxl.styles import Border as openpyxl_Border
    from openpyxl.styles import Side as openpyxl_Side
        
    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_RenameCols import set_col_attr
       
    # Setting useful column sizes
    col_attr, col_set_list = set_col_attr(institute, org_tup)
    columns_list = list(df.columns)
    for col in columns_list:
        if col not in col_set_list: col_attr[col] = col_attr['else']
        
     # Setting list of cell colors   
    cell_colors = [openpyxl_PatternFill(fgColor = pg.ROW_COLORS['odd'], fill_type = "solid"),
                   openpyxl_PatternFill(fgColor = pg.ROW_COLORS['even'], fill_type = "solid")]
    
    # Initialize wb as a workbook and ws its active worksheet
    if not wb : wb = Workbook()
    ws = wb.active
    ws_rows = openpyxl_dataframe_to_rows(df, index=False, header=True)
    
    # Coloring alternatly rows in ws using list of cell colors cell_colors
    for idx_row, row in enumerate(ws_rows):       
        ws.append(row)        
        last_row = ws[ws.max_row]            
        if idx_row >= 1:
            cell_color = cell_colors[idx_row%2]
            for cell in last_row:
                cell.fill = cell_color 
    
    # Setting cell alignement and border using dict of column attributes col_attr
    if if_database:
        align_list = ["left", "center","center","center"]
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            for cell in ws[column_letter]:
                cell.alignment = openpyxl_Alignment(horizontal=align_list[idx_col], vertical="center")
                cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick', color='FFFFFF'),
                                              right=openpyxl_Side(border_style='thick', color='FFFFFF'))  
    else: 
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            for cell in ws[column_letter]:
                cell.alignment = openpyxl_Alignment(horizontal=col_attr[col][1], vertical="center")
                cell.border = openpyxl_Border(left=openpyxl_Side(border_style='thick', color='FFFFFF'),
                                              right=openpyxl_Side(border_style='thick', color='FFFFFF'))                    
    
    # Setting the format of the columns heading
    cells_list = ws['A'] + ws[1]
    if if_database: cells_list = ws[1]
    for cell in cells_list:
        cell.font = openpyxl_Font(bold=True)
        cell.alignment = openpyxl_Alignment(wrap_text=True, horizontal="center", vertical="center")
    
    # Setting de columns width using dict of column attributes col_attr if if_database = None
    if if_database:
        col_width_list = [60,15,15,15]
        for idx_col, col in enumerate(columns_list):
            column_letter = openpyxl_get_column_letter(idx_col + 1)
            ws.column_dimensions[column_letter].width = col_width_list[idx_col]
    else:
        for idx_col, col in enumerate(columns_list):
            if idx_col >= 1:
                column_letter = openpyxl_get_column_letter(idx_col + 1)
                try:
                    ws.column_dimensions[column_letter].width = col_attr[col][0]
                except:
                    ws.column_dimensions[column_letter].width = 20
    
    
    # Setting height of first row
    first_row_num = 1
    if if_database:
        ws.row_dimensions[first_row_num].height = 20 
    else:
        ws.row_dimensions[first_row_num].height = 50

    return wb, ws


def save_shaped_homonyms_file(df_homonyms, out_path):
    """
    
    """
    # 3rd party imports
    from openpyxl import Workbook as openpyxl_Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows as openpyxl_dataframe_to_rows
    from openpyxl.styles import PatternFill as openpyxl_PatternFill
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    
    # Setting useful column names 
    col_homonyms = list(df_homonyms.columns)
    
    # Useful aliases of renamed columns names 
    name_alias      = col_homonyms[12] 
    firstname_alias = col_homonyms[13] 
    homonym_alias   = col_homonyms[18] 
    
    wb = openpyxl_Workbook()
    ws = wb.active    
    ws.title = 'Consolidation Homonymes'    
    yellow_ft = openpyxl_PatternFill(fgColor = pg.ROW_COLORS['highlight'], fill_type = "solid")

    for indice, r in enumerate(openpyxl_dataframe_to_rows(df_homonyms, index=False, header=True)):
        ws.append(r)
        last_row = ws[ws.max_row]
        if r[col_homonyms.index(homonym_alias)] == pg.HOMONYM_FLAG and indice > 0:
            cell      = last_row[col_homonyms.index(name_alias)] 
            cell.fill = yellow_ft
            cell      = last_row[col_homonyms.index(firstname_alias)] 
            cell.fill = yellow_ft
            
    wb.save(out_path)

    
def solving_homonyms(institute, org_tup, in_path, out_path): 
    """
    Uses the local function 'save_shaped_homonyms_file' 
    to shape then save the homonyms df.
    """
    # 3rd party imports
    import pandas as pd

    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_RenameCols import set_homonym_col_names
    
    # Setting useful column names 
    col_homonyms = set_homonym_col_names(institute, org_tup)
    homonym_col_alias = col_homonyms[18]
    
    # Reading the submit file #
    df_submit = pd.read_excel(in_path)
    
    # Getting rid of the columns we don't want #
    df_homonyms = df_submit[col_homonyms].copy()    
    
    # Setting homonyms status
    homonyms_status = False
    if pg.HOMONYM_FLAG in df_homonyms[homonym_col_alias].to_list(): homonyms_status = True
    
    # Saving shaped df_homonyms
    save_shaped_homonyms_file(df_homonyms, out_path)
    
    end_message = f"File for solving homonymies saved in folder: \n  '{out_path}'"
    return (end_message, homonyms_status)


def _add_authors_name_list(institute, org_tup, in_path, out_path): 
    ''' The function ` _add_authors_name_list` adds two columns to the dataframe get 
    from the Excel file pointed by 'in_path'.
    The columns contain respectively the full name of each author as "NAME, Firstname" 
    and the institute co-authors list with their job type as 
    "NAME1, Firstame1 (job type); NAME2, Firstame2 (job type);...".
    
    Args:
        in_path (path): Fullpath of the working excel file. 
        out_path (path): Fullpath of the processed dataframe as an Excel file 
                         saved after going through its treatment.
    
    Returns:
        (str): end message recalling out_path.
    
    Notes:
        The global 'COL_NAMES' is imported from 'BiblioSpecificGlobals' module 
        of 'BiblioParsing' package.
        The global 'EMPLOYEES_USEFUL_COLS' is imported from 'BM_EmployeesGlobals' 
        module of 'BiblioMeter_FUNCTS' package.  
        The global 'COL_NAMES_BONUS' is imported from 'BM_PubGlobals' 
        module of 'BiblioMeter_FUNCTS' package.
    '''
    # 3rd party imports
    import pandas as pd
    import BiblioParsing as bp
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_EmployeesGlobals as eg
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_RenameCols import build_col_conversion_dic
    
    # Internal functions
    def _get_dpt_key(dpt_raw):
        for key, values in dpt_label_dict.items():
            if dpt_raw in values: return key 
        
    # Setting institute parameters
    col_names_dpt  = org_tup[0]
    dpt_label_dict = org_tup[1]
    inst_col_list  = org_tup[4]  
    
    # Setting useful column names
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    bm_col_rename_dic = col_rename_tup[2]       
    
    # Setting useful aliases
    pub_id_alias          = bm_col_rename_dic[bp.COL_NAMES['pub_id']]
    idx_authors_alias     = bm_col_rename_dic[bp.COL_NAMES['authors'][1]]
    nom_alias             = bm_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['name']]
    prenom_alias          = bm_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['first_name']]
    full_name_alias       = bm_col_rename_dic[pg.COL_NAMES_BONUS['nom prénom'] + institute]
    author_type_col_alias = bm_col_rename_dic[pg.COL_NAMES_BONUS['author_type']]
    full_name_list_alias  = bm_col_rename_dic[pg.COL_NAMES_BONUS['nom prénom liste'] + institute]
    dept_col_alias        = bm_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['dpt']]
    
    # Reading the excel file
    df_in = pd.read_excel(in_path)
    
    # Adding the column 'Nom Prénom' that will be used to create the authors fullname list
    df_in[prenom_alias]    = df_in[prenom_alias].apply(lambda x: x.capitalize())
    df_in[full_name_alias] = df_in[nom_alias] + ', ' + df_in[prenom_alias]
    
    df_out = pd.DataFrame()
    for pub_id, pub_id_df in df_in.groupby(pub_id_alias):

        authors_tup_list = sorted(list(set(zip(pub_id_df[idx_authors_alias], 
                                               pub_id_df[full_name_alias], 
                                               pub_id_df[author_type_col_alias],
                                               pub_id_df[dept_col_alias]))))

        authors_str_list = [f'{x[1]} ({x[2]},{_get_dpt_key(x[3])})' for x in  authors_tup_list]
        authors_full_str ="; ".join(authors_str_list)
        pub_id_df[full_name_list_alias] = authors_full_str
        df_out = pd.concat([df_out, pub_id_df])

    # Saving 'df_out' in an excel file 'out_path'
    df_out.to_excel(out_path, index = False)
    
    end_message = f"Column with co-authors list is added to the file: \n  '{out_path}'"
    return end_message


def _save_dpt_OTP_file(institute, org_tup, dpt, df_dpt, dpt_otp_list,
                       OTP_alias, excel_dpt_path, col_otp):

    ''' Create and store an Excel file under 'excel_dpt_path' for the department labelled 'dpt'.
    The OPTs of the choosen department are added in a new column named 'OTP_alias'. 
    A list data validation rules is added to each celles of the column
    'OTP_alias'. The data frame column are renamed using 'col_otp'. The Excel frame is
    configurated by the `mise_en_page` function.

    '''
    # 3rd party imports
    from openpyxl.worksheet.datavalidation import DataValidation as openpyxl_DataValidation
    from openpyxl.utils import get_column_letter as openpyxl_get_column_letter
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_ConsolidatePubList import mise_en_page
    
    # Building validation list of OTP for 'dpt' department
    validation_list = '"'+','.join(dpt_otp_list) + '"' 
    data_val = openpyxl_DataValidation(type = "list", 
                                       formula1 = validation_list, 
                                       showErrorMessage = False)

    # Adding a column containing OTPs of 'dpt' department
    df_dpt[OTP_alias] = validation_list

    # Renaming the columns 
    df_dpt = df_dpt.reindex(columns = col_otp)
    # to replace by :
    # COL_OTP_DICT {col_old_name : col_new_name}
    # df_dpt = df_dpt.rename(columns = COL_OTP_DICT)

    # Formatting the EXCEL file
    wb, ws = mise_en_page(institute, org_tup, df_dpt)
    ws.title = pg.OTP_SHEET_NAME_BASE + " " +  dpt 

    # Setting num of first col and first row in EXCEL files
    excel_first_col_num = 1
    excel_first_row_num = 2

    # Getting the column letter for the OTPs column 
    OTP_alias_df_index = list(df_dpt.columns).index(OTP_alias)        
    OTP_alias_excel_index = OTP_alias_df_index + excel_first_col_num 
    OTP_alias_column_letter = openpyxl_get_column_letter(OTP_alias_excel_index)

    # Activating the validation data list in all cells of the OTPs column
    if len(df_dpt):
        # Adding a validation data list
        ws.add_data_validation(data_val)
        for df_index_row in range(len(df_dpt)):  
            OTP_cell_alias = OTP_alias_column_letter + str(df_index_row + excel_first_row_num)
            data_val.add(ws[OTP_cell_alias])

    wb.save(excel_dpt_path)


def add_OTP(institute, org_tup, in_path, out_path, out_file_base):     
    '''    
    Args:
        in_path (path): fullpath of the working excel file. 
        out_path (path): fullpath of the saved prosseced file
    
    Returns:
        None.
    
    Notes:
        The global 'COL_NAMES' is imported from the module 'BiblioSpecificGlobals'  
        of the package 'BiblioParsing'.
        The functions `_add_authors_name_list` and `mise_en_page` are imported 
        from the module 'BiblioMeterFonctions' of the package 'BiblioMeter_FUNCTS'.
        The global 'EMPLOYEES_USEFUL_COLS' is imported from the module 'BM_EmployeesGlobals' 
        of the package 'BiblioMeter_FUNCTS'. 
        The globals 'COL_NAMES_BONUS' and 'DPT_ATTRIBUTS_DICT' are imported 
        from the module 'BM_PubGlobals' of the package 'BiblioMeter_FUNCTS'. 
    '''
    
    # Standard library imports
    from pathlib import Path

    # 3rd party imports
    import pandas as pd
    import BiblioParsing as bp
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_EmployeesGlobals as eg
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_ConsolidatePubList import _add_authors_name_list
    from BiblioMeter_FUNCTS.BM_RenameCols import set_otp_col_names
    from BiblioMeter_FUNCTS.BM_RenameCols import build_col_conversion_dic

    # Setting institute parameters
    col_names_dpt      = org_tup[0]
    dpt_attributs_dict = org_tup[2]

    # Setting useful column names 
    col_otp = set_otp_col_names(institute, org_tup) 
    col_rename_tup = build_col_conversion_dic(institute, org_tup)
    bm_col_rename_dic = col_rename_tup[2]
    
    # Setting useful aliases
    pub_id_alias     = bm_col_rename_dic[bp.COL_NAMES['pub_id']]           # Pub_id
    idx_author_alias = bm_col_rename_dic[bp.COL_NAMES['authors'][1]]       # Idx_author
    dpt_alias        = bm_col_rename_dic[eg.EMPLOYEES_USEFUL_COLS['dpt']]  # Dpt/DOB (lib court)
    OTP_alias        = bm_col_rename_dic[pg.COL_NAMES_BONUS['list OTP']]   # Choix de l'OTP
    dpt_label_alias  = ig.DPT_LABEL_KEY
    dpt_otp_alias    = ig.DPT_OTP_KEY 
        
    # Adding a column with a list of the authors in the file where homonymies 
    # have been solved and pointed by in_path
    end_message = _add_authors_name_list(institute, org_tup, in_path, in_path)
    print('\n ',end_message)
    
    solved_homonymies_df = pd.read_excel(in_path)
    solved_homonymies_df.fillna('', inplace = True)
    
    dpt_list = list(dpt_attributs_dict.keys())
     
    # For each department adding a column containing 1 or 0 
    # depending if the author belongs or not to the department
    for dpt in dpt_list:
        dpt_label_list = dpt_attributs_dict[dpt][dpt_label_alias]
        solved_homonymies_df[dpt] = solved_homonymies_df[dpt_alias].apply(lambda x: 1 
                                                                          if x in dpt_label_list 
                                                                          else 0)    
    
    # Building 'df_out' out of 'solved_homonymies_df' with a row per pub_id 
    # 1 or 0 is assigned to each department column depending 
    # on if at least one co-author is a member of this department,
    # the detailed information is related to the first author only
    df_out = pd.DataFrame()
    for pub_id, dg in solved_homonymies_df.groupby(pub_id_alias):
        dg = dg.sort_values(by = [idx_author_alias])
        for dpt in dpt_list:
            x = dg[dpt].any().astype(int)
            dg[dpt] = x      
        df_out = pd.concat([df_out, dg.iloc[:1]]) 
    
    # Configuring an Excel file per department with the list of OTPs
    for dpt in sorted(dpt_list):
        # Setting df_dpt with only pub_ids for which the first author
        # is from the 'dpt' department
        filtre_dpt = False
        for dpt_value in dpt_attributs_dict[dpt][dpt_label_alias]:
            filtre_dpt = filtre_dpt | (df_out[dpt_alias] == dpt_value)
        df_dpt = df_out[filtre_dpt].copy()
        
        # Setting the list of OTPs for the 'dpt' department
        dpt_otp_list = dpt_attributs_dict[dpt][dpt_otp_alias]
               
        # Setting the full path of the EXCEl file for the 'dpt' department
        OTP_file_name_dpt = f'{out_file_base}_{dpt}.xlsx'
        excel_dpt_path    = out_path / Path(OTP_file_name_dpt)
        
        # Adding a column with validation list for OTPs and saving the file
        _save_dpt_OTP_file(institute, org_tup, dpt, df_dpt, dpt_otp_list, 
                           OTP_alias, excel_dpt_path, col_otp)

    end_message  = f"Files for setting publication OTPs per department "
    end_message += f"saved in folder: \n  '{out_path}'"
    return end_message
    
def _create_if_column(issn_column, if_dict, if_empty_word):
    ''' The function `_create_if_column` builds a dataframe column 'if_column'
    using the column 'issn_column' of this dataframe and the dict 'if_dict' 
    that make the link between ISSNs ('if_dict' keys) and IFs ('if_dict' values). 
    The 'nan' values in the column 'if_column' are replaced by 'empty_word'.

    Args:
        issn_column (pandas serie): The column of the dataframe of interest 
                                    that contains the ISSNs values.
        if_dict (dict): The dict which keys are ISSNs and values are IFs.
        if_empty_word (str): The word that will replace nan values in column the returned column.

    Returns:
        (pandas serie): The column of the dataframe of interest 
                        that contains the IFs values. 
    '''
    if_column = issn_column.map(if_dict)
    if_column = if_column.fillna(if_empty_word)
    return if_column


def _build_inst_issn_df(if_db_df, use_col_list):
    """
    
    """

    # 3rd party import
    import pandas as pd
    import BiblioParsing as bp

    # Setting useful aliases
    journal_col_alias = use_col_list[0]
    issn_col_alias    = use_col_list[1]
    eissn_col_alias   = use_col_list[2]
    unknown_alias     = bp.UNKNOWN    
    
    years_list = list(if_db_df.keys())

    init_inst_issn_df = pd.DataFrame(columns = use_col_list)

    for year in years_list:
        year_sub_df = if_db_df[year][use_col_list].copy()
        init_inst_issn_df = pd.concat([init_inst_issn_df, year_sub_df])
    init_inst_issn_df[journal_col_alias] = init_inst_issn_df.apply(lambda row: row[journal_col_alias].upper(), axis=1)

    inst_issn_df = pd.DataFrame() 
    for _ , dg in init_inst_issn_df.groupby(journal_col_alias):

        issn_list = list(set(dg[issn_col_alias].to_list()) - {unknown_alias})
        if not issn_list: issn_list = [unknown_alias]
        dg[issn_col_alias] = issn_list[0]

        eissn_list = list(set(dg[eissn_col_alias].to_list()) - {unknown_alias})
        if not eissn_list: eissn_list = [unknown_alias]
        dg[eissn_col_alias] = eissn_list[0]

        inst_issn_df = pd.concat([inst_issn_df,dg.iloc[:1]])

    inst_issn_df.sort_values(by=[journal_col_alias], inplace = True)    
    inst_issn_df.drop_duplicates(inplace = True)    
    
    return inst_issn_df


def get_if_db(institute, org_tup, bibliometer_path):
    
    # Standard imports
    from pathlib import Path
    
    # 3rd party imports
    import pandas as pd 
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg 
        
    ## Setting institute parameters    
    if_db_status = org_tup[5]
    
    # Setting useful aliases
    if_root_folder_alias   = pg.ARCHI_IF["root"]
    if_filename_alias      = pg.ARCHI_IF["all IF"]    
    inst_if_filename_alias = institute + pg.ARCHI_IF["institute_if_all_years"]
    
    if if_db_status: if_filename_alias = inst_if_filename_alias
    
    # Setting useful paths
    if_root_folder_path = bibliometer_path / Path(if_root_folder_alias)
    if_path             = if_root_folder_path / Path(if_filename_alias)

    # Getting the df of the IFs database
    if_df = pd.read_excel(if_path, sheet_name = None)
    
    # Setting list of years for which IF are available
    if_available_years_list  = list(if_df.keys())
    
    if_most_recent_year = if_available_years_list[-1]
    
    return if_df, if_available_years_list, if_most_recent_year


def add_if(institute, org_tup, bibliometer_path, in_file_path, out_file_path, 
           missing_if_path, missing_issn_path, corpus_year): 

    '''The function `add_if` adds two new columns containing impact factors
    to the corpus dataframe 'corpus_df' got from a file which full path is 'in_file_path'. 
    The two columns are named through 'corpus_year_if_col_name' and 'most_recent_year_if_col_name'.
    The impact factors are got using `get_if_db` function that returns in particular the dataframe 'if_df'.
    The column 'corpus_year_if_col_name' is filled with the IFs values 
    of the corpus year 'corpus_year' if available in the dataframe 'if_df', 
    else the values are set to 'not_available_if_alias'.
    The column 'most_recent_year_if_col_name' is filled with the IFs values 
    of the most recent year available in the dataframe 'if_df'.
    In these columns, the NaN values of IFs are replaced by 'unknown_if_fill_alias'.
    
    Args:
        in_file_path (path): The full path to get the corpus dataframe 'corpus_df'. 
        out_file_path (path): The full path to save the new corpus dataframe with the two columns.
        missing_if_path (path): The full path to save the missing IFs information
        missing_issn_path (path): The full path to save the missing ISSNs information.
        corpus_year (int): The year of the corpus to be appended with the two new IF columns.
        
    Returns:
        (str): Message indicating which file has been affected and how. 
    
    Notes:
        The globals 'COL_NAMES_BONUS', 'FILL_EMPTY_KEY_WORD' and 'NOT_AVAILABLE_IF' 
        are imported from the module 'BM_PubGlobals' 
        of the package 'BiblioMeter_FUNCTS'.    
    '''
    
    # 3rd party imports
    import pandas as pd
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    from BiblioMeter_FUNCTS.BM_RenameCols import set_if_col_names
    
    # Internal functions
    
    def _fullfill_issn(corpus_df_bis):
        for corpus_idx, corpus_row in corpus_df_bis.iterrows():
            if corpus_row[issn_col_alias] == unknown_alias:
                corpus_journal = corpus_row[journal_col_alias].upper()
                for inst_idx, inst_row in inst_issn_df.iterrows():
                    inst_journal = inst_row[journal_col_alias].upper()
                    if corpus_journal == inst_journal:
                        if inst_row[issn_col_alias] != unknown_alias:
                            corpus_df_bis.loc[corpus_idx,issn_col_alias] = inst_row[issn_col_alias]
                        elif inst_row[eissn_col_alias] != unknown_alias:
                            corpus_df_bis.loc[corpus_idx,issn_col_alias] = inst_row[eissn_col_alias]
                        else:
                            pass 
        return corpus_df_bis
    
    def _build_if_dict(if_year, issn_col, eissn_col, if_col):
        issn_if_dict  = dict(zip(if_df[if_year][issn_col], 
                                 if_df[if_year][if_col]))
        if unknown_alias in issn_if_dict.keys(): del issn_if_dict[unknown_alias]
        
        eissn_if_dict = {}
        if eissn_col_alias in list(if_df[if_year].columns):
            eissn_if_dict = dict(zip(if_df[if_year][eissn_col], 
                                     if_df[if_year][if_col]))
            if unknown_alias in eissn_if_dict.keys(): del eissn_if_dict[unknown_alias]

        if_dict = {**issn_if_dict, **eissn_if_dict}
        return if_dict
    
    def _get_id(journal, col):
        id_lower_df = inst_issn_df[inst_issn_df[journal_col_alias]==journal.lower()][col]
        id_lower = unknown_alias
        if not id_lower_df.empty: id_lower = id_lower_df.to_list()[0]   
        id_upper_df = inst_issn_df[inst_issn_df[journal_col_alias]==journal.upper()][col]
        id_upper = unknown_alias
        if not id_upper_df.empty: id_upper = id_upper_df.to_list()[0]
        journal_id = list(set([id_lower,id_upper]) - set([unknown_alias]))[0]
        return journal_id
    
    def _format_and_save(results_df, df_full_path, add_cols):
        # Formatting 'results_df'
        results_df.rename(columns  = {year_col_alias          : final_year_col,
                                      corpus_year_if_col_name : year_db_if_col_name,}, 
                          inplace = True)

        if add_cols:
            results_df.rename(columns  = {issn_col_alias : corpus_issn_col_alias,}, 
                              inplace = True)

            results_df[issn_col_alias]  = unknown_alias
            results_df[eissn_col_alias] = unknown_alias 
            results_df = results_df[final_order_col_list]
        else:
            if results_df.empty: results_df[eissn_col_alias] = unknown_alias
            results_df = results_df[final_order_col_list[:-1]]
        sorted_results_df = results_df.sort_values(by=[journal_col_alias])

        # Saving 'results_df' as EXCEL file at full path 'df_full_path'
        wb, _ = mise_en_page(institute, org_tup, sorted_results_df)
        wb.save(df_full_path)
    
    # Setting useful column names
    col_final_list = set_final_col_names(institute, org_tup)
    col_base_if, col_maj_if = set_if_col_names(institute, org_tup)

    # Setting useful column aliases
    pub_id_col_alias          = col_final_list[0]   
    year_col_alias            = col_final_list[1] 
    journal_col_alias         = col_final_list[6]
    doctype_col_alias         = col_final_list[7]
    issn_col_alias            = col_final_list[10]
    otp_col_alias             = col_final_list[16]
    
    current_if_col_name_alias = col_maj_if[17]
    corpus_year_if_col_name   = col_maj_if[18]
    
    corpus_issn_col_alias     = pg.COL_NAMES_BONUS["database ISSN"]
    database_if_col_alias     = pg.COL_NAMES_BONUS['IF clarivate']
    eissn_col_alias           = pg.COL_NAMES_BONUS['e-ISSN']
    otp_col_new_alias         = pg.COL_NAMES_BONUS['final OTP']
    pub_id_nb_col_alias       = pg.COL_NAMES_BONUS['pub number']      
  
    # Setting globals aliases
    doc_type_dict_alias        = pg.DOC_TYPE_DICT
    not_available_if_alias     = pg.NOT_AVAILABLE_IF
    unknown_if_fill_alias      = pg.FILL_EMPTY_KEY_WORD
    unknown_alias              = pg.FILL_EMPTY_KEY_WORD
            
    # Setting institute parameters
    if_db_status            = org_tup[5] 
    no_if_doctype_keys_list = org_tup[6] 
    
    no_if_doctype = sum([doc_type_dict_alias[x] for x in no_if_doctype_keys_list] , [])
    doctype_to_drop_list_alias = [x.upper() for x in no_if_doctype]
    
    # Getting the df of the IFs database
    if_df, if_available_years_list, if_most_recent_year = get_if_db(institute, org_tup, bibliometer_path)    
    
    # Taking care all IF column names in if_df are database_if_col_alias
    if if_db_status: 
        for year in if_available_years_list: 
            if_df[year].rename(columns = {database_if_col_alias + " " + year : database_if_col_alias},
                               inplace = True) 
            
    # Replacing NAN in if_df
    values_dict = {issn_col_alias       : unknown_alias,
                   eissn_col_alias      : unknown_alias,
                   database_if_col_alias: not_available_if_alias,
                  }
    for year in if_available_years_list: if_df[year].fillna(value = values_dict, inplace = True)
    
    # Building the IF dict keyed by issn or e-issn of journals for the most recent year
    most_recent_year_if_dict = _build_if_dict(if_most_recent_year, issn_col_alias, eissn_col_alias, database_if_col_alias)
    
    # Setting local column names    
    most_recent_year_if_col_name = current_if_col_name_alias + ', ' + if_most_recent_year 
    year_db_if_col_name          = database_if_col_alias + ' ' + corpus_year
    final_year_col               = year_col_alias[0:5]    
    journal_upper_col            = journal_col_alias + '_Upper'
    
    # Setting the ordered final columns
    final_order_col_list = [final_year_col, journal_col_alias, 
                            issn_col_alias, eissn_col_alias,
                            most_recent_year_if_col_name,
                            year_db_if_col_name,
                            pub_id_nb_col_alias,
                            corpus_issn_col_alias, 
                           ]
    
    # Getting the df where to add IFs
    corpus_df = pd.read_excel(in_file_path)
    
    # Recasting column names
    new_col_base_if = list(map(lambda x: x.replace(otp_col_alias, otp_col_new_alias), col_base_if))
    if otp_col_alias in corpus_df.columns: 
        corpus_df.rename(columns = {otp_col_alias : otp_col_new_alias}, inplace = True)
        
    # Setting type of values in 'year_col_alias' as string                
    corpus_df = corpus_df.astype({year_col_alias : str})
        
    # Initializing 'corpus_df_bis' as copy of 'corpus_df'
    corpus_df_bis = corpus_df[new_col_base_if].copy()   
    
    # Getting the df of ISSN and eISSN database of the institut
    use_col_list = [journal_col_alias, issn_col_alias, eissn_col_alias]
    inst_issn_df = _build_inst_issn_df(if_df, use_col_list)
    
    # Filling unknown ISSN in 'corpus_df_bis' using 'inst_issn_df' 
    # through internal function _fullfill_issn
    corpus_df_bis = _fullfill_issn(corpus_df_bis)    
    
    # Adding 'most_recent_year_if_col_name' column to 'corpus_df_bis' 
    # with values defined by internal function '_create_if_column'
    corpus_df_bis[most_recent_year_if_col_name] = _create_if_column(corpus_df_bis[issn_col_alias],
                                                                    most_recent_year_if_dict,
                                                                    unknown_if_fill_alias)
    
    # Adding 'corpus_year_if_col_name' column to 'corpus_df_bis' 
    if corpus_year in if_available_years_list:
        # with values defined by internal function '_create_if_column'
        # Building the IF dict keyed by issn or e-issn of journals for the corpus year
        current_year_if_dict = _build_if_dict(corpus_year, issn_col_alias, eissn_col_alias, database_if_col_alias)
        corpus_df_bis[corpus_year_if_col_name] = _create_if_column(corpus_df_bis[issn_col_alias],
                                                                   current_year_if_dict,
                                                                   unknown_if_fill_alias)
    else:
        # with 'not_available_if_alias' value
        corpus_df_bis[corpus_year_if_col_name] = not_available_if_alias
    
    # Formatting and saving 'corpus_df_bis' as EXCEL file at full path 'out_file_path'
    corpus_df_bis.sort_values(by = [pub_id_col_alias], inplace = True)  
    wb, _ = mise_en_page(institute, org_tup, corpus_df_bis)
    wb.save(out_file_path)
    
    # Building 'year_pub_if_df' with subset of 'corpus_df_bis' columns
    subsetcols = [pub_id_col_alias,
                  year_col_alias,
                  journal_col_alias,
                  doctype_col_alias,
                  issn_col_alias,
                  most_recent_year_if_col_name,
                  corpus_year_if_col_name,]
    year_pub_if_df = corpus_df_bis[subsetcols].copy()   
    
    # Building 'year_article_if_df' by keeping only rows which doc type has usually an IF
    # then droping the doc type column
    year_article_if_df = pd.DataFrame(columns = year_pub_if_df.columns)
    for doc_type, doc_type_df in year_pub_if_df.groupby(doctype_col_alias):
        if doc_type.upper() not in doctype_to_drop_list_alias:
            year_article_if_df = pd.concat([year_article_if_df, doc_type_df])
    year_article_if_df.drop(doctype_col_alias, axis = 1, inplace = True)
        
    # Building 'year_if_df' by keeping one row for each issn adding a column with number of related articles
    # then droping "Pub_id" column   
    year_if_df = pd.DataFrame(columns = year_article_if_df.columns.to_list() [1:] + [pub_id_nb_col_alias])    
    for issn, issn_df in year_article_if_df.groupby(issn_col_alias):
        pub_id_nb = len(issn_df)
        issn_df[pub_id_nb_col_alias] = pub_id_nb   
        issn_df.drop(pub_id_col_alias, axis = 1, inplace = True)    
        issn_df[journal_upper_col] = issn_df[journal_col_alias].astype(str).str.upper()
        issn_df.drop_duplicates(subset=[journal_upper_col], keep='first', inplace = True)
        issn_df.drop([journal_upper_col], axis = 1, inplace = True)
        year_if_df = pd.concat([year_if_df, issn_df])
    
    # Removing from 'year_if_df' the rows which ISSN value is not in IF database and keeping them in 'year_missing_issn_df'
    year_missing_issn_df = pd.DataFrame(columns = year_if_df.columns)
    year_missing_if_df   = pd.DataFrame(columns = year_if_df.columns)
    inst_issn_list  = inst_issn_df[issn_col_alias].to_list()
    inst_eissn_list = inst_issn_df[eissn_col_alias].to_list()
    for _, row in year_if_df.iterrows():     
        row_issn = row[issn_col_alias]
        row_most_recent_year_if = row[most_recent_year_if_col_name]
        row_corpus_year_if      = row[corpus_year_if_col_name]
        if row_issn not in inst_issn_list and row_issn not in inst_eissn_list:
            year_missing_issn_df = pd.concat([year_missing_issn_df, row.to_frame().T])
        elif unknown_alias in [row_most_recent_year_if, row_corpus_year_if]:
            row_journal = row[journal_col_alias]
            row[issn_col_alias]  = _get_id(row_journal, issn_col_alias)
            row[eissn_col_alias] = _get_id(row_journal, eissn_col_alias)
            year_missing_if_df   = pd.concat([year_missing_if_df, row.to_frame().T])
    
    if_database_complete = True
    if not year_missing_issn_df.empty or not year_missing_if_df.empty: 
        if_database_complete = False
    
    _format_and_save(year_missing_issn_df, missing_issn_path, add_cols = True)
    _format_and_save(year_missing_if_df, missing_if_path, add_cols = False)
    
    end_message = f"IFs added for year {year} in file : \n  '{out_file_path}'"
    return end_message, if_database_complete


def split_pub_list_by_doc_type(institute, org_tup, bibliometer_path, corpus_year): 
    """
    
    """
    # Standard library imports
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg 
    from BiblioMeter_FUNCTS.BM_ConsolidatePubList import mise_en_page
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names

    pub_list_path_alias      = pg.ARCHI_YEAR["pub list folder"]
    pub_list_file_base_alias = pg.ARCHI_YEAR["pub list file name base"]

    year_pub_list_file_alias = pub_list_file_base_alias + " " + corpus_year
    pub_list_file_alias      =  year_pub_list_file_alias + ".xlsx"
    corpus_year_path         = bibliometer_path / Path(corpus_year)                  
    pub_list_path            = corpus_year_path / Path(pub_list_path_alias)
    pub_list_file_path       = pub_list_path / Path(pub_list_file_alias)

    # Setting useful column names
    col_final_list   = set_final_col_names(institute, org_tup)
    pub_id_col_alias = col_final_list[0]
    doc_type_alias   = col_final_list[7]   

    full_pub_list_df = pd.read_excel(pub_list_file_path)
    pub_nb = len(full_pub_list_df)    
    key_pub_nb = 0
    for key, doctype_list in pg.DOCTYPE_TO_SAVE_DICT.items():
        doctype_list = [x.upper() for x in  doctype_list]
        key_dg = pd.DataFrame(columns = full_pub_list_df.columns)
        
        for doc_type, dg in full_pub_list_df.groupby(doc_type_alias):
            if doc_type.upper() in doctype_list: key_dg = pd.concat([key_dg, dg])  

        key_pub_nb += len(key_dg)
        
        key_dg_file_alias = year_pub_list_file_alias + "_" + key + ".xlsx"
        key_dg_path = pub_list_path / Path(key_dg_file_alias)
        
        key_dg.sort_values(by=[pub_id_col_alias], inplace = True)  
        wb, _ = mise_en_page(institute, org_tup, key_dg)
        wb.save(key_dg_path)
    
    split_ratio = 100
    if pub_nb != 0:    
        split_ratio = round(key_pub_nb / pub_nb*100)
    return split_ratio


def consolidate_pub_list(institute, org_tup, bibliometer_path, in_path, out_path, 
                         out_file_path, in_file_base, corpus_year):  
    '''    
    Args : 
        in_path
        out_file_path
        in_file_base
        
    Returns :
        un fichier excel
    '''
    
    # Standard imports
    from pathlib import Path
    import os
    
    # 3rd party imports
    import pandas as pd
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_InstituteGlobals as ig 
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    from BiblioMeter_FUNCTS.BM_RenameCols import set_final_col_names
    from BiblioMeter_FUNCTS.BM_UsePubAttributes import save_otps 
    
    # internal functions
    def _set_df_OTP_dpt(dpt_label):        
        OTP_file_name_dpt_ok = in_file_base + '_' + dpt_label + '_ok.xlsx'
        
        dpt_path = in_path / Path(OTP_file_name_dpt_ok)       
        if not os.path.exists(dpt_path):
            OTP_file_name_dpt = in_file_base + '_' + dpt_label + '.xlsx'
            dpt_path = in_path / Path(OTP_file_name_dpt)
        dpt_df = pd.read_excel(dpt_path)
        return dpt_df
    
    # Setting useful column names
    col_final_list = set_final_col_names(institute, org_tup)
    
    # Setting useful aliases
    missing_if_filename_base_alias   = pg.ARCHI_IF["missing_if_base"]
    missing_issn_filename_base_alias = pg.ARCHI_IF["missing_issn_base"]
    pub_id_alias                     = col_final_list[0] 
    OTP_alias                        = col_final_list[16]   # Choix de l'OTP
       
    # Setting useful paths
    missing_if_path   = out_path / Path(corpus_year + missing_if_filename_base_alias)
    missing_issn_path = out_path / Path(corpus_year + missing_issn_filename_base_alias)
        
    # Setting institute parameters
    dpt_label_dict = org_tup[1]   
    
    ### Charger les df et les concatener 
    dpt_label_list = list(dpt_label_dict.keys())
    OTP_df_init_status = True
    for dpt_label in dpt_label_list:
        dpt_df =  _set_df_OTP_dpt(dpt_label)
        if OTP_df_init_status:
            OTP_df = dpt_df.copy()            
        else:
            OTP_df = pd.concat([OTP_df, dpt_df])  
        OTP_df_init_status = False

    # Deduplicating rows on Pub_id
    OTP_df.drop_duplicates(subset = [pub_id_alias], inplace = True)
    
    # Selecting useful columns using col_final_list
    consolidate_pub_list_df = OTP_df[col_final_list].copy()
    
    # Saving df to EXCEL file
    consolidate_pub_list_df.to_excel(out_file_path, index = False)
    
    # Saving set OTPs
    otp_message = save_otps(institute, org_tup, bibliometer_path, corpus_year)
    
    # Setting pub ID as index for unique identification of rows
    consolidate_pub_list_df.set_index(pub_id_alias, inplace = True)
    
    # Droping invalide publications by pub Id as index
    consolidate_pub_list_df.drop(consolidate_pub_list_df[consolidate_pub_list_df[OTP_alias] == ig.INVALIDE].index, 
                                 inplace = True)
    
    # Resetting pub ID as a standard column
    consolidate_pub_list_df.reset_index(inplace = True)
    
    # Re_saving df to EXCEL file
    consolidate_pub_list_df.to_excel(out_file_path, index = False)

    # Adding Impact Factors and saving new consolidate_pub_list_df 
    # this also for saving results files to complete IFs database
    _, if_database_complete = add_if(institute, 
                                     org_tup,
                                     bibliometer_path, 
                                     out_file_path, 
                                     out_file_path, 
                                     missing_if_path,
                                     missing_issn_path, 
                                     corpus_year)
    
    # Splitting saved file by documents types (ARTICLES, BOOKS and PROCEEDINGS)
    split_ratio = split_pub_list_by_doc_type(institute, org_tup, bibliometer_path, corpus_year)
    
    end_message  = f"\n" + otp_message
    end_message += f"\nOTPs identification integrated in file : \n  '{out_file_path}'"
    end_message += f"\n\nPublications list for year {corpus_year} has been {split_ratio} % splitted "
    end_message += f"in 2 files by group of document types"
    
    return end_message, split_ratio, if_database_complete 
                    
    
def concatenate_pub_lists(institute, org_tup, bibliometer_path, years_list):
    
    """
    """
    # Standard library imports
    import os
    from pathlib import Path
    from datetime import datetime
    
    # 3rd library imports
    import pandas as pd    
    
    # Local imports
    import BiblioMeter_FUNCTS.BM_PubGlobals as pg
    
    # Setting useful aliases
    pub_list_path_alias      = pg.ARCHI_YEAR["pub list folder"]
    pub_list_file_base_alias = pg.ARCHI_YEAR["pub list file name base"]
    bdd_multi_annuelle_folder_alias = pg.ARCHI_BDD_MULTI_ANNUELLE["root"]
    bdd_multi_annuelle_file_alias   = pg.ARCHI_BDD_MULTI_ANNUELLE["concat file name base"]
    
    # Building the concatenated dataframe of available publications lists
    df_concat = pd.DataFrame()    
    available_liste_conso = ""    
    for i in range(len(years_list)):
        try:
            year = years_list[i]
            pub_list_file_name = f"{pub_list_file_base_alias} {year}.xlsx"
            pub_list_path = bibliometer_path / Path(year) / Path(pub_list_path_alias) / Path(pub_list_file_name)
            df_inter  = pd.read_excel(pub_list_path)
            df_concat = pd.concat([df_concat, df_inter])        
            available_liste_conso += f" {year}"
        
        except FileNotFoundError:
            pass
    
    # Formatting and saving the concatenated dataframe in an EXCEL file
    date = str(datetime.now())[:16].replace(':', 'h')
    out_file = f"{date} {bdd_multi_annuelle_file_alias} {os.getlogin()}_{available_liste_conso}.xlsx"
    out_path = bibliometer_path / Path(bdd_multi_annuelle_folder_alias)
    out_file_path = out_path / Path(out_file)
    
    wb, _ = mise_en_page(institute, org_tup, df_concat)
    wb.save(out_file_path)
    
    end_message  = f"Concatenation of consolidated pub lists saved in folder: \n  '{out_path}'"
    end_message += f"\n\n under filename: \n  '{out_file}'."
    return end_message




