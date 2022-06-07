__all__ = ['create_PageTwo']

def build_year_month_dpt(current_year, sheet_names_all, bibliometer_path, save_case=False):

    '''The excel workbook, with full pass `file`, contains a worksheet per month labelled mmyyyy 
    where mm stands for the month (01,02,...,12) and yyyy stands for the year (2019,2020,...). 
    All the worksheets must at least contain the two columns entitled: `Matricule` and `Dpt/DOB (lib court)`. 
    The function `build_month_dpt` builds out, using at most 12 worksheets, a new dataframe with
    an additional column entitled: 'Matricule' and 'Dpt_month'.
    The columns 'Dpt_month' contains the lists of at most 12 tuples
    [(current_year,month_1,dept_1),...,(current_year,month_n,dept_n)].
    If a matricule is not present in a spreadsheat the tuple is replace by (year,month_1,None).

    Uses the following globals:
    PATH_TO_EFFECTIFS
    COL_NAMES_RH
    COL_NAMES_BM

    Args:
       current_year (string): string formatted as yyyy to build the df on a period of 12 months

    Returns:
       (dataframe)
    '''

    #Standard Library imports
    import os
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_BM
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_RH

    def _recast_to_tuple(df,col):
        df[col] = df[col].apply(lambda x: x if isinstance(x, list) else [x])

    def _tuples_to_lists(df,cols):
        col_in, col_out = cols[0], cols[1]
        df[col_out] = df[col_in].apply(lambda x: [list(x) for x in list(zip(*x))])

    def _get_firstname_initiales(row):
        row = row.replace('-',' ') 
        row_list = row.split(' ')
        initiale_list = [x[0] for x in row_list]
        initiales = ''.join(initiale_list)        
        return initiales

    def _manage_duplicate_matricule(df,col):
        # Initializing dataframe 'df_effectif' as a copy of the dataframe 'df'
        df_effectif = df.copy()
        dup_matricule_idx = (df[df[col].apply(lambda x:len(x)>1)]).index
        for index in dup_matricule_idx:
            save = df.iloc[index]
            save_copy = save.copy()
            sub_names = save[col].copy()    
            for s_name_idx,sub_name in enumerate(sub_names): 
                # For all different sub names for the same matricule
                # Replacing the sub_names list with the current sub name from sub_names list
                save_copy[col] = [sub_name]
                if not s_name_idx: 
                    # Updating the dataframe 'df_effectif' at the index 'index' for the first sub-name
                    df_effectif.iloc[index] = save_copy
                else:            
                    # Adding a copy of info at the end of the dataframe 'df_effectif' with the other sub names
                    df_effectif = df_effectif.append(save_copy.to_frame().T)                    
        return df_effectif

    # Building the list 'sheet_names' of the worksheet's name pertening to the year 'current_year'.
    sheet_names = [x for x in sheet_names_all if str(current_year) in x]

    # Reading the selected sheet_names from the excel file as a dict {sheet-name: sheet-content dataframe}
    df_dict = pd.read_excel(Path(bibliometer_path) / Path(STOCKAGE_ARBORESCENCE['effectif'][0]) / Path(STOCKAGE_ARBORESCENCE['effectif'][2]),
                            sheet_name = sheet_names)

    # Building a dataframe 'df_rh_year' as a concatenation of sheets from the 'sheet_names' list
    # by sweeping on each sheet 'sheet_name' of dict 'df_dict'
    # and getting the dataframe 'df_rh_month' with the content of the sheet 'sheet_name'
    for sheet_idx,(sheet_name,df_rh_month) in enumerate(df_dict.items()):

        # Getting the month and year for the sheet 'sheet_name'
        month = sheet_name[0:2]
        year = sheet_name[2:] # Should be always the same as 'current_year'

        # For the dataframe 'df_rh_month' that contents the sheet 'sheet_name' content
        # replacing each cell of column 'COL_NAMES_RH['Dpt']' that specifies the employee departement dpt
        # by a tuple (month,year,dpt)
        col = COL_NAMES_RH['Dpt']
        df_rh_month[col] = df_rh_month[col].apply(lambda x:(month,year,x))

        # For the dataframe 'df_rh_month' that contents the sheet 'sheet_name' content
        # replacing each cell of column 'COL_NAMES_RH['Service']' that specifies the employee service serv
        # by a tuple (month,year,serv) 
        col = COL_NAMES_RH['Service']
        df_rh_month[col] = df_rh_month[col].apply(lambda x:(month,year,x))

        # Getting the columns names of 'df_rh_month' 
        # To Do: selecting the useful columns to keep
        rh_columns = list(df_rh_month.columns)
        rh_columns_to_keep = rh_columns

        # Building the dataframe 'df_rh_year'
        if sheet_idx == 0:    
            # Initializing 'df_rh_year' with first sheet content
            df_rh_year = df_rh_month[rh_columns_to_keep].copy()
        else:
            # Appending the next sheet content to 'df_rh_year'
            df_rh_year = df_rh_year.append(df_rh_month[rh_columns_to_keep].copy())

    #### Saving the dataframe 'df_rh_year' as an EXCEL file for checking
    ###if ###:df_rh_year.to_excel(PATH_OF_CHECKS / Path('df_rh_year.xlsx'))    

    # Building a dict 'rh_dict' from dataframe 'df_rh_year' 
    # with information keyed by employee matricule COL_NAMES_RH['Matricule']
    # and its structure is {matricule : dataframe which content is df_rh_year for COL_NAMES_RH['Matricule'] equal to matricule}
    rh_dict = {}
    for matricule, df_matricule in df_rh_year.groupby([COL_NAMES_RH['Matricule']]):
        # Setting unique occcurence of items in a given column of the dataframe df_matricule
        rh_dict[matricule] = [df_matricule[item].unique() for item in df_matricule.columns]

        # Recasting array into list if len(array)>1 otherwise into string  
        rh_dict[matricule] = [x[0] if len(x)==1 else list(x) for x in rh_dict[matricule]]     

    # Recasting the dict 'rh_dict' into a dataframe 'df_rh_year_singlemat' with ad hoc columns names and index
    # in this dataframe there is only one row per matricule
    df_rh_year_singlemat = pd.DataFrame.from_dict(rh_dict)
    df_rh_year_singlemat = df_rh_year_singlemat.T
    df_rh_year_singlemat.reset_index(drop=True, inplace=True)
    df_rh_year_singlemat.columns = rh_columns_to_keep

    # Recasting into list of items whatever the number of items for selected columns
    # To Do: Doing the recasting for all columns
    col_list = [COL_NAMES_RH['Dpt'],
                COL_NAMES_RH['Service'],
                COL_NAMES_RH['Nom'],
                COL_NAMES_RH['Prénom']]
    for col in col_list:
        _recast_to_tuple(df_rh_year_singlemat,col)

    #### Saving the dataframe 'df_rh_year_singlemat' as an EXCEL file for checking
    ###if save_case:df_rh_year_singlemat.to_excel(PATH_OF_CHECKS / Path('df_rh_year_singlemat.xlsx')) 

    # Transforming list of tuples into lists of lists ex [(x,c,d),((e,f,g))] -->[[x,e],[c,f],[d,g]] for 'Dpt' and 'Service' columns
    # and putting them in 'Dpts' and 'Servs' new columns
    cols_tup_list = [(COL_NAMES_RH['Dpt'],COL_NAMES_BM['Dpts']), 
                     (COL_NAMES_RH['Service'],COL_NAMES_BM['Servs'])]
    for cols_tup in cols_tup_list:
        _tuples_to_lists(df_rh_year_singlemat,cols_tup) 

    # Spliting list of n lists into n different columns for 'Dpts' and 'Servs' columns
    for col in [COL_NAMES_BM['Dpts'], COL_NAMES_BM['Servs']]:
        df_rh_year_singlemat[['months', 'years', col]] = pd.DataFrame(df_rh_year_singlemat[col].tolist())

    #### Saving the dataframe 'df_rh_year_singlemat' as an EXCEL file for checking
    ###if save_case:df_rh_year_singlemat.to_excel(PATH_OF_CHECKS / Path('df_rh_year_singlemat_addedcols.xlsx'))

    # Dealing with same matricule for different lastnames 
    col = COL_NAMES_RH['Nom']
    df_effectif = _manage_duplicate_matricule(df_rh_year_singlemat,col)

    # Dealing with same matricule for different firstnames
    col = COL_NAMES_RH['Prénom']
    df_effectif = _manage_duplicate_matricule(df_rh_year_singlemat,col)

    # Creating a column with first name initials as a list
    # ex PIERRE -->P, JEAN-PIERRE --> JP , JEAN-PIERRE MARIE --> JPM 
    col_in, col_out = COL_NAMES_RH['Prénom'], COL_NAMES_BM['First_name']
    df_effectif[col_out] = df_effectif.apply(lambda row : 
                                             [_get_firstname_initiales(x) for x in row[col_in]],
                                             axis=1)

    # Recasting single-element list into string for specific columns
    cols_list = [COL_NAMES_RH['Nom'],COL_NAMES_RH['Prénom'],COL_NAMES_BM['First_name']]
    for col in cols_list:
        df_effectif[col] = df_effectif[col].apply(lambda x : x[0])

    # Creating the column ['Full_name'] by combining COL_NAMES_RH['Nom'] and COL_NAMES_BM['First_name']
    new_col = COL_NAMES_RH['Full_name']
    df_effectif[new_col] = df_effectif[COL_NAMES_RH['Nom']] + ' ' + df_effectif[COL_NAMES_BM['First_name']]

    # Keeping only the last element of the first tuple (the first occurrence in the year) for specific columns
    cols_list = [COL_NAMES_RH['Dpt'],COL_NAMES_RH['Service']]
    for col in cols_list:
        df_effectif[col] = df_effectif[col].apply(lambda x: x[0][-1])    

    df_effectif = df_effectif.reset_index()

    #### Saving the dataframe 'df_effectif' as an EXCEL file for checking
    ###if ###:df_effectif.to_excel(PATH_OF_CHECKS / Path('df_effectif_test.xlsx'))  

    return df_effectif

def build_fichier_rh_all_years(years, file_name, sheet_names_all, bibliometer_path):

    '''Function `build_fichier_rh_all_years` builds employees dataframes for a list of years `years`
       and saves the results in 'file_name' file

    Uses the following globals:
    PATH_OF_RESULTS

    '''

    #Standard Library imports
    import os
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    list_fichier_rh = [(year, build_year_month_dpt(year, sheet_names_all, bibliometer_path)) for year in years]

    path = Path(bibliometer_path) / Path('Results') / Path(file_name)
    with pd.ExcelWriter(path, engine='xlsxwriter') as writer:
        for sheet_name,x in list_fichier_rh:
            x.to_excel(writer, sheet_name=sheet_name)

def build_df_submit(df_eff, df_pub, bibliometer_path, test_case='No test'):

    """
    Description à venir
    """

    #Standard Library imports
    import os
    from pathlib import Path

    # 3rd party import
    import numpy as np
    import pandas as pd

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_BM
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_RH

    # Création chemin vers résultats pour effectuer des checks étapes
    PATH_OF_CHECKS = Path(bibliometer_path) / Path('Results')

    def _orphan_reduction(orphan_lastname,eff_lastname):
        # A bug with "if ' TRAN ' in ' TUAN TRAN ':"
        orphan_lastname = ' ' + orphan_lastname + ' '
        lastname_match_list = []
        for eff_name in eff_lastname:
            if (orphan_lastname in eff_name) or (eff_name in orphan_lastname):
                lastname_match_list.append(eff_name.strip())
        return lastname_match_list

    def _test_full_match():
        if len(df_eff_pub_match)!=0:
            print('\nMatch found for author lastname:',pub_lastname)
            print(' Nb of matches:',len(df_eff_pub_match))
            print(' Employee matricule:',df_eff_pub_match[COL_NAMES_RH['Matricule']].to_list()[0])
            print(' Employee lastname:',df_eff_pub_match[COL_NAMES_RH['Nom']].to_list()[0])
        else:
            print('\nNo match for author lastname:',pub_lastname)
            print('  Nb first matches:',len(df_eff_pub_match))

    def _test_similarity():
        print('\nSimilarities by orphan reduction for author lastname:',pub_lastname)
        print('  Lastname flag match:', flag_lastname_match)
        print('  Nb similarities by orphan reduction:',len(lastname_match_list))
        print('  List of lastnames with similarities:', lastname_match_list)
        print('  Employee matricules:',df_eff_pub_match[COL_NAMES_RH['Matricule']].to_list())
        print('  Employee lastnames:',df_eff_pub_match[COL_NAMES_RH['Nom']].to_list())
        print('  Employee firstnames:',df_eff_pub_match[COL_NAMES_RH['Prénom']].to_list())
        print('  Employee fullnames:',df_eff_pub_match[COL_NAMES_RH['Full_name']].to_list())    

    def _test_no_similarity():
        print('\nNo similarity by orphan reduction for author lastname:',pub_lastname)
        print('  Lastname flag match:', flag_lastname_match)
        print('  Nb similarities by orphan reduction:',len(lastname_match_list))
        print('  Orphan full author name:',df_pub_row[COL_NAMES['authors'][2]])
        print('  Orphan author lastname:',df_pub_row[COL_NAMES_BM['Last_name']])
        print('  Orphan author firstname initiales:',df_pub_row[COL_NAMES_BM['First_name']])

    def _test_match_of_firstname_initiales():
        print('\nInitiales for author lastname:',pub_lastname)
        print('  Author fullname:', df_pub_row[COL_NAMES['authors'][2]]) 
        print('  Author firstname initiales:',pub_firstname)
        print('\nInitiales of matching employees for author lastname:',pub_lastname)
        print('  Employees firstname initiales list:',eff_firstnames)                              
        print('\nChecking initiales matching for author lastname:',pub_lastname)
        print('  Nb of matching initiales:', len(list_idx))
        print('  Index list of matching initiales:',list_idx)
        print('  Employees lastnames list:',eff_lastnames_spec)

    def _save_spec_dfs():                       
        df_temp.to_excel(PATH_OF_CHECKS / Path('df_temp_' + test_name + '.xlsx'))
        df_eff_pub_match.to_excel(PATH_OF_CHECKS / Path('df_eff_pub_match_' + test_name + '.xlsx'))     

    # Initializing a Data frame that will contains all matches 
    # between 'df_pub' author-name and 'df_eff' emmployee-name
    df_submit = pd.DataFrame() 

    # Initializing a Data frame that will contains all 'df_pub' author-names 
    # which do not match with any of the 'df_eff' emmployee-names
    df_orphan = pd.DataFrame()

    # Building the set of lastnames (without duplicates) of the dataframe 'df_eff' 
    eff_lastnames = set(df_eff[COL_NAMES_RH['Nom']].to_list())
    eff_lastnames = [' ' + x + ' ' for x in eff_lastnames]

    # Setting the useful info for testing the function if verbose = True
    # Setting a dict keyyed by type of test with values for test states and 
    # test name from column [COL_NAMES_BM['Last_name']] of the dataframe 'df_pub'
    # for testing this function for year 2021
    test_dict = {'Full match'            : [True, True, True,True,True,'SIMONATO'],
                 'Lower value similarity': [False,True, True,True,True,'SILVA' ],
                 'Upper value similarity': [False,True, True,True,True,'TUAN TRAN'],
                 'No similarity'         : [False,False,True,True,True,'LUIS GABRIEL'],
                 'No test'               : [False,False,False,False,False,'None']
                 }

    test_nb = len(test_dict[test_case])-1
    test_name = test_dict[test_case][test_nb]
    test_states = test_dict[test_case][0:test_nb]    

    # Building df_submit and df_orphan dataframes
    for _,df_pub_row in df_pub.iterrows(): 

        # Building a dataframe 'df_match_eff_publi' with rows of 'df_eff'
        # where name in column COL_NAMES_BM['Last_name'] of the dataframe 'df_pub' 
        # matches with name in column COL_NAMES_RH['Nom'] of the dataframe 'df_eff'

        # Initializing 'df_eff_pub_match' as dataframe
        df_eff_pub_match = pd.DataFrame()

        # Initializing the flag 'flag_lastname_match' as True by default
        flag_lastname_match = True

        # Getting the lastname from df_pub_row row of the dataframe df_pub
        pub_lastname = df_pub_row[COL_NAMES_BM['Last_name']]

        # Building the dataframe 'df_eff_pub_match' with rows of dataframe df_eff 
        # where item at COL_NAMES_RH['Nom'] matches author lastname 'pub_lastname'
        df_eff_pub_match = df_eff[df_eff[COL_NAMES_RH['Nom']] == pub_lastname].copy()

        # Adding column COL_NAMES_BM['Full_name'] + '_eff' by combination of 
        # df_eff_pub_match[COL_NAMES_RH['Nom']] and df_eff_pub_match[COL_NAMES_BM['First_name']]
        #new_col = COL_NAMES_BM['Full_name']
        #df_eff_pub_match[new_col] =  df_eff_pub_match[COL_NAMES_RH['Nom']] + ' ' + df_eff_pub_match[COL_NAMES_BM['First_name']] #------------------------------------------

        # Test of lastname full match
        if pub_lastname == test_name and test_states[0]: _test_full_match()          

        if len(df_eff_pub_match) == 0: # No match found
            flag_lastname_match = False
            lastname_match_list = _orphan_reduction(pub_lastname,eff_lastnames) # check for a similarity

            if lastname_match_list: 
                # Concatenating in the dataframe 'df_eff_pub_match', the rows of the dataframe 'df_eff'
                # corresponding to each of the found similarities by orphan reduction
                col = COL_NAMES_RH['Nom']
                frames = []
                for lastname_match in lastname_match_list:
                    df_temp = df_eff[df_eff[COL_NAMES_RH['Nom']] == lastname_match].copy()
                    # Replacing the employee last name by the publication last name
                    # for df_pub_rh_join building
                    df_temp[COL_NAMES_RH['Full_name']] = pub_lastname + ' ' + df_temp[COL_NAMES_BM['First_name']] #------------------------------------------
                    frames.append(df_temp )

                df_eff_pub_match = pd.concat(frames, ignore_index=True)
                flag_lastname_match = True

                # Test of lastnames similarity found by '_orphan_reduction' function
                if pub_lastname == test_name and test_states[1]: _test_similarity()

            else: 
                # Appending to dataframe df_orphan the row 'df_pub_row'  as effective orphan after orphan reduction
                df_orphan = df_orphan.append(df_pub_row)
                flag_lastname_match = False

                # Test of lastnames no-similarity by '_orphan_reduction' function
                if pub_lastname == test_name and test_states[2]: _test_no_similarity             

        # Checking match for a given lastname between the publication first-name and the employee first-name
        if flag_lastname_match:

            # Finding the author name initiales for the current publication
            pub_firstname = df_pub_row[COL_NAMES_BM['First_name']]

            # List of firstnames initiales of a given name in the rh effectif
            eff_firstnames = df_eff_pub_match[COL_NAMES_BM['First_name']].to_list()
            eff_lastnames_spec = df_eff_pub_match[COL_NAMES_RH['Nom']].to_list()

            #if pub_lastname == 'MARTIN': 
                #print(eff_firstnames)

            # Building the list of index of firsnames initiales 
            list_idx = []
            for idx,eff_firstname in enumerate(eff_firstnames):

                #if pub_lastname == 'MARTIN': 
                    #print()
                    #print('pub_lastname',pub_lastname, 'pub_firstname',pub_firstname)
                    #print('eff_firstnames',eff_firstnames,'eff_firstname',eff_firstname)

                if (pub_firstname == eff_firstname):
                    list_idx.append(idx)

                elif (pub_firstname[0] == eff_firstname[0]):
                    # Replacing the employee first name initials by the publication first name initials
                    # for df_pub_rh_join building
                    df_eff_pub_match[COL_NAMES_RH['Full_name']].iloc[idx] = pub_lastname + ' ' + pub_firstname
                    list_idx.append(idx)

                #if pub_lastname == 'MARTIN': 
                    #print(list_idx)
                    #print()

            # Test of match of firstname initiales for lastname match or similarity
            if pub_lastname == test_name and test_states[3]: _test_match_of_firstname_initiales()

            if list_idx: 
                # Building a dataframe df_temp with the row 'df_pub_row'related to a given publication 
                # and adding the item value 'HOMONYM' at column COL_NAMES_BM['Homonym'] 
                # when several matches on firstname initiales are found
                df_temp = df_pub_row.to_frame().T            
                df_temp[COL_NAMES_BM['Homonym']] = 'HOMONYM' if len(list_idx)>1 else ''

                # Saving specific dataframes 'df_temp' and 'df_eff_pub_match' for function testing
                if pub_lastname == test_name and test_states[4]: _save_spec_dfs()                       

                # Merging the dataframe 'df_eff_pub_match' to the dataframe 'df_temp' 
                # by matching column '[COL_NAMES_BM['Last_name']]' of the dataframe 'df_temp'
                # to the column '[COL_NAMES_RH['Nom']]' of the dataframe 'df_eff_pub_match'
                df_pub_rh_join = pd.merge(df_temp,
                                          df_eff_pub_match, 
                                          how = 'left',
                                          left_on = [COL_NAMES_BM['Full_name']],
                                          right_on = [COL_NAMES_RH['Full_name']])

                # Appending to the dataframe 'df_submit' the dataframe 'df_pub_rh_join'
                # which is specific to a given publication
                df_submit = df_submit.append(df_pub_rh_join, ignore_index = True)
            else:
                # Appending to the dataframe df_orphan the row 'df_pub_row' as effective orphan 
                # after complementary checking of match via firsname initiales
                df_orphan = df_orphan.append(df_pub_row)

    return df_submit, df_orphan

def _single_year_search(year, bibliometer_path):

    """
    Description à venir
    """

    #Standard Library imports
    import os
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_BM
    from BiblioMeter_GUI.Globals_GUI import SUBMIT_FILE_NAME
    from BiblioMeter_GUI.Globals_GUI import ORPHAN_FILE_NAME #TO DO : faire une seule variable pour ça 

    ###################################
    # Building the articles dataframe #
    ###################################

    df_pub = _build_pubs_authors_Liten(year, bibliometer_path)

    #################################################
    # Setting the file names for saving the results #
    #################################################

    year_submit_file_name = year + '_' + SUBMIT_FILE_NAME
    year_orphan_file_name = year + '_' + ORPHAN_FILE_NAME

    ##################################################################
    # Building the list of available years in the employees database #
    ##################################################################

    df = pd.ExcelFile(Path(bibliometer_path) / Path(STOCKAGE_ARBORESCENCE['effectif'][0]) / Path(STOCKAGE_ARBORESCENCE['effectif'][2]))
    sheet_names_all = df.sheet_names
    sheet_names_all = sheet_names_all[::-1] # inverses the list of sheet names
    years_db = set([sheet_names_all[index][2:] for index in range(len(sheet_names_all))])

    if year not in years_db :
        print('Year ',year,'is not available in the employees database which is limited to',years_db)
        print('Please define a new testing year.')
    else:
        #######################################################
        # Building the 'df_eff' dataframe for year `year` #
        #######################################################

        save_case=False
        df_eff = build_year_month_dpt(year, sheet_names_all, bibliometer_path, save_case=save_case)

        ##############################################################################################
        # Building the `df_submit` and `df_orphan` dataframes using `df_eff` dataframe of year `year` #
        ##############################################################################################

        test_list = ['Full match',            
                     'Lower value similarity',
                     'Upper value similarity',
                     'No similarity',
                     'No test'
                    ]
        test_case = 'Upper value similarity'
        df_submit, df_orphan =  build_df_submit(df_eff, df_pub, bibliometer_path, test_case = test_case)

        #############################################################################
        # Save results in 'year_submit_file_name' and 'year_orphan_file_name' files #
        #############################################################################

        df_submit.to_excel(Path(bibliometer_path) / Path(year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(year_submit_file_name))
        df_orphan.to_excel(Path(bibliometer_path) / Path(year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(year_orphan_file_name))  
        print('The dataframe to be submitted is built and saved')

def _recursive_year_search(corpus_year, bibliometer_path):

    """
    Description à venir
    """

    # Standard library imports
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_BM
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_ORPHAN
    from BiblioMeter_GUI.Globals_GUI import SUBMIT_FILE_NAME
    from BiblioMeter_GUI.Globals_GUI import ORPHAN_FILE_NAME

    ###################################
    # Building the articles dataframe #
    ###################################

    df_pub = _build_pubs_authors_Liten(corpus_year, bibliometer_path)

    ##################################################################
    # Building the list of available years in the employees database #
    ##################################################################

    df = pd.ExcelFile(Path(bibliometer_path) / Path(STOCKAGE_ARBORESCENCE['effectif'][0]) / Path(STOCKAGE_ARBORESCENCE['effectif'][2]))
    sheet_names_all = df.sheet_names
    sheet_names_all = sheet_names_all[::-1] # inverses the list of sheet names
    years_db = set([sheet_names_all[index][2:] for index in range(len(sheet_names_all))])

    ##########################################################
    # Building the search time extension of Liten co-authors #
    ##########################################################

    today_year = 2022
    time_line_history = 5
    years = [str(i) for i in range(today_year - time_line_history,today_year)]
    years = years[::-1]

    ##################################################################################
    # building an employees dataframe for the list of years and saving as EXCEL file #
    ##################################################################################

    #build_fichier_rh_all_years(years,EFFECTIFS_FILE,sheet_names_all)

    #####################################################################################################
    # Building recursively the `df_submit` and `df_orphan` dataframes using `df_eff` files of years #
    #####################################################################################################

    # Initializing the dataframes to be built
    df_submit = pd.DataFrame() # Data frame containing all match between publi author name name and rh name name
    df_orphan = pd.DataFrame() # No match found between article LITEN author and rh names

    # Setting the test case
    test_list = ['Full match',            
                 'Lower value similarity',
                 'Upper value similarity',
                 'No similarity',
                 'No test'
                ]

    test_case = 'Upper value similarity'

    for step, year in enumerate(years): 

        # Read the sheet `year` of `EFFECTIFS_FILE` file
        effectifs_path = Path(bibliometer_path) / Path(STOCKAGE_ARBORESCENCE['effectif'][0]) / Path(STOCKAGE_ARBORESCENCE['effectif'][1]) # Récupère ALL_Effectifs.xlsx

        df_eff = pd.read_excel(effectifs_path, sheet_name = year)
        df_eff.drop(['Unnamed: 0'], axis=1, inplace = True)

        if step == 0:
            df = pd.ExcelFile(Path(bibliometer_path) / Path(STOCKAGE_ARBORESCENCE['effectif'][0]) / Path(STOCKAGE_ARBORESCENCE['effectif'][2]))
            sheet_names_all = df.sheet_names
            sheet_names_all = sheet_names_all[::-1] # inverses the list of sheet names
            df_eff = build_year_month_dpt(year,sheet_names_all, bibliometer_path)

            # Building the initial dataframes
            df_submit, df_orphan =  build_df_submit(df_eff, df_pub, bibliometer_path, test_case=test_case)

        else:
            # Updating the dataframes 
            df_submit_add, df_orphan =  build_df_submit(df_eff, df_orphan, bibliometer_path, test_case)
            # Updating df_submit 
            df_submit = df_submit.append(df_submit_add)        

        year_submit_file_name = year + '_' + SUBMIT_FILE_NAME
        year_orphan_file_name = year + '_' + ORPHAN_FILE_NAME

        df_submit.to_excel(Path(bibliometer_path) / Path(corpus_year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(year_submit_file_name))
        df_orphan.to_excel(Path(bibliometer_path) / Path(corpus_year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(year_orphan_file_name)) 

    #_=[bmu.you_got_OTPed(df_submit,i) for i in range(len(df_submit))]

    #####################################################################
    # Saving results in `SUBMIT_FILE_NAME` and `ORPHAN_FILE_NAME` files #
    #####################################################################

    df_submit.to_excel(Path(bibliometer_path) / Path(corpus_year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(SUBMIT_FILE_NAME)) 
    df_orphan = df_orphan.reindex(columns = COL_NAMES_ORPHAN)
    df_orphan.to_excel(Path(bibliometer_path) / Path(corpus_year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(ORPHAN_FILE_NAME))         
    print('Results saved')

def _mise_en_forme(year, bibliometer_path):

    # Standard library imports
    from pathlib import Path
    import math

    # 3rd party import
    import pandas as pd
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.styles.colors import Color

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioMeter_GUI.Globals_GUI import SUBMIT_FILE_NAME
    from BiblioMeter_GUI.Globals_GUI import ORPHAN_FILE_NAME

    ###########################
    # Getting the submit file #
    ###########################

    df_submit = pd.read_excel(Path(bibliometer_path) / Path(year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path(SUBMIT_FILE_NAME))

    # Remettre les Pub_ID dans l'ordre
    df_submit.sort_values(by=['Pub_id', 'Dpt/DOB (lib court)'], ascending=False, inplace = True)
    df_submit.sort_values(by=['Pub_id', 'Idx_author'], ascending=True, inplace = True)
    #df_submit.sort_values(by=['Pub_id'], inplace = True)

    ###############################
    # Réajustement de List_of_OTP #
    ###############################

    _=[bmu.you_got_OTPed(df_submit,i) for i in range(len(df_submit))]

    ############################################
    # Getting rid of the columns we don't want #
    ############################################

    columns_to_keep = ['Pub_id', 
                       'Idx_author', 
                       'Authors',  
                       'DOI', 
                       'ISSN', 
                       'LITEN_France',  
                       'Document_type',
                       'Matricule', 
                       'Nom', 
                       'Prénom', 
                       'Dpt/DOB (lib court)', 
                       'Service (lib court)', 
                       'Laboratoire (lib court)', 
                       'Laboratoire (lib long)',
                       'List_of_OTP',
                       'HOMONYM']

    df_submit = df_submit[columns_to_keep].copy()

    # La colonne 0/1 pour afficher ou pas
    list_de_1 = [1] * len(df_submit)
    df_submit['Affichage'] = list_de_1

    columns_to_keep = df_submit.columns.to_list()

    # Liste unique des Pub_id conservé
    list_of_Pub_id = df_submit['Pub_id'].to_list()
    list_of_Pub_id = bmu.get_unique_numbers(list_of_Pub_id)

    #########################
    # Ouverture du workbook #
    #########################

    wb = Workbook()
    ws = wb.active
    ws.title = 'Publi x Effectifs'

    yellow_ft = PatternFill(fgColor = '00FFFF00', fill_type = "solid")
    red_ft = PatternFill(fgColor = '00FF0000', fill_type = "solid")
    blue_ft = PatternFill(fgColor = '0000FFFF', fill_type = "solid")
    bd = Side(style='medium', color="000000")
    active_color = 'red'

    for indice, r in enumerate(dataframe_to_rows(df_submit, index=False, header=True)):
        ws.append(r)
        last_row = ws[ws.max_row]
        if r[columns_to_keep.index('HOMONYM')] == 'HOMONYM' and indice > 0:
            cell = last_row[columns_to_keep.index('Nom')]
            cell.fill = yellow_ft
            cell = last_row[columns_to_keep.index('Prénom')]
            cell.fill = yellow_ft

            column_letter = get_column_letter(columns_to_keep.index('Nom')+1)
            excel_cell_id = column_letter + str(indice+1)

            # Adds comments to the highlighted cells ---------------------> CREER UNE FONCTION
            comment = Comment("Homonymy detected, please chose one of the following yellow rows by choosing 1 or 0 in the 'Affichage' column", "Author")
            comment.width = 300
            comment.height = 60

            ws[excel_cell_id].comment = comment

            cell = last_row[columns_to_keep.index('Affichage')]
            cell.fill = yellow_ft
            column_letter = get_column_letter(columns_to_keep.index('Affichage')+1)
            excel_cell_id = column_letter + str(indice+1)

            # Adds comments to the highlighted cells ---------------------> CREER UNE FONCTION
            comment = Comment("1 to keep, 0 to get rid of", "Author")
            comment.width = 155
            comment.height = 20

            ws[excel_cell_id].comment = comment


        if ws.max_row > 1 and r[0] in list_of_Pub_id:
            if list_of_Pub_id.index(r[0])%2 == 0:
                cell = last_row[0]
                cell.fill = red_ft
                if active_color != 'red':
                    for cell_bis in last_row:
                        cell_bis.border = Border(top=bd)
                        active_color = 'red'
            else:
                cell = last_row[0]
                cell.fill = blue_ft
                if active_color != 'blue':
                    for cell_bis in last_row:
                        cell_bis.border = Border(top=bd)
                        active_color = 'blue'


    for cell in ws['A'] + ws[1]:
        cell.font = Font(bold=True)
        cell.border = Border(left=bd, top=bd, right=bd, bottom=bd)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in range(0,ws.max_row-2):
        #if df_jp['Pub_id'].iloc[r-1] == df_jp['Pub_id'].iloc[r] and df_jp['Dpt/DOB (lib court)'].iloc[r-1] == df_jp['Dpt/DOB (lib court)'].iloc[r]: VERSION AMAL
        if df_submit['Pub_id'].iloc[r-1] == df_submit['Pub_id'].iloc[r]: # VERSION JEAN PIERRE
            'Yo'
        else:        
            validation_list = bmu.liste_de_validation(df_submit,r)

            data_val = DataValidation(type="list",formula1=validation_list, showErrorMessage=False)
            ws.add_data_validation(data_val)

            data_val.add(ws[get_column_letter(columns_to_keep.index('List_of_OTP')+1)+str(r+2)])

    wb.save(Path(bibliometer_path) / Path(year) / Path(STOCKAGE_ARBORESCENCE['general'][0]) / Path('PubliXEffectifs.xlsx'))
    
def _build_pubs_authors_Liten(year, bibliometer_path):

    """ 
    Description : 

    Uses the following globals :
    - PATH_DAT_DEDUPLICATED
    - LIST_INSTITUTIONS

    Args : 

    Returns : 

    """

    #Standard Library imports
    import os
    from pathlib import Path

    # 3rd party import
    import pandas as pd

    # Local library imports
    import BiblioMeter_Utils as bmu

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import COL_NAMES

    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_BM
    from BiblioMeter_GUI.Globals_GUI import COL_NAMES_ORPHAN

    # Création des alias pour simplifier les accès
    corpus_path_alias = FOLDER_NAMES['corpus']
    wos_alias = FOLDER_NAMES['wos']
    scopus_alias = FOLDER_NAMES['scopus']

    scopus_path_alias = Path(corpus_path_alias) / Path(scopus_alias)
    wos_path_alias = Path(corpus_path_alias) / Path(wos_alias)

    parsing_path_alias = FOLDER_NAMES['parsing']
    rawdata_path_alias = FOLDER_NAMES['rawdata']

    concat_path_alias = Path(corpus_path_alias) / FOLDER_NAMES['concat']
    dedupli_path_alias = Path(corpus_path_alias) / FOLDER_NAMES['dedup']

    article_path_alias = DIC_OUTDIR_PARSING['A']
    index_alias = COL_NAMES['address'][0]

    # Building useful paths
    PATH_DAT_DEDUPLICATED = Path(bibliometer_path) / Path(year) / Path(dedupli_path_alias) / Path(parsing_path_alias)

    def _retain_firstname_initiales(row):
        row = row.replace('-',' ')
        initiales = ''.join(row.split(' '))
        return initiales 

    def _split_lastname_firstname(row,digits_min=4):
        names_list = row.split()
        last_names_list = names_list[:-1]
        first_names_list = names_list[-1:]

        for name_idx,name in enumerate(last_names_list):

            if len(name)<digits_min and ('-' in name):
                first_names_list.append(name)
                first_names_list = first_names_list[::-1]
                last_names_list = last_names_list[:name_idx] + last_names_list[name_idx+1:]

        first_name_initiales = _retain_firstname_initiales((' ').join(first_names_list))
        last_name = (' ').join(last_names_list)

        return (last_name, first_name_initiales)

    def _build_filt_authors_inst():

        '''The function `_build_filt_authors_inst` builds the `filt_authors_inst` filter
        to select the authors by their institution.
        Returns a filter, as a Pandas Serie, with:
                - true if any institution in the institution list is equal to the author institution;
                - false elsewhere.
        '''

        filt_authors_inst = df_authorsinst_authors['LITEN_France']==1 # TO DO : faire en sorte qu'on construise en fonction de ce l'on souhaite chercher

        return filt_authors_inst

    # Getting ID of each author with institution by publication ID
    df_authorsinst = pd.read_csv(PATH_DAT_DEDUPLICATED / Path(DIC_OUTDIR_PARSING['I2']), 
                                 sep="\t")

    # Getting ID of each author with author name  
    df_authors = pd.read_csv(PATH_DAT_DEDUPLICATED / Path(DIC_OUTDIR_PARSING['AU']), 
                             sep="\t")

    # Getting ID of each publication with complementary info 
    df_articles = pd.read_csv(PATH_DAT_DEDUPLICATED / Path(DIC_OUTDIR_PARSING['A']), 
                              sep="\t")

    # Combining name of author to author ID with institution by publication ID
    df_authorsinst_authors = pd.merge(df_authorsinst, 
                                      df_authors, 
                                      how = 'inner', 
                                      left_on = [COL_NAMES['authors'][0],COL_NAMES['authors'][1]], 
                                      right_on = [COL_NAMES['authors'][0],COL_NAMES['authors'][1]])

    # Building the LITEN authors filter
    filt_authors_LITEN = _build_filt_authors_inst() 

    # Associating each publication (with its complementary info) with each of its LITEN authors
    # The resulting dataframe contents on a row for each LITEN author with the corresponding publication info 
    merged_df_Liten = pd.merge(df_authorsinst_authors[filt_authors_LITEN], 
                               df_articles, 
                               how = 'left', 
                               left_on = [COL_NAMES['authors'][0]], 
                               right_on = [COL_NAMES['authors'][0]])

    # Transforming to uppercase the LITEN author name which is in column COL_NAMES_BAU['co_author']
    col = COL_NAMES['authors'][2]
    merged_df_Liten[col] = merged_df_Liten[col].str.upper()

    # Spliting the LITEN author name to firstname initiales and lastname
    # and putting them as a tupple in column COL_NAMES_BM['Last_name']
    col_in, col_out = COL_NAMES['authors'][2], COL_NAMES_BM['Full_name'] #COL_NAMES_BM['Last_name'] 
    merged_df_Liten[col_out] = merged_df_Liten.apply(lambda row: 
                                                     _split_lastname_firstname(row[col_in]),
                                                     axis=1)

    # Spliting tuples of column COL_NAMES_BM['Full_name']
    # into the two columns COL_NAMES_BM['Last_name'] and COL_NAMES_BM['First_name']
    col_in = COL_NAMES_BM['Full_name'] #Last_name + firstname initials
    col1_out, col2_out = COL_NAMES_BM['Last_name'], COL_NAMES_BM['First_name']
    merged_df_Liten[[col1_out, col2_out]] = pd.DataFrame(merged_df_Liten[col_in].tolist())

    # Recasting tuples (NAME, INITIALS) into a single string 'NAME INITIALS'
    col_in = COL_NAMES_BM['Full_name'] #Last_name + firstname initials
    merged_df_Liten[col_in] = merged_df_Liten[col_in].apply(lambda x : ' '.join(x))

    return  merged_df_Liten

def create_PageTwo(self, bibliometer_path):
    
    """
    Description : 
    
    Uses the following globals :
    
    Args :
    
    Returns :
    
    """
    
    # Standard library imports
    import os
    from pathlib import Path
    from datetime import datetime

    # 3rd party imports
    import tkinter as tk
    from tkinter import filedialog
    from tkinter import messagebox
    import pandas as pd

    # Local imports
    import BiblioAnalysis_Utils as bau
    from BiblioMeter_GUI.BiblioMeter_AllPagesFunctions import five_last_available_years
    from BiblioMeter_GUI.Globals_GUI import STOCKAGE_ARBORESCENCE

    from BiblioAnalysis_Utils.BiblioSpecificGlobals import DIC_OUTDIR_PARSING
    from BiblioAnalysis_Utils.BiblioSpecificGlobals import FOLDER_NAMES  

    # Useful alias
    bdd_mensuelle_alias = STOCKAGE_ARBORESCENCE['general'][0]
    bdd_annuelle_alias = STOCKAGE_ARBORESCENCE['general'][1]

    # Récupérer les corpus disponibles
    list_annee = five_last_available_years(bibliometer_path)

    variable = tk.StringVar(self)
    variable.set(list_annee[0])

    OptionButton = tk.OptionMenu(self, variable, *list_annee)
    OptionButton.place(anchor = 'center', relx = 0.50, rely = 0.20)

    Button_merged = tk.Button(self, 
                              text = "Création de la liste des auteurs Liten à \n partir des fichiers concaténés et dédoublonnés \n pour l'année sélectionnée", 
                              font = ("Helvetica", 9), 
                              command = lambda: _build_pubs_authors_Liten(variable.get(), bibliometer_path))
    Button_merged.place(anchor = 'center', relx = 0.5, rely = 0.35)

    Button_single_year = tk.Button(self, 
                                   text = 'PubliXEffectifs Single Year', 
                                   font = ("Helvetica", 18), 
                                   command = lambda: _single_year_search(variable.get(), bibliometer_path))
    Button_single_year.place(anchor = 'center', relx = 0.5, rely = 0.50)

    Button_recursive = tk.Button(self, 
                                 text = 'PubliXEffectifs Recursive', 
                                 font = ("Helvetica", 18), 
                                 command = lambda: _recursive_year_search(variable.get(), bibliometer_path))
    Button_recursive.place(anchor = 'center', relx = 0.5, rely = 0.65)

    Button_mise_en_forme = tk.Button(self, 
                                     text = 'Concat 5 dernières années des submits', 
                                     font = ("Helvetica", 18), 
                                     command = lambda: _concat_submit())
    Button_mise_en_forme.place(anchor = 'center', relx = 0.5, rely = 0.80)
    
    def _concat_submit():
        
        df_concat = pd.DataFrame()
    
        for i in range(len(list_annee)):
            path = Path(bibliometer_path) / Path(list_annee[i]) / Path(bdd_mensuelle_alias) / Path(f'submit.xlsx')
            df_inter = pd.read_excel(path)
            df_concat = df_concat.append(df_inter)
            
        date = str(datetime.now())[:16].replace(':', '')
        df_concat.to_excel(Path(bibliometer_path) / Path(bdd_annuelle_alias) / Path(f'{date}_concat_submit_{os.getlogin()}.xlsx'))